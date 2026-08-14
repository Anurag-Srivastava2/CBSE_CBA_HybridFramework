import base64
import html
import json
import os
import re
import shutil
import socket
import tempfile
import time
import uuid
import webbrowser
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
import pytest
from pytest_html import extras
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from utilities.environment_health import (
    is_infrastructure_failure,
    wait_for_environment,
)
from utilities.logger import LogGenerator
from utilities.pdf_report_utils import PDFReportUtils
from utilities.read_config import ReadConfig
from utilities.screenshot_utils import ScreenshotUtils


EXTENT_RESULTS = []
ALLURE_RESULTS = []


def get_reports_dir():
    """Allow concurrent pytest sessions to keep report artifacts isolated."""
    configured = os.getenv("PYTEST_REPORTS_DIR", "").strip()
    return Path(configured).resolve() if configured else Path.cwd() / "reports"


def get_partial_results_dir():
    """Where xdist workers hand their results back to the controller."""
    return get_reports_dir() / "partial_results"


def get_xdist_worker_id(config):
    """Return 'gw0', 'gw1', ... inside an xdist worker; None otherwise."""
    workerinput = getattr(config, "workerinput", None)
    return workerinput["workerid"] if workerinput else None


def is_xdist_run(config):
    if hasattr(config, "workerinput"):
        return True
    if getattr(config.option, "numprocesses", None):
        return True
    return getattr(config.option, "dist", "no") not in ("no", None)


def dump_worker_results(worker_id):
    """Each xdist worker reports only the tests it ran, so persist them for the
    controller to merge instead of letting every process write its own report."""
    partial_dir = get_partial_results_dir()
    partial_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        # Screenshots are re-encoded from their on-disk paths at render time,
        # so the base64 blobs are dropped rather than shipped between processes.
        "extent": [dict(result, screenshot_base64=None) for result in EXTENT_RESULTS],
        "allure": ALLURE_RESULTS,
        "module_counts": MODULE_TEST_COUNTS,
    }
    (partial_dir / f"{worker_id}.json").write_text(
        json.dumps(payload), encoding="utf-8"
    )


def merge_worker_results():
    partial_dir = get_partial_results_dir()
    if not partial_dir.exists():
        return
    for path in sorted(partial_dir.glob("*.json")):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            print(f"WARNING: could not read partial results file {path}.")
            continue
        EXTENT_RESULTS.extend(payload.get("extent") or [])
        ALLURE_RESULTS.extend(payload.get("allure") or [])
        # Every worker collects the whole suite and runs a slice of it, so the
        # module totals are the per-module maximum rather than the sum.
        for module, count in (payload.get("module_counts") or {}).items():
            MODULE_TEST_COUNTS[module] = max(MODULE_TEST_COUNTS.get(module, 0), count)
    shutil.rmtree(partial_dir, ignore_errors=True)


PROJECT_NAME = "CBSE_CBA"
MODULE_NAMES = {
    "m1_item_bank_mgmt": "M1 - Item Bank Mgmt",
    "m2_web_portal_admin": "M2 - Web Portal Admin",
    "m3_item_testing": "M3 - Item Testing",
    "m4_qp_creation": "M4 - QP Creation",
    "m5_teacher_contribution": "M5 - Teacher Contribution",
}
MODULE_TEST_COUNTS = {display_name: 0 for display_name in MODULE_NAMES.values()}
logger = LogGenerator.loggen()

ENVIRONMENT_REACHABILITY = {}
SESSION_TIMING = {"start": None, "end": None}
# Set when a test fails with an infrastructure signature, and cleared by the
# next test's setup once the environment answers again. Without this, one
# outage cascades: every remaining test fails in seconds against a dead
# environment, and pytest-rerunfailures spends its retries there too.
INFRA_OUTAGE = {"pending": False}


def get_environment_info():
    """Best-effort snapshot of the environment/URL and browser under test."""
    try:
        base_url = ReadConfig.get_base_url()
    except Exception:
        base_url = "Not configured"
    try:
        browser_name = ReadConfig.get_browser_name()
    except Exception:
        browser_name = "Unknown"
    try:
        reachable = is_configured_environment_reachable() if base_url != "Not configured" else None
    except Exception:
        reachable = None
    return {"base_url": base_url, "browser": browser_name, "reachable": reachable}


def format_duration(seconds):
    seconds = int(round(seconds))
    hours, remainder = divmod(seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def get_execution_timing():
    start = SESSION_TIMING.get("start")
    end = SESSION_TIMING.get("end") or time.time()
    duration = (end - start) if start else 0
    return {
        "start": datetime.fromtimestamp(start).strftime("%Y-%m-%d %H:%M:%S") if start else "Unknown",
        "end": datetime.fromtimestamp(end).strftime("%Y-%m-%d %H:%M:%S") if start else "Unknown",
        "duration": format_duration(duration) if start else "Unknown",
    }


@pytest.hookimpl(tryfirst=True)
def pytest_collection_modifyitems(items):
    """Apply consistent suite markers from stable test naming conventions.

    Runs tryfirst because xdist turns an `xdist_group` marker into a nodeid
    suffix from its own collection hook - a marker added after that hook has
    run is simply ignored, and the grouping below silently does nothing.
    """
    for item in items:
        path = Path(str(item.fspath))
        if "tests" in path.parts:
            item.add_marker(pytest.mark.ui)
        if path.name == "test_teacher_login.py" or path.name.startswith("test_smoke_"):
            item.add_marker(pytest.mark.smoke)
        if "contracts" in path.stem:
            item.add_marker(pytest.mark.contract)
        if item.name.startswith("test_e2e_"):
            item.add_marker(pytest.mark.e2e)
        if item.get_closest_marker("serial"):
            # `serial` only documented an intent until now - nothing stopped
            # xdist from spreading those tests across workers. Putting them
            # all in one xdist group keeps them on a single worker (under
            # `--dist loadgroup`), which is what "must not run in parallel
            # with other tests" needs to actually mean.
            item.add_marker(pytest.mark.xdist_group("serial"))


def pytest_collection_finish(session):
    """Count the tests that will actually run, per module.

    Deliberately not done in pytest_collection_modifyitems above: that hook is
    tryfirst, so it sees the collected items *before* pytest applies `-m`
    deselection. Counting there attributed every deselected test to its module
    - the 25 `nightly` tests excluded by the default addopts, or the whole
    suite on a `-m smoke` run - and the report then flagged its own totals as
    inconsistent with the executed count. This hook runs once collection is
    final, so the totals only ever describe the selected run.
    """
    MODULE_TEST_COUNTS.clear()
    MODULE_TEST_COUNTS.update({display_name: 0 for display_name in MODULE_NAMES.values()})
    for item in session.items:
        module_name = get_module_name(item)
        MODULE_TEST_COUNTS[module_name] = MODULE_TEST_COUNTS.get(module_name, 0) + 1


def record_extent_result(entry):
    """pytest-rerunfailures replays the same test, so keep one row per test and
    let the retry badge carry the attempt count instead of stacking rows."""
    key = (entry.get("module"), entry.get("name"))
    for index, existing in enumerate(EXTENT_RESULTS):
        if (existing.get("module"), existing.get("name")) == key:
            EXTENT_RESULTS[index] = entry
            return
    EXTENT_RESULTS.append(entry)


def screenshot_as_base64(screenshot_path):
    return base64.b64encode(Path(screenshot_path).read_bytes()).decode("utf-8")


def one_line(value, max_length=240):
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text if len(text) <= max_length else text[: max_length - 1].rstrip() + "…"


def get_failure_one_line(report):
    lines = [line.strip() for line in str(report.longrepr or "").splitlines() if line.strip()]
    # Say it up front, so nobody triages an outage as a product bug.
    prefix = "[ENVIRONMENT] " if getattr(report, "infra_failure", False) else ""
    for line in reversed(lines):
        cleaned = re.sub(r"^E\s+", "", line).strip()
        if any(
            marker in cleaned
            for marker in (
                "AssertionError",
                "TimeoutException",
                "Exception:",
                "Error:",
            )
        ):
            return prefix + one_line(cleaned)
    return prefix + one_line(lines[-1]) if lines else "Test assertion failed."


def build_step_evidence_html(evidence_screenshots):
    image_cards = []
    for evidence in evidence_screenshots:
        evidence_path = evidence.get("path")
        if not evidence_path:
            continue
        try:
            evidence_base64 = screenshot_as_base64(evidence_path)
        except Exception:
            continue
        evidence_name = html.escape(evidence.get("name") or "Evidence Screenshot")
        image_cards.append(
            "<figure style='margin:12px 0;padding:10px;border:1px solid #d8dee9;border-radius:6px;'>"
            f"<figcaption style='font-weight:700;margin-bottom:8px;'>{evidence_name}</figcaption>"
            f"<img src='data:image/png;base64,{evidence_base64}' "
            "style='max-width:100%;border:1px solid #d8dee9;border-radius:6px;' "
            f"alt='{evidence_name}'>"
            "</figure>"
        )
    if not image_cards:
        return ""
    return (
        "<section>"
        "<h3 style='margin:8px 0;'>Step Evidence Screenshots</h3>"
        + "".join(image_cards)
        + "</section>"
    )


def _evidence_context_lines(evidence_screenshots):
    return [
        f"{e['name']}: {e['path']}"
        for e in evidence_screenshots
        if e.get("name") and e.get("path")
    ]


def _append_shared_extras(report, props):
    for value, name in (
        (props.get("item_set_id"), "Item Set ID"),
        (props.get("baseline_item_set_id"), "Baseline/IS10 Item Set ID"),
        (props.get("is12_item_set_id"), "Duplicate/IS12 Item Set ID"),
        (props.get("baseline_toast"), "Baseline/IS10 Toast"),
        (props.get("item_set_status"), "Initial Item Set Status"),
        (props.get("item_set_status_after_teacher_resubmit"), "Status After Teacher Resubmit"),
        (props.get("post_approval_status"), "Post Approval Status"),
    ):
        if value:
            report.extras.append(extras.text(value, name=name))


def _append_evidence_extras(report, evidence_screenshots):
    step_evidence_html = build_step_evidence_html(evidence_screenshots)
    if step_evidence_html:
        report.extras.append(extras.html(step_evidence_html))
    for evidence in evidence_screenshots:
        try:
            evidence_path = evidence.get("path")
            evidence_name = evidence.get("name") or "Evidence Screenshot"
            if evidence_path:
                report.extras.append(
                    extras.image(screenshot_as_base64(evidence_path), name=evidence_name)
                )
        except Exception:
            continue


def get_module_name(item):
    """Return the Excel/module folder name for report grouping."""
    path = Path(str(item.fspath))
    parts = [part.casefold() for part in path.parts]
    for folder, display_name in MODULE_NAMES.items():
        if folder in parts:
            return display_name
    return "Other"


def is_environment_connection_failure(report):
    """Identify infra/browser failures caused by an unreachable test environment."""
    failure_text = str(getattr(report, "longrepr", "") or "").casefold()
    return any(
        marker in failure_text
        for marker in (
            "m1_environment_unavailable",
            "net::err_connection_timed_out",
            "net::err_name_not_resolved",
            "net::err_connection_refused",
            "tcp connect",
        )
    )


def _tcp_reachable(url, timeout, attempts, retry_delay):
    parsed = urlparse(url)
    host = parsed.hostname
    port = parsed.port or (443 if parsed.scheme == "https" else 80)
    if not host:
        return False

    for attempt in range(attempts):
        try:
            with socket.create_connection((host, port), timeout=timeout):
                return True
        except OSError:
            if attempt < attempts - 1:
                time.sleep(retry_delay)
    return False


def is_configured_environment_reachable(timeout=5.0, attempts=3, retry_delay=2.0):
    """Small TCP preflight so unreachable UI environments fail fast.

    Retries before concluding the environment is down. A negative verdict is
    cached for the whole session and skips every M1 test with the run still
    exiting 0, so a single dropped packet used to turn a whole suite into a
    silent pass. Only a positive result is cheap to be wrong about.

    Both hosts are checked. The SPA host alone is not enough: during a
    2026-08-14 outage it served HTTP 200 throughout while the REST API on its
    separate host was unreachable, which strands the browser on the app's
    global "Loading..." screen - the SPA never gets the data it needs to render
    even a login form. A preflight that only saw the SPA host called that
    environment healthy and let 8 tests run into it.
    """
    base_url = ReadConfig.get_base_url()
    if base_url in ENVIRONMENT_REACHABILITY:
        return ENVIRONMENT_REACHABILITY[base_url]

    urls = [base_url]
    try:
        urls.append(ReadConfig.get_api_base_url())
    except Exception:
        # API host not configured - the SPA host is all there is to check.
        pass

    reachable = all(
        _tcp_reachable(url, timeout, attempts, retry_delay) for url in urls
    )
    ENVIRONMENT_REACHABILITY[base_url] = reachable
    return reachable


def compute_module_stats():
    module_stats = {
        module: {"total": total, "passed": 0, "failed": 0, "skipped": 0}
        for module, total in MODULE_TEST_COUNTS.items()
    }
    for result in EXTENT_RESULTS:
        module = result.get("module") or "Other"
        if module not in module_stats:
            module_stats[module] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
        module_stats[module][result["status"].lower()] += 1
    return module_stats


def compute_report_health(module_stats):
    """Cross-check report data for internal consistency and flag suspicious patterns."""
    integrity_warnings = []
    anomaly_modules = set()

    modules_total = sum(stats["total"] for stats in module_stats.values())
    executed_total = len(EXTENT_RESULTS)
    if modules_total and modules_total != executed_total:
        integrity_warnings.append(
            f"Module totals ({modules_total}) do not match executed test count ({executed_total})."
        )

    seen_names = {}
    for result in EXTENT_RESULTS:
        key = (result.get("module"), result.get("name"))
        seen_names[key] = seen_names.get(key, 0) + 1
    duplicates = [
        f"{module} / {name} (x{count})"
        for (module, name), count in seen_names.items()
        if count > 1
    ]
    if duplicates:
        integrity_warnings.append("Duplicate test entries detected: " + "; ".join(duplicates))

    missing_message = [
        result["name"]
        for result in EXTENT_RESULTS
        if result.get("status") == "FAILED" and not (result.get("message") or "").strip()
    ]
    if missing_message:
        integrity_warnings.append(
            "Failed tests with no captured failure message: " + ", ".join(missing_message)
        )

    for module, stats in module_stats.items():
        if stats["total"] > 0 and stats["passed"] == 0 and stats["failed"] > 0:
            anomaly_modules.add(module)

    failures_by_module = {}
    for result in EXTENT_RESULTS:
        if result.get("status") == "FAILED":
            failures_by_module.setdefault(result.get("module"), []).append(
                result.get("message", "")
            )
    for module, messages in failures_by_module.items():
        if len(messages) > 1 and len(set(messages)) == 1:
            anomaly_modules.add(module)
            integrity_warnings.append(
                f"All {len(messages)} failures in '{module}' share the identical error "
                f"({messages[0]!r}) — likely an environment/infra issue rather than distinct bugs."
            )

    return integrity_warnings, anomaly_modules


def parse_stage_item_counts(user_properties):
    """Read the optional 'stage_item_counts' test property: a JSON list of
    {"stage": name, "counts": {label: count}} entries recorded by a test as an
    item set moves across review stages (e.g. RWG -> SME -> Teacher)."""
    raw = user_properties.get("stage_item_counts") if hasattr(user_properties, "get") else None
    if not raw:
        return []
    try:
        stages = json.loads(raw)
    except Exception:
        return []
    for stage in stages:
        stage.setdefault("total", sum((stage.get("counts") or {}).values()))
    return stages


def build_stage_counts_html(stages):
    if not stages:
        return ""
    rows = []
    previous_total = None
    for stage in stages:
        counts = stage.get("counts") or {}
        total = stage.get("total", sum(counts.values()))
        reconciled = previous_total is None or previous_total == total
        flag = "Reconciled" if reconciled else f"Mismatch (expected {previous_total})"
        flag_class = "passed-text" if reconciled else "failed-text"
        counts_text = ", ".join(f"{label}: {value}" for label, value in counts.items())
        rows.append(
            "<tr>"
            f"<td>{html.escape(str(stage.get('stage') or ''))}</td>"
            f"<td>{html.escape(counts_text)}</td>"
            f"<td>{total}</td>"
            f"<td class='{flag_class}'>{html.escape(flag)}</td>"
            "</tr>"
        )
        previous_total = total
    return (
        "<h3 style='margin:8px 0;'>Item Count Reconciliation</h3>"
        "<table class='summary-table'>"
        "<thead><tr><th>Stage</th><th>Item Counts</th><th>Total</th><th>Reconciled</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody>"
        "</table>"
    )


def build_stage_counts_text(stages):
    if not stages:
        return ""
    lines = ["Item Count Reconciliation:"]
    previous_total = None
    for stage in stages:
        counts = stage.get("counts") or {}
        total = stage.get("total", sum(counts.values()))
        counts_text = ", ".join(f"{label}={value}" for label, value in counts.items())
        if previous_total is None or previous_total == total:
            flag = "reconciled"
        else:
            flag = f"MISMATCH (expected {previous_total})"
        lines.append(f"  {stage.get('stage')}: {counts_text} (total={total}) — {flag}")
        previous_total = total
    return "\n".join(lines)


def render_message_html(message, status):
    """Escape a test message and color the text after the PASSED:/FAILED: prefix."""
    text = message or ""
    color_class = {
        "PASSED": "passed-text",
        "FAILED": "failed-text",
        "SKIPPED": "skipped-text",
    }.get((status or "").upper())
    if not color_class:
        return html.escape(text)
    prefix, sep, rest = text.partition(":")
    if sep:
        return f"{html.escape(prefix)}:<span class='{color_class}'>{html.escape(rest)}</span>"
    return f"<span class='{color_class}'>{html.escape(text)}</span>"


def build_extent_report():
    reports_dir = get_reports_dir()
    reports_dir.mkdir(exist_ok=True)
    report_path = reports_dir / "extent_report.html"

    timing = get_execution_timing()
    environment = get_environment_info()
    reachable = environment.get("reachable")
    reachable_text = "Unknown" if reachable is None else ("Reachable" if reachable else "Unreachable")
    reachable_class = (
        "" if reachable is None else ("passed-text" if reachable else "failed-text")
    )
    run_info_html = (
        "<table class='summary-table run-info-table'>"
        "<tbody>"
        f"<tr><th>Environment URL</th><td>{html.escape(environment['base_url'])}</td></tr>"
        f"<tr><th>Environment Status</th><td class='{reachable_class}'>{reachable_text}</td></tr>"
        f"<tr><th>Browser</th><td>{html.escape(environment['browser'])}</td></tr>"
        f"<tr><th>Start Time</th><td>{html.escape(timing['start'])}</td></tr>"
        f"<tr><th>End Time</th><td>{html.escape(timing['end'])}</td></tr>"
        f"<tr><th>Duration</th><td>{html.escape(timing['duration'])}</td></tr>"
        "</tbody>"
        "</table>"
    )

    module_stats = compute_module_stats()
    integrity_warnings, anomaly_modules = compute_report_health(module_stats)
    warnings_html = ""
    if integrity_warnings:
        warning_items = "".join(f"<li>{html.escape(w)}</li>" for w in integrity_warnings)
        warnings_html = (
            "<div class='integrity-warnings'>"
            "<strong>&#9888; Report Validation Warnings</strong>"
            f"<ul>{warning_items}</ul>"
            "</div>"
        )

    total_all = sum(stats["total"] for stats in module_stats.values())
    passed_all = sum(stats["passed"] for stats in module_stats.values())
    failed_all = sum(stats["failed"] for stats in module_stats.values())
    skipped_all = sum(stats["skipped"] for stats in module_stats.values())
    executed_all = passed_all + failed_all + skipped_all
    pass_rate = (passed_all / executed_all * 100) if executed_all else 0

    stat_cards_html = (
        "<div class='stat-cards'>"
        f"<div class='stat-card total'><div class='stat-value'>{total_all}</div><div class='stat-label'>Total Tests</div></div>"
        f"<div class='stat-card passed'><div class='stat-value'>{passed_all}</div><div class='stat-label'>Passed</div></div>"
        f"<div class='stat-card failed'><div class='stat-value'>{failed_all}</div><div class='stat-label'>Failed</div></div>"
        f"<div class='stat-card skipped'><div class='stat-value'>{skipped_all}</div><div class='stat-label'>Expected Failures / Skipped</div></div>"
        f"<div class='stat-card rate'><div class='stat-value'>{pass_rate:.0f}%</div><div class='stat-label'>Pass Rate</div></div>"
        "</div>"
    )

    module_rows = []
    for module in sorted(module_stats):
        stats = module_stats[module]
        row_class = " class='anomaly-row'" if module in anomaly_modules else ""
        total = stats["total"] or 1
        passed_pct = stats["passed"] / total * 100
        failed_pct = stats["failed"] / total * 100
        skipped_pct = stats["skipped"] / total * 100
        bar_html = (
            "<div class='mini-bar'>"
            f"<span class='mini-bar-passed' style='width:{passed_pct:.1f}%'></span>"
            f"<span class='mini-bar-failed' style='width:{failed_pct:.1f}%'></span>"
            f"<span class='mini-bar-skipped' style='width:{skipped_pct:.1f}%'></span>"
            "</div>"
        )
        module_rows.append(
            f"<tr{row_class}>"
            f"<td>{html.escape(module)}</td>"
            f"<td>{stats['total']}</td>"
            f"<td class='passed-text'>{stats['passed']}</td>"
            f"<td class='failed-text'>{stats['failed']}</td>"
            f"<td class='skipped-text'>{stats['skipped']}</td>"
            f"<td>{bar_html}</td>"
            "</tr>"
        )

    rows = []
    for result in EXTENT_RESULTS:
        status_class = result["status"].lower()
        screenshot_entries = list(result.get("screenshots") or [])
        if result.get("screenshot_base64"):
            screenshot_entries.append(
                {
                    "name": "Screenshot",
                    "base64": result["screenshot_base64"],
                }
            )
        image_html = ""
        image_sections = []
        for screenshot in screenshot_entries:
            screenshot_base64 = screenshot.get("base64")
            screenshot_path = screenshot.get("path")
            if not screenshot_base64 and screenshot_path:
                try:
                    screenshot_base64 = screenshot_as_base64(screenshot_path)
                except Exception:
                    screenshot_base64 = None
            if not screenshot_base64:
                continue
            screenshot_name = html.escape(screenshot.get("name") or "Screenshot")
            image_sections.append(
                "<details>"
                f"<summary>{screenshot_name}</summary>"
                f"<img src='data:image/png;base64,{screenshot_base64}' alt='{screenshot_name}'>"
                "</details>"
            )
        if image_sections:
            image_html = "<div class='screenshots'>" + "".join(image_sections) + "</div>"
        if result.get("screenshot_base64"):
            result["screenshot_base64"] = None

        logged_in_user_html = (
            f"<p><strong>Logged In User:</strong> {html.escape(result['logged_in_user'])}</p>"
            if result.get("logged_in_user") else ""
        )
        stage_counts_html = result.get("stage_counts_html") or ""
        status_icon = {"passed": "&#10003;", "failed": "&#10007;", "skipped": "&#9888;"}.get(status_class, "")
        retry_badge_html = ""
        if result.get("retried"):
            attempt = result.get("execution_count") or 2
            retry_badge_html = f"<span class='retry-badge'>&#8635; Retried &middot; attempt {attempt}</span>"
        rows.append(
            f"<section class='test-card card-{status_class}'>"
            f"<div class='status {status_class}'><span class='status-icon'>{status_icon}</span>{result['status']}</div>"
            f"{retry_badge_html}"
            f"<p><strong>Module:</strong> {html.escape(result.get('module') or 'Other')}</p>"
            f"{logged_in_user_html}"
            f"<h2>{html.escape(result['name'])}</h2>"
            f"<p><strong>Message:</strong> {render_message_html(result.get('message'), result.get('status'))}</p>"
            + (
                "<details class='details-block'>"
                "<summary>Full Details / Traceback</summary>"
                f"<pre>{html.escape(result.get('details') or '')}</pre>"
                "</details>"
                if result.get("details")
                else ""
            )
            + f"{stage_counts_html}"
            f"{image_html}"
            "</section>"
        )

    report_path.write_text(
        f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{PROJECT_NAME} - Extent Report</title>
    <style>
        :root {{
            --passed: #0a7d47; --passed-bg: #e6f7ee; --passed-border: #a9e2c5;
            --failed: #c0261e; --failed-bg: #fdecec; --failed-border: #f3b7b2;
            --skipped: #8a6100; --skipped-bg: #fff6dd; --skipped-border: #f0d264;
            --ink: #1b2436; --muted: #5b6577; --border: #dde3ee;
            --card-bg: #ffffff; --page-bg: #eef1f7;
        }}
        * {{ box-sizing: border-box; }}
        body {{
            font-family: "Segoe UI", Roboto, Arial, sans-serif;
            margin: 0; padding: 28px 32px 48px;
            background: var(--page-bg); color: var(--ink);
            line-height: 1.5;
        }}
        .report-header {{
            background: linear-gradient(135deg, #16324f 0%, #1f6f5c 100%);
            color: #fff; border-radius: 12px; padding: 22px 28px; margin-bottom: 24px;
            box-shadow: 0 6px 18px rgba(20, 40, 70, 0.18);
        }}
        .report-header h1 {{ margin: 0; font-size: 24px; letter-spacing: 0.2px; }}
        .report-header .subtitle {{ margin-top: 4px; opacity: 0.85; font-size: 13px; }}
        h2 {{ font-size: 17px; margin: 28px 0 12px; color: var(--ink); }}
        h2.section-title {{ display: flex; align-items: center; gap: 8px; }}

        .stat-cards {{ display: grid; grid-template-columns: repeat(5, 1fr); gap: 14px; margin-bottom: 24px; }}
        .stat-card {{
            background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px;
            padding: 16px 14px; text-align: center; border-top: 4px solid var(--border);
            box-shadow: 0 1px 3px rgba(20,40,70,0.06);
        }}
        .stat-card .stat-value {{ font-size: 26px; font-weight: 800; }}
        .stat-card .stat-label {{ font-size: 12px; color: var(--muted); margin-top: 4px; text-transform: uppercase; letter-spacing: 0.4px; }}
        .stat-card.total {{ border-top-color: #3b5a8a; }}
        .stat-card.total .stat-value {{ color: #2c4770; }}
        .stat-card.passed {{ border-top-color: var(--passed); }}
        .stat-card.passed .stat-value {{ color: var(--passed); }}
        .stat-card.failed {{ border-top-color: var(--failed); }}
        .stat-card.failed .stat-value {{ color: var(--failed); }}
        .stat-card.skipped {{ border-top-color: var(--skipped); }}
        .stat-card.skipped .stat-value {{ color: var(--skipped); }}
        .stat-card.rate {{ border-top-color: #6b4fa0; }}
        .stat-card.rate .stat-value {{ color: #6b4fa0; }}

        .test-card {{
            background: var(--card-bg); border: 1px solid var(--border); border-left: 5px solid var(--border);
            border-radius: 10px; margin: 14px 0; padding: 18px 20px;
            box-shadow: 0 1px 4px rgba(20,40,70,0.06);
        }}
        .test-card h2 {{ margin: 10px 0 8px; font-size: 16px; }}
        .card-passed {{ border-left-color: var(--passed); }}
        .card-failed {{ border-left-color: var(--failed); }}
        .card-skipped {{ border-left-color: var(--skipped); }}

        .status {{
            display: inline-flex; align-items: center; gap: 6px;
            padding: 4px 12px; border-radius: 20px; font-weight: 700; font-size: 12px;
            letter-spacing: 0.3px; text-transform: uppercase; border: 1px solid transparent;
        }}
        .status-icon {{ font-size: 12px; }}
        .passed {{ background: var(--passed-bg); color: var(--passed); border-color: var(--passed-border); }}
        .failed {{ background: var(--failed-bg); color: var(--failed); border-color: var(--failed-border); }}
        .skipped {{ background: var(--skipped-bg); color: var(--skipped); border-color: var(--skipped-border); }}
        .retry-badge {{
            display: inline-flex; align-items: center; gap: 4px; margin-left: 8px;
            padding: 4px 12px; border-radius: 20px; font-weight: 700; font-size: 12px;
            letter-spacing: 0.3px; background: #eef1ff; color: #4338ca; border: 1px solid #c7d2fe;
        }}

        .summary-table {{ width: 100%; border-collapse: collapse; background: var(--card-bg); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(20,40,70,0.05); }}
        .summary-table th, .summary-table td {{ padding: 11px 16px; border-bottom: 1px solid #eef1f7; text-align: left; }}
        .summary-table th {{ background: #f3f6fb; font-size: 12px; text-transform: uppercase; letter-spacing: 0.4px; color: var(--muted); }}
        .summary-table tbody tr:last-child td {{ border-bottom: none; }}
        .summary-table tbody tr:hover {{ background: #f8fafd; }}
        .passed-text {{ color: var(--passed); font-weight: 700; }}
        .failed-text {{ color: var(--failed); font-weight: 700; }}
        .skipped-text {{ color: var(--skipped); font-weight: 700; }}

        .mini-bar {{ display: flex; width: 140px; height: 10px; border-radius: 6px; overflow: hidden; background: #eef1f7; }}
        .mini-bar-passed {{ background: var(--passed); height: 100%; }}
        .mini-bar-failed {{ background: var(--failed); height: 100%; }}
        .mini-bar-skipped {{ background: var(--skipped); height: 100%; }}

        pre {{ white-space: pre-wrap; background: #f4f6fa; padding: 12px 14px; border-radius: 8px; overflow: auto; font-size: 13px; border: 1px solid var(--border); }}
        img {{ max-width: 100%; border: 1px solid var(--border); border-radius: 8px; margin-top: 12px; }}
        summary {{ cursor: pointer; font-weight: 700; margin-top: 12px; color: #2c4770; }}
        summary:hover {{ color: #16324f; }}
        .screenshots details {{ margin-top: 12px; }}
        .run-info-table th {{ width: 180px; background: #f3f6fb; }}
        .anomaly-row {{ background: #fff2f2 !important; }}
        .integrity-warnings {{ background: var(--skipped-bg); border: 1px solid var(--skipped-border); border-radius: 10px; padding: 14px 18px; margin-bottom: 20px; }}
        .integrity-warnings ul {{ margin: 8px 0 0 0; padding-left: 20px; }}

        @media (max-width: 900px) {{
            .stat-cards {{ grid-template-columns: repeat(2, 1fr); }}
        }}
    </style>
</head>
<body>
    <div class="report-header">
        <h1>{PROJECT_NAME} &mdash; Extent Report</h1>
        <div class="subtitle">Generated {timing['end']} &middot; Duration {timing['duration']}</div>
    </div>
    {stat_cards_html}
    <h2 class="section-title">Run Details</h2>
    {run_info_html}
    {warnings_html}
    <h2 class="section-title">Module Summary</h2>
    <table class="summary-table">
        <thead>
            <tr>
                <th>Module</th>
                <th>Total tests</th>
                <th>Passed</th>
                <th>Failed</th>
                <th>Expected failures / skipped</th>
                <th>Distribution</th>
            </tr>
        </thead>
        <tbody>
            {"".join(module_rows)}
        </tbody>
    </table>
    <h2 class="section-title">Test Details</h2>
"""
        + "\n".join(rows)
        + """
</body>
</html>
""",
        encoding="utf-8",
    )
    return report_path


def excel_safe_text(value):
    text = str(value or "")
    return text[:32767]


def style_excel_sheet_header(row):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autofit_excel_columns(worksheet, max_width=80):
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = str(cell.value or "")
            max_length = max(max_length, min(len(value), max_width))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)


def build_excel_report():
    reports_dir = get_reports_dir()
    reports_dir.mkdir(exist_ok=True)
    report_path = reports_dir / "excel_report.xlsx"

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    results_sheet = workbook.create_sheet("Test Results")

    status_counts = {"PASSED": 0, "FAILED": 0, "SKIPPED": 0}
    module_counts = {}
    for result in EXTENT_RESULTS:
        status = result.get("status", "").upper()
        module = result.get("module") or "Other"
        status_counts[status] = status_counts.get(status, 0) + 1
        module_counts.setdefault(module, {"PASSED": 0, "FAILED": 0, "SKIPPED": 0})
        module_counts[module][status] = module_counts[module].get(status, 0) + 1

    timing = get_execution_timing()
    environment = get_environment_info()
    reachable = environment.get("reachable")
    reachable_text = "Unknown" if reachable is None else ("Reachable" if reachable else "Unreachable")
    module_stats = compute_module_stats()
    integrity_warnings, anomaly_modules = compute_report_health(module_stats)

    summary_sheet["A1"] = f"{PROJECT_NAME} Test Execution Summary"
    summary_sheet["A1"].font = Font(size=16, bold=True, color="1F4E78")
    summary_sheet.merge_cells("A1:E1")
    summary_sheet.append([])
    summary_sheet.append(["Environment URL", environment["base_url"], "Browser", environment["browser"]])
    summary_sheet.append(["Environment Status", reachable_text])
    summary_sheet.append(["Start Time", timing["start"], "End Time", timing["end"]])
    summary_sheet.append(["Duration", timing["duration"]])
    summary_sheet.append([])
    summary_sheet.append(["Total Tests", "Passed", "Failed", "Skipped", "Generated Report"])
    style_excel_sheet_header(summary_sheet[8])
    summary_sheet.append(
        [
            len(EXTENT_RESULTS),
            status_counts.get("PASSED", 0),
            status_counts.get("FAILED", 0),
            status_counts.get("SKIPPED", 0),
            str(report_path),
        ]
    )
    summary_sheet.append([])
    summary_sheet.append(["Module", "Passed", "Failed", "Skipped", "Total"])
    style_excel_sheet_header(summary_sheet[11])
    anomaly_fill = PatternFill("solid", fgColor="FDECEC")
    for module, counts in sorted(module_counts.items()):
        passed = counts.get("PASSED", 0)
        failed = counts.get("FAILED", 0)
        skipped = counts.get("SKIPPED", 0)
        summary_sheet.append([module, passed, failed, skipped, passed + failed + skipped])
        if module in anomaly_modules:
            for cell in summary_sheet[summary_sheet.max_row]:
                cell.fill = anomaly_fill

    if integrity_warnings:
        summary_sheet.append([])
        summary_sheet.append(["Report Validation Warnings"])
        summary_sheet[summary_sheet.max_row][0].font = Font(bold=True, color="B42318")
        for warning in integrity_warnings:
            summary_sheet.append([warning])

    result_headers = [
        "Status",
        "Module",
        "Logged In User",
        "Test Name",
        "Message",
        "Details",
        "Screenshots",
        "Retried",
    ]
    results_sheet.append(result_headers)
    style_excel_sheet_header(results_sheet[1])
    status_fills = {
        "PASSED": PatternFill("solid", fgColor="E2F0D9"),
        "FAILED": PatternFill("solid", fgColor="FCE4D6"),
        "SKIPPED": PatternFill("solid", fgColor="FFF2CC"),
    }
    message_font_colors = {
        "PASSED": "087443",
        "FAILED": "B42318",
        "SKIPPED": "8A6100",
    }

    for result in EXTENT_RESULTS:
        screenshots = []
        for screenshot in result.get("screenshots") or []:
            screenshot_path = screenshot.get("path")
            if screenshot_path:
                name = screenshot.get("name") or "Screenshot"
                screenshots.append(f"{name}: {screenshot_path}")
        execution_count = result.get("execution_count") or 1
        retried_text = f"Yes (attempt {execution_count})" if result.get("retried") else ""
        row = [
            result.get("status", ""),
            result.get("module", ""),
            result.get("logged_in_user", "") or "",
            result.get("name", ""),
            excel_safe_text(result.get("message", "")),
            excel_safe_text(result.get("details", "")),
            excel_safe_text("\n".join(screenshots)),
            retried_text,
        ]
        results_sheet.append(row)
        status_cell = results_sheet.cell(row=results_sheet.max_row, column=1)
        status_upper = str(status_cell.value).upper()
        status_cell.fill = status_fills.get(status_upper, PatternFill())
        status_cell.font = Font(bold=True)
        message_color = message_font_colors.get(status_upper)
        if message_color:
            message_cell = results_sheet.cell(row=results_sheet.max_row, column=5)
            message_cell.font = Font(color=message_color, bold=True)
        if retried_text:
            retried_cell = results_sheet.cell(row=results_sheet.max_row, column=8)
            retried_cell.fill = PatternFill("solid", fgColor="EEF1FF")
            retried_cell.font = Font(color="4338CA", bold=True)

    for worksheet in (summary_sheet, results_sheet):
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        autofit_excel_columns(worksheet)

    results_sheet.auto_filter.ref = results_sheet.dimensions
    results_sheet.column_dimensions["F"].width = 80
    results_sheet.column_dimensions["G"].width = 80
    workbook.save(report_path)

    try:
        saved_workbook = load_workbook(report_path, read_only=True)
        saved_rows = saved_workbook["Test Results"].max_row - 1
        saved_workbook.close()
        if saved_rows != len(EXTENT_RESULTS):
            print(
                f"WARNING: excel_report.xlsx has {saved_rows} result rows but "
                f"{len(EXTENT_RESULTS)} tests were recorded; the report file may be corrupt "
                "or truncated."
            )
    except Exception as error:
        print(f"WARNING: could not verify excel_report.xlsx after saving: {error}")


def reset_allure_results_dir():
    allure_results_dir = get_reports_dir() / "allure-results"
    allure_results_dir.mkdir(parents=True, exist_ok=True)
    for path in allure_results_dir.iterdir():
        if path.is_file():
            path.unlink()
    return allure_results_dir


def write_allure_attachment(allure_results_dir, content, extension):
    attachment_name = f"{uuid.uuid4()}-attachment.{extension}"
    attachment_path = allure_results_dir / attachment_name
    mode = "wb" if isinstance(content, bytes) else "w"
    encoding = None if isinstance(content, bytes) else "utf-8"
    with attachment_path.open(mode, encoding=encoding) as attachment_file:
        attachment_file.write(content)
    return attachment_name


def add_allure_result(
    name,
    status,
    message="",
    details="",
    screenshot_path=None,
    screenshot_name="Screenshot",
    manual_item_id=None,
    full_name="",
    suite="",
    module="Other",
    logged_in_user=None,
):
    entry = {
        "name": name,
        "status": status,
        "message": message,
        "details": details,
        "screenshot_path": screenshot_path,
        "screenshot_name": screenshot_name,
        "manual_item_id": manual_item_id,
        "full_name": full_name or name,
        "suite": suite or "Pytest",
        "module": module or "Other",
        "logged_in_user": logged_in_user,
    }
    key = (entry["module"], entry["name"])
    for index, existing in enumerate(ALLURE_RESULTS):
        if (existing.get("module"), existing.get("name")) == key:
            ALLURE_RESULTS[index] = entry
            return
    ALLURE_RESULTS.append(entry)


def build_allure_results():
    allure_results_dir = get_reports_dir() / "allure-results"
    allure_results_dir.mkdir(parents=True, exist_ok=True)

    for result in ALLURE_RESULTS:
        result_uuid = str(uuid.uuid4())
        start_time = int(time.time() * 1000)
        attachments = []
        parameters = []

        if result.get("logged_in_user"):
            parameters.append({"name": "Logged In User", "value": result["logged_in_user"]})
        if result.get("manual_item_id"):
            parameters.append({"name": "Manual Item ID", "value": result["manual_item_id"]})

        if result.get("message"):
            source = write_allure_attachment(allure_results_dir, result["message"], "txt")
            attachments.append(
                {
                    "name": "One-line Outcome",
                    "source": source,
                    "type": "text/plain",
                }
            )

        if result.get("screenshot_path"):
            screenshot_bytes = Path(result["screenshot_path"]).read_bytes()
            source = write_allure_attachment(allure_results_dir, screenshot_bytes, "png")
            attachments.append(
                {
                    "name": result.get("screenshot_name") or "Screenshot",
                    "source": source,
                    "type": "image/png",
                }
            )

        status_details = {}
        if result.get("details"):
            status_details = {"message": result.get("message") or "", "trace": result["details"]}

        allure_result = {
            "uuid": result_uuid,
            "historyId": result["name"],
            "testCaseId": result["name"],
            "name": result["name"],
            "fullName": result["full_name"],
            "description": result.get("message") or "",
            "status": result["status"].lower(),
            "statusDetails": status_details,
            "stage": "finished",
            "start": start_time,
            "stop": int(time.time() * 1000),
            "labels": [
                {"name": "framework", "value": "pytest"},
                {"name": "parentSuite", "value": PROJECT_NAME},
                {"name": "suite", "value": result["module"]},
                {"name": "subSuite", "value": result["suite"]},
            ],
            "parameters": parameters,
            "attachments": attachments,
        }

        result_path = allure_results_dir / f"{result_uuid}-result.json"
        result_path.write_text(json.dumps(allure_result, indent=2), encoding="utf-8")


def open_extent_report_if_enabled(report_path):
    if not ReadConfig.should_auto_open_extent():
        return
    if not report_path or not Path(report_path).exists():
        print("Extent report was not generated; nothing to auto-open.")
        return
    try:
        opened = webbrowser.open_new_tab(Path(report_path).resolve().as_uri())
        if not opened:
            print(f"Could not auto-open Extent report: {report_path}")
    except Exception as error:
        print(f"Could not auto-open Extent report: {error}")


def pytest_sessionstart(session):
    if not getattr(session.config.option, "collectonly", False):
        SESSION_TIMING["start"] = time.time()
        SESSION_TIMING["end"] = None
        EXTENT_RESULTS.clear()
        ALLURE_RESULTS.clear()
        # Only the controller clears shared output, so a worker starting late
        # cannot wipe results another worker has already handed back.
        if get_xdist_worker_id(session.config) is None:
            reset_allure_results_dir()
            shutil.rmtree(get_partial_results_dir(), ignore_errors=True)


def pytest_runtest_setup(item):
    """Hold the next test until a mid-run outage clears.

    The session-start preflight cannot help here - it already passed, and its
    verdict is cached. This runs before the browser fixture starts, so a test
    that would only have failed on a dead environment waits for it instead.
    """
    if not INFRA_OUTAGE["pending"]:
        return
    INFRA_OUTAGE["pending"] = False
    timeout = int(os.getenv("CBSE_INFRA_RECOVERY_TIMEOUT", "300"))
    if timeout <= 0:
        return
    logger.warning(
        "Previous test failed on an infrastructure error; waiting up to %ds for "
        "the environment before running %s.",
        timeout,
        item.name,
    )
    wait_for_environment(timeout=timeout)


def pytest_runtest_call(item):
    if (
        get_module_name(item) == "M1 - Item Bank Mgmt"
        and "setup" in item.fixturenames
        and not is_configured_environment_reachable()
    ):
        pytest.skip(
            "M1_ENVIRONMENT_UNAVAILABLE: configured CBSE UI base URL is not reachable."
        )


@pytest.fixture()
def setup(request):
    if get_module_name(request.node) == "M1 - Item Bank Mgmt" and not is_configured_environment_reachable():
        class UnreachableEnvironmentDriver:
            def quit(self):
                return None

        driver = UnreachableEnvironmentDriver()
        request.cls.driver = driver
        yield driver
        return

    browser = ReadConfig.get_browser_name().lower()
    browser_profile_dir = None
    headless = os.getenv("CBSE_HEADLESS", "").strip().casefold() in {
        "1",
        "true",
        "yes",
        "on",
    }

    # Give every test its own download folder so export/download assertions can
    # look for the file without colliding with the user's real Downloads.
    download_dir = Path(tempfile.mkdtemp(prefix="cbse_downloads_"))
    download_prefs = {
        "download.default_directory": str(download_dir),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }

    if browser == "chrome":
        browser_profile_dir = Path(tempfile.mkdtemp(prefix="cbse_chrome_profile_"))
        chrome_options = ChromeOptions()
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--no-first-run")
        chrome_options.add_argument("--no-default-browser-check")
        chrome_options.add_argument("--disable-background-timer-throttling")
        chrome_options.add_argument("--disable-backgrounding-occluded-windows")
        chrome_options.add_argument("--disable-renderer-backgrounding")
        chrome_options.add_argument("--disable-features=CalculateNativeWinOcclusion")
        chrome_options.add_argument("--remote-allow-origins=*")
        chrome_options.add_argument(f"--user-data-dir={browser_profile_dir}")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_experimental_option("prefs", download_prefs)
        if headless:
            chrome_options.add_argument("--headless=new")
        driver = webdriver.Chrome(
            service=ChromeService(ChromeDriverManager().install()),
            options=chrome_options,
        )
    elif browser == "edge":
        edge_options = EdgeOptions()
        edge_options.add_argument("--disable-dev-shm-usage")
        edge_options.add_argument("--disable-extensions")
        edge_options.add_argument("--disable-gpu")
        edge_options.add_argument("--no-sandbox")
        edge_options.add_argument("--window-size=1920,1080")
        edge_options.add_experimental_option("prefs", download_prefs)
        if headless:
            edge_options.add_argument("--headless=new")
        driver = webdriver.Edge(
            service=EdgeService(EdgeChromiumDriverManager().install()),
            options=edge_options,
        )
    else:
        raise ValueError(f"Unsupported browser: {browser}")

    if not headless:
        driver.maximize_window()
    driver.implicitly_wait(5)
    driver._download_dir = str(download_dir)
    request.cls.driver = driver

    yield driver

    try:
        driver.quit()
    except WebDriverException:
        pass
    if browser_profile_dir:
        shutil.rmtree(browser_profile_dir, ignore_errors=True)
    shutil.rmtree(download_dir, ignore_errors=True)


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item):
    outcome = yield
    report = outcome.get_result()
    report.extras = getattr(report, "extras", [])

    # An outage is not a defect. Flag it so the report says so, and so the next
    # test waits for the environment instead of failing into it.
    report.infra_failure = bool(
        report.failed and is_infrastructure_failure(report.longrepr)
    )
    if report.infra_failure:
        INFRA_OUTAGE["pending"] = True
        logger.warning(
            "%s failed on an infrastructure error, not a product defect.", item.name
        )

    user_properties = dict(item.user_properties)
    manual_item_id = user_properties.get("manual_item_id")
    item_set_id = user_properties.get("item_set_id")
    item_set_reviewer = user_properties.get("item_set_reviewer")
    item_set_status = user_properties.get("item_set_status")
    baseline_item_set_id = user_properties.get("baseline_item_set_id")
    baseline_toast = user_properties.get("baseline_toast")
    is12_item_set_id = user_properties.get("is12_item_set_id")
    item_set_status_after_teacher_resubmit = user_properties.get(
        "item_set_status_after_teacher_resubmit"
    )
    revised_item_ids = user_properties.get("revised_item_ids")
    rerun_qar_message = user_properties.get("rerun_qar_message")
    reviewer_approval_message = user_properties.get("reviewer_approval_message")
    pit_approval_message = user_properties.get("pit_approval_message")
    post_approval_status = user_properties.get("post_approval_status")
    qar_success_message = user_properties.get("qar_success_message")
    qar_success_screenshot = user_properties.get("qar_success_screenshot")
    evidence_screenshots = []
    if user_properties.get("evidence_screenshots"):
        try:
            evidence_screenshots = json.loads(user_properties["evidence_screenshots"])
        except Exception:
            evidence_screenshots = []
    stage_item_counts = parse_stage_item_counts(user_properties)
    stage_counts_html = build_stage_counts_html(stage_item_counts)
    stage_counts_text = build_stage_counts_text(stage_item_counts)
    display_name = item.name
    suite_name = item.cls.__name__ if item.cls else Path(str(item.fspath)).stem
    module_name = get_module_name(item)
    environment_status = user_properties.get("environment_status")
    result_description = one_line(user_properties.get("result_description"))
    result_checkpoint = one_line(user_properties.get("result_checkpoint"))
    driver = item.funcargs.get("setup")
    logged_in_user = getattr(driver, '_logged_in_user', None) if driver else None
    execution_count = getattr(item, "execution_count", 1) or 1
    is_retry = execution_count > 1
    retry_note = f"Retried — passed on attempt {execution_count}" if is_retry else ""

    if report.when == "call" and report.passed:
        passed_detail = result_description or one_line(qar_success_message)
        if not passed_detail:
            passed_detail = f"{display_name} completed successfully."
        pass_summary = f"PASSED: {passed_detail}"
        pass_screenshot_name = f"PASS — {passed_detail}"
        details_lines = [pass_summary]
        if is_retry:
            details_lines.append(retry_note)
        if logged_in_user:
            details_lines.append(f"Logged In User: {logged_in_user}")
        if environment_status:
            details_lines.append(environment_status)
        if item_set_id:
            details_lines.append(f"Item Set ID: {item_set_id}")
        if baseline_item_set_id:
            details_lines.append(f"Baseline/IS10 Item Set ID: {baseline_item_set_id}")
        if is12_item_set_id:
            details_lines.append(f"Duplicate/IS12 Item Set ID: {is12_item_set_id}")
        if baseline_toast:
            details_lines.append(f"Baseline/IS10 Toast: {baseline_toast}")
        if item_set_reviewer:
            details_lines.append(f"Reviewer/RWG: {item_set_reviewer}")
        if item_set_status:
            details_lines.append(f"Initial Status: {item_set_status}")
        if item_set_status_after_teacher_resubmit:
            details_lines.append(
                f"Status After Teacher Resubmit: {item_set_status_after_teacher_resubmit}"
            )
        if revised_item_ids:
            details_lines.append(f"Revised Item IDs: {revised_item_ids}")
        if rerun_qar_message:
            details_lines.append(f"Re-run QAR Message: {rerun_qar_message}")
        if reviewer_approval_message:
            details_lines.append(f"Reviewer Approval: {reviewer_approval_message}")
        if pit_approval_message:
            details_lines.append(f"PIT Approval: {pit_approval_message}")
        if post_approval_status:
            details_lines.append(f"Post Approval Status: {post_approval_status}")
        if manual_item_id:
            details_lines.append(f"Manual Item ID: {manual_item_id}")
        details_lines += _evidence_context_lines(evidence_screenshots)
        if stage_counts_text:
            details_lines.append(stage_counts_text)
        details = "\n".join(details_lines)
        _append_shared_extras(report, {
            "item_set_id": item_set_id,
            "baseline_item_set_id": baseline_item_set_id,
            "is12_item_set_id": is12_item_set_id,
            "baseline_toast": baseline_toast,
            "item_set_status": item_set_status,
            "item_set_status_after_teacher_resubmit": item_set_status_after_teacher_resubmit,
            "post_approval_status": post_approval_status,
        })
        if item_set_reviewer:
            report.extras.append(extras.text(item_set_reviewer, name="Reviewer/RWG"))
        if revised_item_ids:
            report.extras.append(extras.text(revised_item_ids, name="Revised Item IDs"))
        if rerun_qar_message:
            report.extras.append(extras.text(rerun_qar_message, name="Re-run QAR Message"))
        if reviewer_approval_message:
            report.extras.append(extras.text(reviewer_approval_message, name="Reviewer Approval"))
        if pit_approval_message:
            report.extras.append(extras.text(pit_approval_message, name="PIT Approval"))
        if manual_item_id:
            report.extras.append(extras.text(manual_item_id, name="Manual Item ID"))
        if qar_success_message:
            report.extras.append(extras.text(qar_success_message, name="QAR Success Message"))
        report.extras.append(extras.text(pass_summary, name="One-line Outcome"))
        if logged_in_user:
            report.extras.append(extras.text(logged_in_user, name="Logged In User"))
        if stage_counts_html:
            report.extras.append(extras.html(stage_counts_html))
        _append_evidence_extras(report, evidence_screenshots)
        screenshot_base64 = None
        pass_screenshot_path = qar_success_screenshot
        if not pass_screenshot_path:
            if driver:
                try:
                    pass_screenshot_path = ScreenshotUtils.capture(
                        driver, f"{item.name}_passed"
                    )
                except Exception:
                    pass
        if pass_screenshot_path:
            try:
                screenshot_base64 = screenshot_as_base64(pass_screenshot_path)
                report.extras.append(
                    extras.image(
                        screenshot_base64,
                        name=(
                            pass_screenshot_name
                        ),
                    )
                )
            except Exception:
                pass_screenshot_path = None
                screenshot_base64 = None
        if module_name == "M1 - Item Bank Mgmt" and item.name.startswith("test_e2e_") and driver:
            try:
                pdf_path = PDFReportUtils.generate(driver, item.name)
                report.extras.append(extras.text(pdf_path, name="PDF Report"))
            except Exception:
                pass
        extent_screenshots = [
            {"name": evidence.get("name"), "path": evidence.get("path")}
            for evidence in evidence_screenshots
            if evidence.get("path")
        ]
        if pass_screenshot_path:
            extent_screenshots.append(
                {
                    "name": (
                        pass_screenshot_name
                    ),
                    "path": pass_screenshot_path,
                }
            )
        record_extent_result(
            {
                "name": display_name,
                "status": "PASSED",
                "message": pass_summary,
                "details": details,
                "screenshot_base64": screenshot_base64,
                "screenshots": extent_screenshots,
                "module": module_name,
                "logged_in_user": logged_in_user,
                "stage_counts_html": stage_counts_html,
                "retried": is_retry,
                "execution_count": execution_count,
            }
        )
        add_allure_result(
            display_name,
            "PASSED",
            message=pass_summary,
            details=details,
            screenshot_path=pass_screenshot_path,
            screenshot_name=pass_screenshot_name,
            manual_item_id=manual_item_id,
            full_name=item.nodeid,
            suite=suite_name,
            module=module_name,
            logged_in_user=logged_in_user,
        )

    if report.when == "call" and report.failed:
        failure_reason = get_failure_one_line(report)
        if result_checkpoint:
            failure_summary = f"FAILED at {result_checkpoint}: {failure_reason}"
        else:
            failure_summary = f"FAILED: {failure_reason}"
        if is_retry:
            failure_summary = f"{failure_summary} (still failing after {execution_count} attempts)"
        failure_screenshot_name = f"FAIL — {one_line(failure_summary, max_length=180)}"
        screenshot_base64 = None
        screenshot_path = None
        if driver:
            try:
                screenshot_path = ScreenshotUtils.capture(driver, item.name)
                screenshot_base64 = screenshot_as_base64(screenshot_path)
                report.extras.append(
                    extras.image(
                        screenshot_base64,
                        name=failure_screenshot_name,
                    )
                )
            except Exception as e:
                # If screenshot capture fails (e.g., invalid session), continue without screenshot
                pass
        failure_details = str(report.longrepr)
        context_lines = [failure_summary]
        if is_retry:
            context_lines.append(f"Retry: failed again on attempt {execution_count}")
        if logged_in_user:
            context_lines.append(f"Logged In User: {logged_in_user}")
        report.extras.append(extras.text(failure_summary, name="One-line Outcome"))
        if logged_in_user:
            report.extras.append(extras.text(logged_in_user, name="Logged In User"))
        if item_set_id:
            context_lines.append(f"Item Set ID: {item_set_id}")
        if baseline_item_set_id:
            context_lines.append(f"Baseline/IS10 Item Set ID: {baseline_item_set_id}")
        if is12_item_set_id:
            context_lines.append(f"Duplicate/IS12 Item Set ID: {is12_item_set_id}")
        if baseline_toast:
            context_lines.append(f"Baseline/IS10 Toast: {baseline_toast}")
        if item_set_status:
            context_lines.append(f"Initial Status: {item_set_status}")
        if item_set_status_after_teacher_resubmit:
            context_lines.append(
                f"Status After Teacher Resubmit: {item_set_status_after_teacher_resubmit}"
            )
        if reviewer_approval_message:
            context_lines.append(f"Reviewer Approval: {reviewer_approval_message}")
        if pit_approval_message:
            context_lines.append(f"PIT Approval: {pit_approval_message}")
        if post_approval_status:
            context_lines.append(f"Post Approval Status: {post_approval_status}")
        context_lines += _evidence_context_lines(evidence_screenshots)
        if stage_counts_text:
            context_lines.append(stage_counts_text)
        if stage_counts_html:
            report.extras.append(extras.html(stage_counts_html))
        if context_lines:
            failure_details = "\n".join(context_lines) + "\n\n" + failure_details
            _append_shared_extras(report, {
                "item_set_id": item_set_id,
                "baseline_item_set_id": baseline_item_set_id,
                "is12_item_set_id": is12_item_set_id,
                "baseline_toast": baseline_toast,
                "item_set_status": item_set_status,
                "item_set_status_after_teacher_resubmit": item_set_status_after_teacher_resubmit,
                "post_approval_status": post_approval_status,
            })
            _append_evidence_extras(report, evidence_screenshots)
        extent_screenshots = [
            {"name": evidence.get("name"), "path": evidence.get("path")}
            for evidence in evidence_screenshots
            if evidence.get("path")
        ]
        if screenshot_path:
            extent_screenshots.append(
                {
                    "name": failure_screenshot_name,
                    "path": screenshot_path,
                }
            )
        record_extent_result(
            {
                "name": display_name,
                "status": "FAILED",
                "message": failure_summary,
                "details": failure_details,
                "screenshot_base64": screenshot_base64,
                "screenshots": extent_screenshots,
                "module": module_name,
                "logged_in_user": logged_in_user,
                "stage_counts_html": stage_counts_html,
                "retried": is_retry,
                "execution_count": execution_count,
            }
        )
        add_allure_result(
            display_name,
            "FAILED",
            message=failure_summary,
            details=failure_details,
            screenshot_path=screenshot_path,
            screenshot_name=failure_screenshot_name,
            full_name=item.nodeid,
            suite=suite_name,
            module=module_name,
            logged_in_user=logged_in_user,
        )

    if report.when in ("setup", "call") and report.skipped:
        if hasattr(report, "wasxfail"):
            skip_message = "Expected failure / known issue."
            skip_details = getattr(report, "wasxfail", "Expected failure")
        else:
            # A plain skip reports longrepr as (path, lineno, reason).
            longrepr = getattr(report, "longrepr", None)
            skip_details = (
                str(longrepr[2])
                if isinstance(longrepr, tuple) and len(longrepr) == 3
                else str(longrepr or "")
            )
            skip_message = one_line(skip_details) or "Test was skipped."
        record_extent_result(
            {
                "name": display_name,
                "status": "SKIPPED",
                "message": skip_message,
                "details": skip_details,
                "screenshot_base64": None,
                "module": module_name,
                "logged_in_user": logged_in_user,
            }
        )
        add_allure_result(
            display_name,
            "SKIPPED",
            message=skip_message,
            details=skip_details,
            full_name=item.nodeid,
            suite=suite_name,
            module=module_name,
        )

    if report.when == "call" and logged_in_user:
        print(f"\n  [Login] User: {logged_in_user}", flush=True)


def pytest_sessionfinish(session, exitstatus):
    if getattr(session.config.option, "collectonly", False):
        return
    SESSION_TIMING["end"] = time.time()
    worker_id = get_xdist_worker_id(session.config)
    if worker_id is not None:
        dump_worker_results(worker_id)
        return
    if is_xdist_run(session.config):
        merge_worker_results()
    extent_report_path = build_extent_report()
    build_excel_report()
    build_allure_results()
    open_extent_report_if_enabled(extent_report_path)
