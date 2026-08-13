import json
import re
from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from selenium.webdriver.common.by import By

from pages.common.login_page import LoginPage
from pages.rwg.review_queue_page import RWGReviewQueuePage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.qar_retry_flow import run_qar_need_improvement_retry_loop
from utilities.read_config import ReadConfig


# Logs in across the configured RWG accounts to prove routing, so it must not run
# alongside another flow driving the same reviewer (one session per account).
@pytest.mark.serial
@pytest.mark.e2e
@pytest.mark.usefixtures("setup")
class TestE2EQARNeedImprovementRetryFlow:
    """QAR pass, bounded correction, and terminal blocking workflows."""

    QUESTION_COUNT = 3
    MAX_QAR_RETRIES = 3

    # Byte-identical rows in one workbook are collapsed into a single item during
    # upload, so QAR never sees a duplicate pair and passes the set at 100%.
    # These paraphrases survive ingest as distinct items while still being
    # duplicate content for QAR to flag.
    NEAR_DUPLICATE_QUESTIONS = (
        "Does one hour contain sixty minutes?",
        "Is one hour the same as sixty minutes?",
        "Does one hour have sixty minutes in it?",
    )

    @staticmethod
    def make_prefix(scenario):
        return f"QAR_AUTO_RETRY_{scenario.upper()}_{uuid4().hex[:10]}"

    @classmethod
    def build_workbook(cls, tmp_path, prefix, questions):
        assert len(questions) == cls.QUESTION_COUNT
        source = Path(ReadConfig.get_upload_item_file_path())
        target = tmp_path / f"{prefix}.xlsx"
        copy2(source, target)
        workbook = load_workbook(target)
        worksheet = workbook.active

        for offset, question in enumerate(questions):
            row = offset + 2
            if row > 2:
                for column in range(1, 25):
                    source_cell = worksheet.cell(2, column)
                    target_cell = worksheet.cell(row, column)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell._style = copy_cell_style(source_cell._style)
                    target_cell.alignment = copy_cell_style(source_cell.alignment)
            worksheet.cell(row, 5).value = offset + 1
            worksheet.cell(row, 10).value = "True or False"
            worksheet.cell(row, 11).value = question
            worksheet.cell(row, 12).value = None
            for column in range(13, 21):
                worksheet.cell(row, column).value = None
            worksheet.cell(row, 21).value = "True"
            worksheet.cell(row, 22).value = None
            worksheet.cell(row, 23).value = (
                f"{prefix}: the statement is correct and uses valid Grade 1 content."
            )
            worksheet.cell(row, 24).value = "1"

        for row in range(cls.QUESTION_COUNT + 2, worksheet.max_row + 1):
            for column in range(1, 25):
                worksheet.cell(row, column).value = None

        workbook.save(target)
        workbook.close()
        return target

    @staticmethod
    def record_stage_item_counts(request, stage, counts):
        """Append an item-count snapshot for this stage so the report can show
        the breakdown reconciling across the workflow (e.g. QAR result -> post-edit re-run)."""
        stages = []
        existing_index = None
        for index, (key, value) in enumerate(request.node.user_properties):
            if key == "stage_item_counts":
                existing_index = index
                try:
                    stages = json.loads(value)
                except Exception:
                    stages = []
                break
        stages.append({"stage": stage, "counts": counts})
        entry = ("stage_item_counts", json.dumps(stages))
        if existing_index is not None:
            request.node.user_properties[existing_index] = entry
        else:
            request.node.user_properties.append(entry)

    def reset_and_login(self, username):
        page = UploadItemFilePage(self.driver)
        page.reset_browser_session_to_login()
        LoginPage(self.driver).login_to_application(
            username,
            ReadConfig.get_password_for_username(username),
        )
        page.close_popup_if_open()
        return page

    def login_as_sme(self):
        return self.reset_and_login(ReadConfig.get_sme2_username())

    def upload_scenario(self, workbook_path):
        self.driver.get(ReadConfig.get_base_url())
        sme_username = ReadConfig.get_sme2_username()
        LoginPage(self.driver).login_to_application(
            sme_username,
            ReadConfig.get_password_for_username(sme_username),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()
        _, _, item_ids, qar_message = page.upload_item_file_and_submit_for_qar(
            workbook_path
        )
        assert len(item_ids) == self.QUESTION_COUNT, (
            f"Expected {self.QUESTION_COUNT} uploaded items, received {item_ids}."
        )
        return {
            "item_ids": item_ids,
            "item_set_id": page.get_item_set_id_from_item_ids(item_ids),
            "qar_message": qar_message,
        }

    def open_sme_item_set(self, item_set_id):
        page = self.login_as_sme()
        page.open_sets_module()
        page.open_item_set_from_sets_list(item_set_id)
        return page

    def get_sme_set_row_text(self, item_set_id):
        page = self.login_as_sme()
        page.open_sets_module()
        row = page.wait_utils.until_visible(
            (
                By.XPATH,
                f"//table//tbody/tr[contains(normalize-space(), '{item_set_id}')]",
            ),
            timeout=30,
        )
        return row.text

    def get_rwg_queue_snapshots(self, item_set_id):
        snapshots = {}
        for username in dict.fromkeys(ReadConfig.get_role_usernames("rwg")):
            self.reset_and_login(username)
            queue_text = RWGReviewQueuePage(self.driver).get_queue_body_text()
            snapshots[username] = {
                "contains_item_set": item_set_id in queue_text,
                "text": queue_text,
            }
        return snapshots

    def assert_blocked_from_rwg(self, item_set_id, failed_item_ids):
        row_text = self.get_sme_set_row_text(item_set_id)
        normalized = row_text.casefold()
        assert "under review" not in normalized, (
            f"Item set {item_set_id} moved to Under Review while items "
            f"{failed_item_ids} still Need Improvement: {row_text}"
        )
        assert any(
            marker in normalized
            for marker in ("pending", "qar", "revise", "revision", "improvement")
        ), f"Blocked set {item_set_id} showed an unexpected SME status: {row_text}"

        snapshots = self.get_rwg_queue_snapshots(item_set_id)
        visible_to = [
            username
            for username, evidence in snapshots.items()
            if evidence["contains_item_set"]
        ]
        assert not visible_to, (
            f"Item set {item_set_id} reached RWG queues {visible_to} while "
            f"{failed_item_ids} still Need Improvement."
        )

    def assert_routed_to_rwg(self, item_set_id):
        snapshots = self.get_rwg_queue_snapshots(item_set_id)
        visible_to = [
            username
            for username, evidence in snapshots.items()
            if evidence["contains_item_set"]
        ]
        assert visible_to, (
            f"QAR-passed item set {item_set_id} was absent from every configured "
            f"RWG queue: {list(snapshots)}"
        )

    @staticmethod
    def valid_correction(prefix):
        valid_questions = (
            "A class starts at 8:00 and ends at 9:00. Is its duration one hour?",
            "The minute hand makes one full turn in sixty minutes. Is this true?",
            "Tuesday comes immediately after Monday. Is this statement true?",
            "Seven o'clock is earlier than eight o'clock. Is this statement true?",
            "Noon occurs after morning. Is this statement true?",
        )

        def correction(item_id, retry_number):
            match = re.search(r"-i(\d+)$", item_id, re.IGNORECASE)
            item_number = int(match.group(1)) if match else sum(map(ord, item_id))
            question = valid_questions[(item_number - 1) % len(valid_questions)]
            return {
                "question": f"{prefix} corrected item {item_number}: {question}",
                "revision_note": (
                    f"Retry {retry_number}: replaced duplicate content with a "
                    "known-valid, unique Grade 1 question."
                ),
            }

        return correction

    @classmethod
    def persistent_duplicate_correction(cls, prefix):
        def correction(item_id, retry_number):
            # Keep the items paraphrases of one another rather than identical
            # text, so each retry resubmits duplicate content without the items
            # merging into one.
            variant = cls.NEAR_DUPLICATE_QUESTIONS[
                sum(map(ord, item_id)) % len(cls.NEAR_DUPLICATE_QUESTIONS)
            ]
            return {
                "question": f"{prefix} persistent duplicate: {variant}",
                "revision_note": (
                    f"Retry {retry_number}: intentionally retained duplicate content "
                    "to verify the retry ceiling."
                ),
            }

        return correction

    def run_retry_flow(
        self,
        item_set_id,
        all_item_ids,
        correction_factory,
        max_retries=3,
    ):
        active_page = {"page": None}

        def get_failed_item_ids():
            page = self.open_sme_item_set(item_set_id)
            active_page["page"] = page
            statuses = page.get_qar_item_statuses()
            status_keys = {
                page.compact_item_id(item_id).casefold()
                for item_id in statuses
            }
            missing_item_ids = [
                item_id
                for item_id in all_item_ids
                if page.compact_item_id(item_id).casefold() not in status_keys
            ]
            assert not missing_item_ids, (
                f"QAR results for {item_set_id} did not show statuses for every "
                f"uploaded item. Missing: {missing_item_ids}; visible: {statuses}."
            )
            retry_statuses = {
                status.casefold() for status in page.QAR_RETRY_STATUS_LABELS
            }
            return [
                item_id
                for item_id, status in statuses.items()
                if status.casefold() in retry_statuses
            ]

        def correct_items(item_ids, retry_number):
            page = self.open_sme_item_set(item_set_id)
            active_page["page"] = page
            return page.revise_qar_need_improvement_items(
                item_set_id,
                item_ids,
                correction_factory,
                retry_number,
            )

        def rerun_qar(retry_number):
            page = active_page["page"]
            assert page is not None
            assert page.is_rerun_qar_enabled(), (
                f"Re-run QAR was disabled after correction attempt {retry_number} "
                f"for {item_set_id}."
            )
            result = page.rerun_qar_if_enabled()
            assert "disabled" not in result.casefold() and "not available" not in result.casefold(), (
                f"QAR retry {retry_number} could not start for {item_set_id}: {result}"
            )
            return result

        return run_qar_need_improvement_retry_loop(
            item_set_id=item_set_id,
            get_need_improvement_item_ids=get_failed_item_ids,
            correct_items=correct_items,
            rerun_qar=rerun_qar,
            assert_blocked_from_rwg=self.assert_blocked_from_rwg,
            assert_routed_to_rwg=self.assert_routed_to_rwg,
            max_retries=max_retries,
        )

    def test_e2e_qar_happy_path_routes_to_rwg_without_retry(self, tmp_path):
        prefix = self.make_prefix("happy")
        workbook = self.build_workbook(
            tmp_path,
            prefix,
            [
                f"{prefix}: Is 8:00 earlier than 9:00?",
                f"{prefix}: Does one hour contain sixty minutes?",
                f"{prefix}: Does Tuesday come immediately after Monday?",
            ],
        )
        submission = self.upload_scenario(workbook)

        result = self.run_retry_flow(
            submission["item_set_id"],
            submission["item_ids"],
            self.valid_correction(prefix),
            max_retries=self.MAX_QAR_RETRIES,
        )

        assert result.retry_count == 0
        assert result.final_need_improvement_item_ids == ()

    def test_e2e_qar_retry_path_passes_after_one_edit_and_rerun(self, tmp_path, request):
        prefix = self.make_prefix("retry")
        workbook = self.build_workbook(
            tmp_path,
            prefix,
            [
                f"{prefix}: {self.NEAR_DUPLICATE_QUESTIONS[0]}",
                f"{prefix}: {self.NEAR_DUPLICATE_QUESTIONS[1]}",
                f"{prefix}: Is 7:00 earlier than 8:00?",
            ],
        )
        submission = self.upload_scenario(workbook)
        initial_failed = self.open_sme_item_set(
            submission["item_set_id"]
        ).get_qar_need_improvement_item_ids()
        assert initial_failed, "Duplicate fixture did not produce Need Improvement items."
        total_items = len(submission["item_ids"])
        self.record_stage_item_counts(
            request,
            "Initial QAR Result (SME2 submission)",
            {
                "approved": total_items - len(initial_failed),
                "need_improvement": len(initial_failed),
            },
        )

        result = self.run_retry_flow(
            submission["item_set_id"],
            submission["item_ids"],
            self.valid_correction(prefix),
            max_retries=self.MAX_QAR_RETRIES,
        )

        assert result.retry_count == 1, (
            f"Expected one edit/rerun cycle, received {result.retry_count}."
        )
        assert set(result.revised_item_ids_by_retry[0]) == set(initial_failed)
        self.record_stage_item_counts(
            request,
            "After SME2 Correction + Re-run QAR",
            {
                "approved": total_items - len(result.final_need_improvement_item_ids),
                "need_improvement": len(result.final_need_improvement_item_ids),
            },
        )

    def test_e2e_qar_failure_path_stops_after_three_retries_and_stays_blocked(
        self,
        tmp_path,
    ):
        prefix = self.make_prefix("failure")
        workbook = self.build_workbook(
            tmp_path,
            prefix,
            [
                f"{prefix} persistent duplicate: {question}"
                for question in self.NEAR_DUPLICATE_QUESTIONS
            ],
        )
        submission = self.upload_scenario(workbook)
        item_set_id = submission["item_set_id"]

        with pytest.raises(
            AssertionError,
            match=r"still 'Need Improvement' after 3 retries.*remains blocked from RWG",
        ) as error:
            self.run_retry_flow(
                item_set_id,
                submission["item_ids"],
                self.persistent_duplicate_correction(prefix),
                max_retries=self.MAX_QAR_RETRIES,
            )

        assert item_set_id in str(error.value)
        remaining = self.open_sme_item_set(
            item_set_id
        ).get_qar_need_improvement_item_ids()
        assert remaining, (
            f"Failure fixture unexpectedly cleared all QAR failures for {item_set_id}."
        )
        self.assert_blocked_from_rwg(item_set_id, tuple(remaining))
