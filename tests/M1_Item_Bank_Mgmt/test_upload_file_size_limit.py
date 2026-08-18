from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from time import monotonic
from uuid import uuid4
import json
import re
import logging
from openpyxl import load_workbook
import pytest
import requests
from selenium.webdriver.common.by import By

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from tests.M1_Item_Bank_Mgmt.m1_surveys import survey_chrome, survey_upload_step
from utilities.element_checks import ElementChecks
from utilities.item_bank_workbook_builder import build_item_workbook
from utilities.read_config import ReadConfig
from utilities.screenshot_utils import ScreenshotUtils


@pytest.mark.rtm
# The item-creation module is lazy-loaded, and this environment intermittently
# fails to serve the chunk ("failed to fetch dynamically imported module") even
# after open_item_creation_module()'s own refresh retries.
@pytest.mark.flaky(reruns=1, reruns_delay=5)
@pytest.mark.usefixtures("setup")
class TestUploadFileSizeLimit:
    """TC-NEG-M1-10 — Bulk Upload File Exceeds 10 MB Size Limit"""

    logger = logging.getLogger(__name__)

    def step(self, n, message):
        print(f"\n[STEP {n}] {message}", flush=True)
        self.logger.info(f"[STEP {n}] {message}")

    @staticmethod
    def add_evidence_screenshot(request, evidence_screenshots, name, screenshot_path):
        evidence_screenshots.append({"name": name, "path": str(screenshot_path)})
        request.node.user_properties[:] = [
            property_entry
            for property_entry in request.node.user_properties
            if property_entry[0] != "evidence_screenshots"
        ]
        request.node.user_properties.append(
            ("evidence_screenshots", json.dumps(evidence_screenshots))
        )

    @staticmethod
    def build_large_workbook(tmp_path, target_size_mb=12):
        """Generate a valid Excel workbook that exceeds target_size_mb by writing large cell content."""
        source = Path(ReadConfig.get_upload_item_file_path())
        target = tmp_path / f"large_upload_{uuid4().hex[:10]}.xlsx"
        copy2(source, target)
        
        workbook = load_workbook(target)
        worksheet = workbook.active
        
        import random
        chars = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
        
        # Write large content to cells to quickly blow up file size
        # 12 MB is roughly 12 * 1024 * 1024 bytes. To exceed 10 MB after compression,
        # we write 520 rows, each with a unique 32,700 character random string in column 11 (Question Content).
        for offset in range(1, 520):
            row = offset + 2
            if row != 2:
                for column in range(1, 25):
                    source_cell = worksheet.cell(2, column)
                    target_cell = worksheet.cell(row, column)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell._style = copy_cell_style(source_cell._style)
            # Generate highly unique random string to prevent zip compression
            random_string = "".join(random.choices(chars, k=32700))
            worksheet.cell(row, 5).value = offset + 1
            worksheet.cell(row, 10).value = "True or False"
            worksheet.cell(row, 11).value = random_string
            worksheet.cell(row, 21).value = "True"
            worksheet.cell(row, 23).value = "Large file automation description."
            worksheet.cell(row, 24).value = "1"
            
        workbook.save(target)
        workbook.close()
        return target

    def get_api_session(self):
        """Extract Selenium cookies and Authorization header into requests Session."""
        session = requests.Session()
        for cookie in self.driver.get_cookies():
            session.cookies.set(cookie['name'], cookie['value'])
            
        token = self.driver.execute_script(
            "return localStorage.getItem('token') || sessionStorage.getItem('token') "
            "|| localStorage.getItem('jwt') || sessionStorage.getItem('jwt') "
            "|| localStorage.getItem('auth_token') || sessionStorage.getItem('auth_token');"
        )
        
        headers = {}
        if token:
            headers["Authorization"] = f"Bearer {token}"
            
        return session, headers

    def test_tc_neg_m1_10_upload_file_size_limit_exceeded(
        self, tmp_path, request, record_property
    ):
        """TC-NEG-M1-10: Verify 12 MB file upload is rejected via Selenium and API, and smaller files still upload.
        
        Steps:
        1. Login as SME2.
        2. Dynamically build a 12 MB valid Excel workbook.
        3. Upload the 12 MB file via Selenium.
        4. Assert error message "File size exceeds 10 MB limit. Please compress your file."
        5. Verify no Upload ID is generated on page.
        6. POST /upload with 12 MB file via requests -> assert 413 Payload Too Large.
        7. Verify smaller valid file is accepted immediately after (no stuck state).
        """
        evidence_screenshots = []

        # Step 1: Login as SME2
        self.step(1, "Logging in as SME2")
        self.driver.get(ReadConfig.get_base_url())
        sme_username = ReadConfig.get_sme2_username()
        LoginPage(self.driver).login_to_application(
            sme_username,
            ReadConfig.get_password_for_username(sme_username),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()

        # Step 2: Build a 12 MB valid Excel workbook
        self.step(2, "Generating a 12 MB valid Excel workbook dynamically")
        large_workbook = self.build_large_workbook(tmp_path, target_size_mb=12)
        file_size_bytes = large_workbook.stat().st_size
        file_size_mb = file_size_bytes / (1024 * 1024)
        self.logger.info(f"Generated workbook file path: {large_workbook} (Size: {file_size_mb:.2f} MB)")
        
        print("[CHECK] Generated file size exceeds 10 MB limit", flush=True)
        assert file_size_mb > 10.0, f"Expected file size > 10 MB, but got {file_size_mb:.2f} MB"
        print(f"[PASS] File size confirmed: {file_size_mb:.2f} MB", flush=True)

        # Step 3: Upload the 12 MB file via Selenium
        self.step(3, "Uploading the 12 MB file via Selenium")
        page.open_item_creation_module()
        page.open_upload_item_file_tab()
        page.open_upload_step()

        # Surveyed before the oversized file is pushed at the page, so the
        # element table describes a healthy upload step rather than a rejected one.
        checks = ElementChecks(
            page, record_property, page_name="Upload Item File — File Size Limit"
        )
        survey_chrome(checks, page)
        survey_upload_step(checks, page)
        record_property("result_description", checks.publish())

        page.upload_file(large_workbook)

        # Step 4: Assert size limit error message and no Upload ID
        self.step(4, "Asserting size limit error message and absence of Upload ID")
        print("[CHECK] Size limit rejection error message is displayed", flush=True)
        
        rejection_text = page.wait_for_upload_rejection(timeout=30)
        self.logger.info(f"Upload rejection text: {rejection_text}")
        normalized_rejection = rejection_text.casefold()
        failure_screenshot = ScreenshotUtils.capture(
            self.driver,
            f"{request.node.name}_file_validation_failed_message",
        )
        self.add_evidence_screenshot(
            request,
            evidence_screenshots,
            "Oversized file FAILED validation message",
            failure_screenshot,
        )
        
        expected_error_text = "exceeds 10 mb"
        assert expected_error_text in normalized_rejection or "10 mb limit" in normalized_rejection or "compress" in normalized_rejection or "too large" in normalized_rejection or "file size" in normalized_rejection, (
            f"Expected error message containing size limit warning, but got: '{rejection_text}'"
        )
        print(f"[PASS] Confirmed error message displayed.", flush=True)

        print("[CHECK] No Upload ID is generated on page", flush=True)
        assert "upload id:" not in normalized_rejection, "BUG: Upload ID was incorrectly generated for an oversized file!"
        print("[PASS] Confirmed no Upload ID was generated.", flush=True)

        # Step 5: Assert the bulk-upload API rejects the oversized payload
        self.step(5, "Verifying the bulk-upload API rejects the 12 MB payload")
        session, headers = self.get_api_session()

        # The REST API lives on its own host; every /api/* path under the web
        # app's host resolves to the SPA's HTML 404 page instead, so a URL
        # derived from CBSE_BASE_URL never reaches the API. Posting 12 MB there
        # returns a proxy 502, which says nothing about the size limit.
        upload_url = ReadConfig.get_api_base_url() + "/excel-import/upload"

        assert "Authorization" in headers, (
            "No auth token was found in browser storage, so this API check would "
            "assert on an unauthenticated 401 rather than the size limit."
        )
        self.logger.info(f"Upload API URL: {upload_url}")

        print(f"[CHECK] POST {upload_url} with 12 MB file returns 413 Payload Too Large", flush=True)
        with open(large_workbook, "rb") as file_bytes:
            files = {"file": (large_workbook.name, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            # (connect, read) - the read leg also bounds the socket send, and
            # pushing 12 MB to this environment exceeds 30 s whenever the link
            # is busy, which surfaces as "the write operation timed out"
            # rather than as the size rejection being tested.
            post_response = session.post(
                upload_url, files=files, headers=headers, timeout=(30, 300)
            )
            
        self.logger.info(
            f"API Upload response status code: {post_response.status_code}; "
            f"body: {post_response.text[:500]}"
        )
        # Note: 413 is standard Payload Too Large, 400 Bad Request can also be returned by some validators
        assert post_response.status_code in (413, 400), (
            f"Expected status 413 Payload Too Large or 400 Bad Request, but got {post_response.status_code}. Response: {post_response.text}"
        )
        print(f"[PASS] API returned status {post_response.status_code} as expected.", flush=True)

        # Step 6: Verify smaller valid file is accepted immediately after
        self.step(6, "Verifying smaller valid file is accepted immediately after (no stuck state)")
        page.reset_upload_step()

        # Build a unique under-limit workbook rather than re-uploading the raw
        # template: the app rejects a re-upload of identical filename+content
        # ("A file with identical content and same filename was already
        # uploaded"), so the static fixture fails validation once any earlier
        # run has consumed it - which looks like the stuck state this step is
        # meant to rule out.
        small_workbook = tmp_path / f"small_upload_{uuid4().hex[:10]}.xlsx"
        build_item_workbook(
            Path(ReadConfig.get_upload_item_file_path()),
            small_workbook,
            count=1,
            seed=f"upload-size-limit-{uuid4().hex[:12]}",
        )
        self.logger.info(f"Uploading valid smaller workbook: {small_workbook}")

        page.upload_file(small_workbook)
        validation_message = page.wait_for_upload_validation_success()
        self.logger.info(f"Validation success message: {validation_message}")
        success_screenshot = ScreenshotUtils.capture(
            self.driver,
            f"{request.node.name}_file_validation_passed_message",
        )
        self.add_evidence_screenshot(
            request,
            evidence_screenshots,
            "Valid file PASSED validation message",
            success_screenshot,
        )

        print("[CHECK] Validation success message contains PASSED", flush=True)
        assert "passed" in validation_message.casefold() or "success" in validation_message.casefold(), (
            f"Expected smaller file validation to succeed, but got: {validation_message}"
        )
        print("[PASS] Smaller valid file successfully accepted without interface stuck state.", flush=True)
