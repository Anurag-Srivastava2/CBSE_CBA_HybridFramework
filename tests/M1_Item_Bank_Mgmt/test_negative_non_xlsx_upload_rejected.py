from shutil import copy2
import json
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from selenium.webdriver.common.by import By

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from tests.M1_Item_Bank_Mgmt.m1_surveys import survey_chrome, survey_upload_step
from utilities.element_checks import ElementChecks
from utilities.read_config import ReadConfig
from utilities.screenshot_utils import ScreenshotUtils


@pytest.mark.rtm
@pytest.mark.e2e
@pytest.mark.usefixtures("setup")
class TestNegativeNonXlsxUploadRejected:
    """BUG-M1-001 regression: invalid upload formats must reject cleanly."""

    def login_and_open_upload_step(self):
        self.driver.get(ReadConfig.get_base_url())
        sme_username = ReadConfig.get_sme2_username()
        LoginPage(self.driver).login_to_application(
            sme_username,
            ReadConfig.get_password_for_username(sme_username),
        )

        upload_page = UploadItemFilePage(self.driver)
        upload_page.close_popup_if_open()
        upload_page.open_item_creation_module()
        upload_page.open_upload_item_file_tab()
        upload_page.open_upload_step()
        return upload_page

    @staticmethod
    def make_unique_xlsx(source_file_path, target_file_path, header_override=None):
        copy2(source_file_path, target_file_path)
        workbook = load_workbook(target_file_path)
        worksheet = workbook.active

        if header_override:
            header_lookup = {
                str(worksheet.cell(row=1, column=column).value).strip(): column
                for column in range(1, worksheet.max_column + 1)
                if worksheet.cell(row=1, column=column).value
            }
            for old_header, new_header in header_override.items():
                worksheet.cell(row=1, column=header_lookup[old_header]).value = new_header

        run_id = uuid4().hex[:10]
        for row_number in range(2, worksheet.max_row + 1):
            question_cell = worksheet.cell(row=row_number, column=11)
            if question_cell.value:
                question_cell.value = (
                    f"{question_cell.value} BUG-M1-001 recovery validation {run_id}-{row_number}"
                )

        workbook.save(target_file_path)
        workbook.close()
        return target_file_path

    @staticmethod
    def assert_non_xlsx_rejection(message, extension):
        normalized_message = message.casefold()
        assert extension in normalized_message or "xlsx" in normalized_message
        assert any(
            token in normalized_message
            for token in ("invalid", "only", "not supported", "file format", "error", "failed")
        ), f"Expected explicit non-.xlsx rejection. Actual message:\n{message}"

    def assert_continue_button_disabled(self):
        enabled_continue_buttons = self.driver.find_elements(
            By.XPATH,
            "//button[contains(normalize-space(),'Continue') "
            "and not(@disabled) and not(@aria-disabled='true')]",
        )
        visible_enabled = [
            button
            for button in enabled_continue_buttons
            if button.is_displayed() and button.is_enabled()
        ]
        assert not visible_enabled, "Continue button should remain disabled after invalid upload."

    def add_evidence_screenshot(self, request, name):
        screenshot_path = ScreenshotUtils.capture(self.driver, name)
        evidence = []
        for property_name, property_value in request.node.user_properties:
            if property_name == "evidence_screenshots":
                try:
                    evidence = json.loads(property_value)
                except Exception:
                    evidence = []
                break
        evidence.append({"name": name, "path": screenshot_path})
        request.node.user_properties[:] = [
            item for item in request.node.user_properties if item[0] != "evidence_screenshots"
        ]
        request.node.user_properties.append(("evidence_screenshots", json.dumps(evidence)))
        return screenshot_path

    def upload_assert_capture_and_reset(
        self,
        request,
        upload_page,
        file_path,
        screenshot_name,
        assertion,
        timeout=30,
        reset_after=True,
    ):
        upload_page.upload_file(file_path)
        message = assertion(upload_page, timeout)
        self.add_evidence_screenshot(request, screenshot_name)
        if reset_after:
            upload_page.reset_upload_step()
        return message

    def test_bug_m1_001_non_xlsx_and_bad_header_uploads_rejected_then_valid_xlsx_accepted(
        self, tmp_path, request, record_property
    ):
        pdf_file = tmp_path / "invalid_item_upload.pdf"
        pdf_file.write_bytes(b"%PDF-1.4\n% BUG-M1-001 invalid PDF upload regression\n")

        docx_file = tmp_path / "invalid_item_upload.docx"
        docx_file.write_bytes(
            b"PK\x03\x04 BUG-M1-001 invalid DOCX upload regression placeholder\n"
        )

        bad_header_file = self.make_unique_xlsx(
            ReadConfig.get_upload_item_file_path(),
            tmp_path / "invalid_missing_competency_header.xlsx",
            header_override={"Competency": "cpm"},
        )
        valid_file = self.make_unique_xlsx(
            ReadConfig.get_upload_item_file_path(),
            tmp_path / "valid_after_rejections.xlsx",
        )

        upload_page = self.login_and_open_upload_step()

        # Surveyed on arrival, before any invalid file is pushed at the page, so
        # the element table describes the upload step as it should look rather
        # than as a rejected upload left it.
        checks = ElementChecks(
            upload_page, record_property, page_name="Upload Item File — Invalid Uploads"
        )
        survey_chrome(checks, upload_page)
        survey_upload_step(checks, upload_page)
        record_property("result_description", checks.publish())

        def assert_pdf_rejected(page, timeout):
            pdf_error = page.wait_for_upload_rejection(timeout=timeout)
            self.assert_non_xlsx_rejection(pdf_error, "pdf")
            return pdf_error

        def assert_docx_rejected(page, timeout):
            docx_error = page.wait_for_upload_rejection(timeout=timeout)
            self.assert_non_xlsx_rejection(docx_error, "docx")
            return docx_error

        def assert_bad_header_rejected(page, timeout):
            header_error = page.wait_for_upload_rejection(timeout=timeout)
            normalized_header_error = header_error.casefold()
            # The app words this as "missing required column(s)" when one column
            # is missing, so match the singular stem rather than the plural.
            assert "missing required column" in normalized_header_error
            assert "competency" in normalized_header_error
            self.assert_continue_button_disabled()
            return header_error

        def assert_valid_xlsx_accepted(page, timeout):
            return page.wait_for_upload_validation_success()

        pdf_error = self.upload_assert_capture_and_reset(
            request,
            upload_page,
            pdf_file,
            "BUG-M1-001 PDF rejection passed",
            assert_pdf_rejected,
            timeout=30,
        )
        docx_error = self.upload_assert_capture_and_reset(
            request,
            upload_page,
            docx_file,
            "BUG-M1-001 DOCX rejection passed",
            assert_docx_rejected,
            timeout=30,
        )
        header_error = self.upload_assert_capture_and_reset(
            request,
            upload_page,
            bad_header_file,
            "BUG-M1-001 missing Competency header rejection passed",
            assert_bad_header_rejected,
            timeout=60,
        )
        success_message = self.upload_assert_capture_and_reset(
            request,
            upload_page,
            valid_file,
            "BUG-M1-001 valid XLSX accepted after rejections",
            assert_valid_xlsx_accepted,
            timeout=60,
            reset_after=False,
        )
        normalized_success = success_message.casefold()
        # Single-file uploads report "All N rows added successfully (... Status:
        # PASSED)" instead of the multi-file "validated successfully" wording.
        assert (
            "validated successfully" in normalized_success
            or "added successfully" in normalized_success
        ), f"Valid XLSX was not accepted after the rejections: {success_message!r}"
        assert "passed" in normalized_success
        request.node.user_properties.append(("upload_pdf_rejection", pdf_error))
        request.node.user_properties.append(("upload_docx_rejection", docx_error))
        request.node.user_properties.append(("upload_bad_header_rejection", header_error))
        request.node.user_properties.append(("upload_success_message", success_message))
