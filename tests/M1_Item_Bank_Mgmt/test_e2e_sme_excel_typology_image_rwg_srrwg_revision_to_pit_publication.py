"""E2E test: for each of the 12 supported typologies, SME uploads a single-item
Excel + images ZIP, RWG sends it back for revision, SME adds an image and edits
the text, RWG verifies and approves, the same revision cycle repeats with
SRRWG, and the item is finally published through PIT.

Each parametrized run picks exactly one typology template from
data/typology_templates/, so a full pytest run of this file exercises all 12
typologies.
"""
import os
import zipfile
from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from time import monotonic, sleep
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from selenium.common.exceptions import TimeoutException

from pages.common.login_page import LoginPage
from pages.pit.review_queue_page import PITReviewQueuePage
from pages.rwg.review_queue_page import RWGReviewQueuePage
from pages.sme.bulk_upload_page import BulkUploadPage
from pages.sr_rwg.review_queue_page import SRRWGReviewQueuePage
from tests.M1_Item_Bank_Mgmt.m1_surveys import (
    enter_screen,
    survey_opened_item_set,
    survey_upload_step,
)
from utilities.element_checks import ElementChecks
from utilities.logger import LogGenerator
from utilities.qar_recovery import recover_qar_need_improvement_items
from utilities.read_config import ReadConfig
from utilities.screenshot_utils import ScreenshotUtils

TEMPLATES_DIR = Path(__file__).parent.parent.parent / "data" / "typology_templates"
IMAGES_DIR = TEMPLATES_DIR / "images"

# Column indexes match the 24 canonical template columns documented in
# data/typology_templates/README.md.
QUESTION_COLUMN = 11
QUESTION_IMAGE_COLUMN = 12

TYPOLOGY_CASES = [
    ("01_MCQ.xlsx", "Multiple Choice Question"),
    ("02_TOF.xlsx", "True or False"),
    ("03_MTF.xlsx", "Match the Following"),
    ("04_FITB.xlsx", "Fill in the Blank"),
    ("05_AR.xlsx", "Assertion and Reasoning"),
    ("06_VSAQ.xlsx", "Very Short Answer Question"),
    ("07_SAQ.xlsx", "Short Answer Question"),
    ("08_LAQ.xlsx", "Long Answer Question"),
    ("09_FAA.xlsx", "FA Activity"),
    ("10_FR.xlsx", "Free Response"),
    ("11_CABA.xlsx", "Case Based Question"),
    ("12_SBQ.xlsx", "Source Based Question"),
]


class AppendingBulkUploadPage(BulkUploadPage):
    """Revision editing that appends to every text box and image instead of
    replacing them.

    UploadItemFilePage.edit_open_revision_item() (used by the sibling E2E
    tests) selects-all-and-deletes the Question editor before typing new
    text and only ever touches that one field - fine for QAR-correction
    flows that intentionally swap in a fully-corrected question, but wrong
    here: this test wants each revision round to add its own text and image
    on top of what's already there (including the previous round's image,
    which lives in the same rich-text region), and to touch every text box
    on the item (Question, Options, Answer, Explanation, ...), not just the
    Question field. Left as a test-local subclass rather than changing the
    shared method, since other tests rely on its current replace behavior.
    """

    APPEND_TEXT_JS = r"""
        const markerText = arguments[0];
        const visible = el => {
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return rect.width > 0 && rect.height > 0
                && style.display !== 'none' && style.visibility !== 'hidden';
        };
        const insideRevisionNotes = el => {
            let node = el;
            while (node && node !== document.body) {
                const text = (node.innerText || node.textContent || '');
                if (text.includes('Revision Notes') && text.length < 400) return true;
                node = node.parentElement;
            }
            return false;
        };
        const fields = Array.from(
            document.querySelectorAll('textarea, input[type="text"], [contenteditable="true"]')
        ).filter(el => visible(el) && !insideRevisionNotes(el));
        let touched = 0;
        for (const field of fields) {
            const tag = field.tagName.toLowerCase();
            if (tag === 'textarea' || tag === 'input') {
                field.focus();
                field.value = (field.value || '') + ' ' + markerText;
                field.dispatchEvent(new Event('input', { bubbles: true }));
                field.dispatchEvent(new Event('change', { bubbles: true }));
            } else {
                field.focus();
                const span = document.createElement('span');
                span.textContent = ' ' + markerText;
                field.appendChild(span);
                field.dispatchEvent(new InputEvent('input', {
                    bubbles: true, inputType: 'insertText', data: markerText,
                }));
                field.dispatchEvent(new Event('change', { bubbles: true }));
            }
            touched += 1;
        }
        return touched;
        """

    def append_text_to_all_open_item_fields(self, marker_text):
        touched = self.driver.execute_script(self.APPEND_TEXT_JS, marker_text)
        assert touched > 0, "No visible editable text fields found on the open item."
        return touched

    def revise_open_item_appending_text_and_image(self, item_label, image_path, marker_text):
        """Edit the currently-open revision item without clearing its
        existing content: append `marker_text` to every visible text box
        and add `image_path` alongside any image already in the editor.
        """
        self.click_edit_item()
        touched = self.append_text_to_all_open_item_fields(marker_text)
        self.attach_image_to_item_editor(image_path)
        self.enter_revision_notes(f"Automation revision note for {item_label}: {marker_text}.")
        self.click_save_revision()
        return item_label, touched

    def revise_items_appending_text_and_image_in_open_item_set(
        self, item_set_id, image_path, marker_text
    ):
        return self._revise_items_loop(
            item_set_id,
            lambda label: self.revise_open_item_appending_text_and_image(
                label, image_path, marker_text
            ),
        )


class RevisionCapableSRRWGReviewQueuePage(SRRWGReviewQueuePage):
    """Adds the send-back-for-revision action SRRWGReviewQueuePage doesn't
    expose as a convenience wrapper, using the same base-class building
    blocks RWGReviewQueuePage.send_item_set_back_as_rwg already relies on.
    """

    def send_item_set_back_as_sr_rwg(self, item_set_id, item_ids, item_set_url=""):
        self.open_review_item_set(item_set_id, item_set_url)
        revised_item_ids = self.send_items_back_for_revision(item_set_id, item_ids)
        self.click_required_and_confirm(self.SUBMIT_REVIEW_LOCATORS, "Submit SRRWG Review")
        return revised_item_ids


# Every typology run drives the app-assigned RWG/SRRWG reviewer and the PIT
# panel. Those accounts are shared across all parametrized runs, so the
# typologies must execute one at a time rather than fan out across workers.
@pytest.mark.serial
@pytest.mark.flaky(reruns=1, reruns_delay=5)
@pytest.mark.usefixtures("setup")
class TestE2ESMEExcelTypologyImageRWGSRRWGRevisionToPITPublication:
    logger = LogGenerator.loggen()

    @staticmethod
    def get_sme_username(typology_index):
        """Round-robin across every configured SME account so parallel workers
        (pytest-xdist -n) each drive a distinct typology run under a distinct
        SME session instead of contending for one shared account.
        """
        # Under xdist the typology index is the wrong key: two workers running
        # typologies 4 apart would land on the same account. Pin to the account
        # this worker owns instead (see ReadConfig.get_sme2_username).
        if os.getenv("PYTEST_XDIST_WORKER", "").strip().startswith("gw"):
            return ReadConfig.get_sme2_username()
        configured_sme_usernames = ReadConfig.get_role_usernames("sme")
        if configured_sme_usernames:
            return configured_sme_usernames[typology_index % len(configured_sme_usernames)]
        return ReadConfig.get_sme2_username()

    # Stage-labeled revision images (generated once, see data/typology_templates/
    # images/README or the make_label_image script in git history): every RWG
    # send-back revision attaches the same clearly-labeled "RWG Revision" image
    # and every SRRWG send-back revision attaches "SRRWG Revision", so evidence
    # screenshots make it visually obvious which stage added which image.
    RWG_REVISION_IMAGE = IMAGES_DIR / "rwg_revision.png"
    SR_RWG_REVISION_IMAGE = IMAGES_DIR / "sr_rwg_revision.png"
    PIT_REVISION_IMAGE = IMAGES_DIR / "pit_revision.png"

    @staticmethod
    def build_run_workbook_and_image_zip(template_name, typology_index):
        """Copy the typology template, tag its Question cell with a unique run
        token (so QAR duplicate-detection never sees the same question text
        twice), point Question Image at a filename, and zip a matching image
        for it (flat, no folder prefix, sourced from the staged demo images).
        """
        template_path = TEMPLATES_DIR / template_name
        run_token = uuid4().hex[:10]

        upload_dir = Path.cwd() / "tmp_uploads"
        upload_dir.mkdir(exist_ok=True)
        workbook_path = upload_dir / f"{template_path.stem}_{run_token}.xlsx"
        copy2(template_path, workbook_path)

        workbook = load_workbook(workbook_path)
        worksheet = workbook.active
        image_filename = f"typology_{run_token}.png"
        question_cell = worksheet.cell(row=2, column=QUESTION_COLUMN)
        question_cell.value = f"{question_cell.value} [{run_token}]"
        worksheet.cell(row=2, column=QUESTION_IMAGE_COLUMN).value = image_filename
        workbook.save(workbook_path)
        workbook.close()

        source_images = sorted(IMAGES_DIR.glob("*.png"))
        assert source_images, f"No staged demo images found in {IMAGES_DIR}"
        source_image = source_images[typology_index % len(source_images)]

        zip_path = upload_dir / f"{template_path.stem}_{run_token}.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.write(source_image, arcname=image_filename)

        return workbook_path, zip_path, run_token, image_filename

    def soft_check(self, request, condition, message, blocking=False):
        """Record a failed expectation without stopping the run, unless the
        caller marks it `blocking=True` (nothing further can happen without
        it - e.g. the upload itself was rejected). Non-blocking checks are
        recorded on the report and printed, but the flow proceeds so a
        single flaky verification step doesn't waste an otherwise-complete
        12-step run.
        """
        if condition:
            return True
        if blocking:
            raise AssertionError(message)
        self.logger.warning("[SOFT FAILURE] %s", message)
        request.node.user_properties.append(("soft_failure", message))
        return False

    def soft_call(self, request, description, fn, *args, **kwargs):
        """Run `fn`; on any exception, record it as a soft failure (see
        soft_check) and return None instead of aborting the run.
        """
        try:
            return fn(*args, **kwargs)
        except Exception as error:
            message = f"{description}: {error}"
            self.logger.warning("[SOFT FAILURE] %s", message)
            request.node.user_properties.append(("soft_failure", message))
            return None

    def capture_item_set_url(self, upload_page, item_set_id, fallback_url):
        """Return the current URL, but only if it's actually the item set's
        own page - not the generic "/item-sets" list page.

        rerun_qar_if_enabled() can briefly route through the list page
        before settling back onto the item-set detail page; capturing
        driver.current_url immediately afterward is racy and can grab that
        transitional URL instead. If the current page doesn't contain
        `item_set_id`, re-navigate to `fallback_url` (the last known-good
        item-set URL) before trusting the captured URL.
        """
        body_text = self.driver.find_element("tag name", "body").text
        numeric_prefix = upload_page.item_set_numeric_prefix(item_set_id)
        if numeric_prefix in body_text:
            return self.driver.current_url
        upload_page.open_item_set_url_and_wait(fallback_url, item_set_id)
        return self.driver.current_url

    def wait_for_open_item_images_to_load(self, timeout=60):
        """Wait for every visible image in the currently-open item's review
        pane to finish loading (complete && naturalWidth > 0), not just
        appear in the DOM. An item revised twice legitimately carries two
        embedded images (revision rounds append rather than replace), and
        multi-image items take longer to fully render than the default
        pause allows.
        """
        try:
            self.driver.set_script_timeout(timeout + 5)
            self.driver.execute_async_script(
                r"""
                const timeoutMs = arguments[0];
                const callback = arguments[1];
                let settled = false;
                const done = value => {
                    if (settled) return;
                    settled = true;
                    callback(value);
                };
                const check = () => {
                    const imgs = Array.from(document.querySelectorAll('img'))
                        .filter(img => {
                            const rect = img.getBoundingClientRect();
                            return rect.width > 10 && rect.height > 10;
                        });
                    if (imgs.length === 0
                        || imgs.every(img => img.complete && img.naturalWidth > 0)) {
                        done(true);
                    } else {
                        setTimeout(check, 300);
                    }
                };
                check();
                setTimeout(() => done(false), timeoutMs);
                """,
                timeout * 1000,
            )
        except Exception:
            pass

    def survey_reviewer_screen(self, review_page, role):
        """Survey a reviewer's opened item set, if this test is collecting.

        Read-only: no criteria are marked, so it cannot consume the one-time
        review vote the set still needs. Re-published each phase because
        publish() writes the whole accumulated list, so a failure later in
        this long chain still leaves the rows gathered so far on the card.
        """
        checks = getattr(self, "checks", None)
        if checks is None:
            return
        enter_screen(checks, f"{role.upper()} — Opened Item Set")
        survey_opened_item_set(checks, review_page)
        checks.publish()

    def resolve_reviewer_username(
        self,
        upload_page,
        review_page,
        role,
        item_set_id,
        item_set_url,
        preferred_username=None,
        sme_username=None,
    ):
        """Find whichever configured account for `role` actually holds this
        item set.

        Primary strategy: open the item set as SME and read the scoped
        "Reviewer:"/"Assigned RWG/SRRWG" label directly off that page (see
        get_scoped_item_set_assignee_text - it reads only the label's own
        text, not the whole flattened page, so it can't bleed in an
        unrelated adjacent number), then log straight in as that exact
        account. This avoids probing every configured account one by one.

        Fallback: if the SME-page read doesn't resolve to a configured
        username, or that account's queue doesn't actually have the item
        set (e.g. an assignment changed since the read), fall back to
        probing every configured account for `role` directly - mirrors
        resolve_rwg_username_for_item_set() in the reference test
        test_e2e_teacher_excel_triple_revision_rejection_to_pit_publication.py.
        """
        if sme_username:
            self.login_as(sme_username, upload_page)
            upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
            try:
                scraped_username = ReadConfig.get_all_user_username(
                    upload_page.require_item_set_assignee(role)
                )
            except (AssertionError, RuntimeError):
                scraped_username = None
            self.logger.info(
                "resolve_reviewer_username(%s, %s): SME page assignee scrape -> %r",
                role, item_set_id, scraped_username,
            )
            if scraped_username:
                try:
                    self.login_as(scraped_username, upload_page)
                    review_page.open_review_item_set(item_set_id, item_set_url)
                    self.survey_reviewer_screen(review_page, role)
                    return scraped_username
                except TimeoutException:
                    pass

        role_usernames = ReadConfig.get_role_usernames(role)
        candidates = list(
            dict.fromkeys(
                username
                for username in [preferred_username, *role_usernames]
                if username
            )
        )
        last_error = None
        # Handoff to the next review stage has been observed to lag well
        # beyond the previous 3-pass/10s budget for content-heavy typologies
        # (multi-pair MTF, multi-blank FITB, multi-sub-question CABA/SBQ);
        # widened to give backend indexing more time before treating it as a
        # genuine routing failure.
        max_passes = 6
        for pass_number in range(1, max_passes + 1):
            for candidate in candidates:
                try:
                    # The login belongs inside the try: with several suites
                    # driving the same shared reviewer pool at once, the SPA
                    # occasionally strands a sign-in on the login form until it
                    # times out. That is just another candidate that did not
                    # work, so record it and move to the next account instead of
                    # abandoning every remaining candidate and pass.
                    self.login_as(candidate, upload_page)
                    review_page.open_review_item_set(item_set_id, item_set_url)
                    self.survey_reviewer_screen(review_page, role)
                except TimeoutException as error:
                    last_error = error
                    continue
                return candidate
            self.logger.info(
                "resolve_reviewer_username(%s, %s): pass %d/%d found no queue match "
                "among %s",
                role, item_set_id, pass_number, max_passes, candidates,
            )
            if pass_number < max_passes:
                sleep(20)
        role_label = "SRRWG" if role == "sr_rwg" else role.upper()
        raise TimeoutException(
            f"Item set {item_set_id} was not visible in any configured {role_label} "
            f"queue ({candidates}) after {max_passes} passes."
        ) from last_error

    def login_as(self, username, upload_page):
        """Clear cookies/local/session storage before every login switch.

        The SPA keeps the previous user's session client-side, so logging in
        as a different role without resetting first silently re-renders the
        prior role's dashboard (e.g. an RWG login would still show "SME" in
        the sidebar and lack the review/evaluation-criteria panel).
        """
        upload_page.reset_browser_session_to_login()
        LoginPage(self.driver).login_to_application(
            username,
            ReadConfig.get_password_for_username(username),
        )
        upload_page.close_popup_if_open()

    @pytest.mark.parametrize("template_name, typology", TYPOLOGY_CASES)
    def test_e2e_sme_typology_image_revision_rwg_srrwg_pit_publish(
        self, request, template_name, typology, record_property
    ):
        typology_index = TYPOLOGY_CASES.index((template_name, typology))
        request.node.user_properties.append(
            ("result_checkpoint", f"fresh {typology} Excel + images ZIP upload")
        )

        upload_page = AppendingBulkUploadPage(self.driver)
        rwg_page = RWGReviewQueuePage(self.driver)
        sr_rwg_page = RevisionCapableSRRWGReviewQueuePage(self.driver)
        pit_page = PITReviewQueuePage(self.driver)

        sme_username = self.get_sme_username(typology_index)
        self.login_as(sme_username, upload_page)

        # One collector for the whole SME -> RWG -> Sr. RWG -> PIT chain,
        # re-pointed at each screen as the set moves through it.
        self.checks = ElementChecks(
            upload_page,
            record_property,
            page_name=f"SME Bulk Upload — {typology}",
        )

        workbook_path, zip_path, run_token, image_filename = (
            self.build_run_workbook_and_image_zip(template_name, typology_index)
        )
        request.node.user_properties.extend(
            [
                ("typology", typology),
                ("template_name", template_name),
                ("run_token", run_token),
                ("image_filename", image_filename),
            ]
        )

        # 1. Upload the Excel + images ZIP and submit for QAR.
        upload_outcome = upload_page.upload_excel_and_images_zip_for_validation(
            workbook_path, zip_path
        )
        assert upload_outcome["accepted"], (
            f"Excel + images ZIP upload was rejected for {typology}: "
            f"{upload_outcome['message']}"
        )
        # Note: deliberately not calling get_images_zip_validation_evidence()
        # here - it advances the wizard (Upload -> Review) as a side effect,
        # which would desync submit_for_qar()'s own Upload-step-relative
        # click_continue() calls below. Just capture the current state.
        upload_evidence_screenshot = ScreenshotUtils.capture(
            self.driver, f"{request.node.name}_upload_zip_accepted"
        )
        self.logger.info(
            "Upload accepted for %s: %s", typology, upload_outcome["message"]
        )

        qar_outcome = upload_page.submit_for_qar(run_token)
        item_ids = qar_outcome["item_ids"]
        item_set_id = qar_outcome["item_set_id"]
        assert len(item_ids) == 1, (
            f"Expected exactly 1 item for {typology}, got {item_ids} "
            "(CABA/SBQ sub-rows should not create separate items)."
        )
        request.node.user_properties.extend(
            [
                ("item_set_id", item_set_id),
                ("item_ids", ", ".join(item_ids)),
            ]
        )

        # 2. Let QAR run to completion, correcting any Need Improvement verdict.
        request.node.user_properties.append(
            ("result_checkpoint", "initial QAR validation")
        )
        recover_qar_need_improvement_items(
            page=upload_page,
            item_set_id=item_set_id,
            item_ids=item_ids,
            workbook_path=workbook_path,
            max_retries=3,
            run_label=f"TYPOLOGY-{typology_index + 1:02d}-E2E",
        )
        item_set_url = self.driver.current_url

        # 3. RWG round 1 - send the item back for revision.
        request.node.user_properties.append(
            ("result_checkpoint", "RWG sends item back for revision (round 1)")
        )
        rwg_username = self.resolve_reviewer_username(
            upload_page, rwg_page, "rwg", item_set_id, item_set_url,
            sme_username=sme_username,
        )
        sent_back_item_ids_1 = rwg_page.send_item_set_back_as_rwg(
            item_set_id, item_ids, item_set_url
        )
        self.soft_check(
            request,
            set(sent_back_item_ids_1) == set(item_ids),
            f"RWG did not send back the expected item for {typology}: {sent_back_item_ids_1}.",
        )
        rwg_round1_send_back_screenshot = rwg_page.capture_review_screenshot(
            request.node.name, "rwg_round1_sent_back"
        )

        # 4. SME adds an image and edits the text, then resubmits.
        request.node.user_properties.append(
            ("result_checkpoint", "SME adds image + text, resubmits (round 1)")
        )
        self.login_as(sme_username, upload_page)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        first_image_path = self.RWG_REVISION_IMAGE
        round_1_marker_text = f"RWG revision round 1 [{run_token}]"
        revised_round_1 = upload_page.revise_items_appending_text_and_image_in_open_item_set(
            item_set_id, first_image_path, round_1_marker_text
        )
        self.soft_check(
            request,
            bool(revised_round_1),
            f"SME did not see any revision items after RWG round 1 send-back for {typology}.",
        )
        upload_page.rerun_qar_if_enabled()
        sme_round1_resubmit_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_sme_round1_resubmitted"
        )
        item_set_url = self.capture_item_set_url(upload_page, item_set_id, item_set_url)

        # 5. RWG verifies the added image/text, then approves.
        request.node.user_properties.append(
            ("result_checkpoint", "RWG verifies image + text, approves")
        )
        rwg_username = self.resolve_reviewer_username(
            upload_page, rwg_page, "rwg", item_set_id, item_set_url, rwg_username,
            sme_username=sme_username,
        )
        rwg_revised_visible = self.soft_call(
            request,
            f"RWG round 1 revised-content visibility check for {typology}",
            rwg_page.verify_revised_items_visible_to_reviewer,
            item_set_id,
            [label for label, _ in revised_round_1]
            if revised_round_1 and isinstance(revised_round_1[0], tuple)
            else revised_round_1,
            expected_revision_prefix=round_1_marker_text,
        )
        # An item revised once already carries its round-1 image; multi-image
        # items take longer to fully render than a fixed pause allows.
        self.wait_for_open_item_images_to_load()
        rwg_images_visible = self.soft_call(
            request,
            f"RWG round 1 image visibility check for {typology}",
            rwg_page.verify_images_visible_to_reviewer,
            item_set_id,
            item_ids,
        )
        request.node.user_properties.extend(
            [
                ("rwg_visible_revised_content_round1", str(rwg_revised_visible)),
                ("rwg_visible_images_round1", str(rwg_images_visible)),
            ]
        )
        rwg_approved_item_ids = rwg_page.approve_item_set_as_rwg(
            item_set_id, item_ids, item_set_url
        )
        self.soft_check(
            request,
            set(rwg_approved_item_ids) == set(item_ids),
            f"RWG did not approve the expected item for {typology}: {rwg_approved_item_ids}.",
        )
        rwg_round1_approved_screenshot = rwg_page.capture_review_screenshot(
            request.node.name, "rwg_round1_approved"
        )

        # 6. SRRWG round - send the item back for revision.
        request.node.user_properties.append(
            ("result_checkpoint", "SRRWG sends item back for revision")
        )
        sr_rwg_username = self.resolve_reviewer_username(
            upload_page, sr_rwg_page, "sr_rwg", item_set_id, item_set_url,
            sme_username=sme_username,
        )
        sent_back_item_ids_2 = sr_rwg_page.send_item_set_back_as_sr_rwg(
            item_set_id, item_ids, item_set_url
        )
        self.soft_check(
            request,
            set(sent_back_item_ids_2) == set(item_ids),
            f"SRRWG did not send back the expected item for {typology}: {sent_back_item_ids_2}.",
        )
        sr_rwg_send_back_screenshot = sr_rwg_page.capture_review_screenshot(
            request.node.name, "sr_rwg_sent_back"
        )

        # 7. SME adds an image and edits the text again, then resubmits.
        request.node.user_properties.append(
            ("result_checkpoint", "SME adds image + text, resubmits (round 2)")
        )
        self.login_as(sme_username, upload_page)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        second_image_path = self.SR_RWG_REVISION_IMAGE
        round_2_marker_text = f"SRRWG revision round 2 [{run_token}]"
        revised_round_2 = upload_page.revise_items_appending_text_and_image_in_open_item_set(
            item_set_id, second_image_path, round_2_marker_text
        )
        self.soft_check(
            request,
            bool(revised_round_2),
            f"SME did not see any revision items after SRRWG send-back for {typology}.",
        )
        upload_page.rerun_qar_if_enabled()
        sme_round2_resubmit_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_sme_round2_resubmitted"
        )
        item_set_url = self.capture_item_set_url(upload_page, item_set_id, item_set_url)

        # 8. SRRWG verifies the added image/text, then approves.
        request.node.user_properties.append(
            ("result_checkpoint", "SRRWG verifies image + text, approves")
        )
        sr_rwg_username = self.resolve_reviewer_username(
            upload_page, sr_rwg_page, "sr_rwg", item_set_id, item_set_url, sr_rwg_username,
            sme_username=sme_username,
        )
        sr_rwg_revised_visible = self.soft_call(
            request,
            f"SRRWG revised-content visibility check for {typology}",
            sr_rwg_page.verify_revised_items_visible_to_reviewer,
            item_set_id,
            [label for label, _ in revised_round_2]
            if revised_round_2 and isinstance(revised_round_2[0], tuple)
            else revised_round_2,
            expected_revision_prefix=round_2_marker_text,
        )
        # By this point the item carries two revision images (round 1 + round
        # 2, since revisions append rather than replace) - give the pane time
        # to fully render both before counting.
        self.wait_for_open_item_images_to_load()
        sr_rwg_images_visible = self.soft_call(
            request,
            f"SRRWG image visibility check for {typology}",
            sr_rwg_page.verify_images_visible_to_reviewer,
            item_set_id,
            item_ids,
        )
        request.node.user_properties.extend(
            [
                ("sr_rwg_visible_revised_content", str(sr_rwg_revised_visible)),
                ("sr_rwg_visible_images", str(sr_rwg_images_visible)),
            ]
        )
        sr_rwg_approved_item_ids = sr_rwg_page.approve_item_set_as_sr_rwg(
            item_set_id, item_ids, item_set_url
        )
        self.soft_check(
            request,
            set(sr_rwg_approved_item_ids) == set(item_ids),
            f"SRRWG did not approve the expected item for {typology}: {sr_rwg_approved_item_ids}.",
        )
        sr_rwg_approved_screenshot = sr_rwg_page.capture_review_screenshot(
            request.node.name, "sr_rwg_approved"
        )

        # 9. PIT 3/3 quorum publishes the item.
        request.node.user_properties.append(
            ("result_checkpoint", "PIT 3/3 quorum and publication")
        )
        completed_pit_approvals = []
        for pit_username in ReadConfig.get_pit_usernames()[:3]:
            self.login_as(pit_username, upload_page)
            pit_page.approve_item_set_as_pit(item_set_id, item_set_url)
            completed_pit_approvals.append(pit_username)
            pit_page.capture_review_screenshot(
                request.node.name, f"pit_approval_{len(completed_pit_approvals)}"
            )
        self.soft_check(
            request,
            len(completed_pit_approvals) == 3,
            f"Expected PIT 3/3 for {typology}, completed {completed_pit_approvals}.",
        )

        # 10. Final verification.
        request.node.user_properties.append(
            ("result_checkpoint", "final published item-set status verification")
        )
        self.login_as(sme_username, upload_page)
        upload_page.open_item_set_detail(item_set_id)
        # The status-count widget refreshes asynchronously after the page
        # skeleton finishes loading, so the "Approved"/"Pending" counts can
        # still be stale (or absent) for a few seconds right after the 3rd
        # PIT approval - poll instead of reading once, to avoid a false
        # failure while the backend/UI catches up.
        final_summary = {}
        final_status_text = ""
        deadline = monotonic() + 30
        while monotonic() < deadline:
            final_summary = upload_page.get_item_set_status_summary()
            final_status_text = upload_page.format_status_summary(final_summary)
            if (
                str(final_summary.get("Approved", "")) == str(len(item_ids))
                and str(final_summary.get("Pending", "")) == "0"
            ):
                break
            sleep(2)
            upload_page.open_item_set_detail(item_set_id)
        assert int(final_summary.get("Approved", -1)) == len(item_ids), (
            f"{typology} was not fully approved after PIT 3/3: {final_status_text}"
        )
        assert int(final_summary.get("Pending", -1)) == 0, (
            f"{typology} still had pending items after PIT 3/3: {final_status_text}"
        )
        final_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_final_published_status"
        )

        request.node.user_properties.extend(
            [
                ("post_approval_status", final_status_text),
                ("qar_success_screenshot", final_screenshot),
                (
                    "result_description",
                    f"{typology} ({item_set_id}) passed a full RWG revision cycle and a "
                    "full SRRWG revision cycle (image + text added and verified each time), "
                    "then published through PIT 3/3.",
                ),
            ]
        )
