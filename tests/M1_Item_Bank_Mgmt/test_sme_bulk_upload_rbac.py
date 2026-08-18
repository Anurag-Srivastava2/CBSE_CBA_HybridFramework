from openpyxl import load_workbook
import pytest

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from tests.M1_Item_Bank_Mgmt.m1_surveys import survey_chrome, survey_item_sets
from utilities.element_checks import ElementChecks
from utilities.read_config import ReadConfig


@pytest.mark.rtm
@pytest.mark.usefixtures("setup")
class TestSMEBulkUploadRBAC:
    def test_tc_ibmm_01a_p03_sme_sees_only_assigned_grade_subject_items(
        self, record_property
    ):
        """An SME's item-set listing shows only their assigned grade and subject.

        The listing furniture is surveyed softly; the scope itself is a security
        contract and stays a hard assert — including that rows were actually
        rendered, since "no out-of-scope rows" is trivially true of an empty grid.
        """
        workbook = load_workbook(
            ReadConfig.get_upload_item_file_path(),
            read_only=True,
            data_only=True,
        )
        worksheet = workbook.active
        expected_grade = str(worksheet.cell(row=2, column=1).value).strip()
        expected_subject = str(worksheet.cell(row=2, column=2).value).strip()
        workbook.close()

        self.driver.get(ReadConfig.get_base_url())
        sme_username = ReadConfig.get_sme2_username()
        LoginPage(self.driver).login_to_application(
            sme_username,
            ReadConfig.get_password_for_username(sme_username),
        )

        sets_page = UploadItemFilePage(self.driver)
        sets_page.close_popup_if_open()
        sets_page.open_item_sets_list()

        checks = ElementChecks(
            sets_page, record_property, page_name="My Item Set — RBAC Scope"
        )
        survey_chrome(checks, sets_page)
        survey_item_sets(checks, sets_page)

        # Review-stage tabs re-scope the grid in place; driving them here is
        # read-only and leaves the listing on All for the assertions below.
        for label in ("QAR", "RWG", "Published"):
            checks.check_interaction(
                f"Tab responds — {label}",
                lambda tab=label: sets_page.switch_item_set_tab(tab),
                lambda tab=label: sets_page.is_item_set_tab_active(tab),
            )
        checks.safe_call(lambda: sets_page.switch_item_set_tab("All"))
        record_property("result_description", checks.publish())

        sets_page.open_sets_module()
        scopes = sets_page.verify_visible_item_sets_within_scope(
            expected_grade,
            expected_subject,
        )

        assert all(scope["grade"] != "Grade 10" for scope in scopes)
        assert all(scope["subject"].casefold() != "english" for scope in scopes)
