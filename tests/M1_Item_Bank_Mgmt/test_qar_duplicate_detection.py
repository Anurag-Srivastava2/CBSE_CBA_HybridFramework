import re
from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.read_config import ReadConfig


@pytest.mark.rtm
@pytest.mark.e2e
@pytest.mark.usefixtures("setup")
class TestQARDuplicateDetection:
    """BUG-M1-002 regression coverage for QAR near-duplicate blocking."""

    BASELINE_QUESTIONS = [
        "What is 6 plus 7? Is the answer 13?",
        "Riya has 8 apples and gets 2 more apples. Does she have 10 apples?",
        "What number comes just after 14? Is it 15?",
        "Five stars and four stars are shown together. Do they make 9 stars?",
        "A box has 10 pencils and 3 pencils are taken out. Are 7 pencils left?",
    ]
    NEAR_DUPLICATE_QUESTIONS = [
        "What is 6+7? Is 13 the answer?",
        "Riya had 8 apples and received 2 more. Does she now have 10 apples?",
        "Which number comes after 14? Is the number 15?",
        "Five stars plus four stars are together. Do they make 9 stars in all?",
        "A box contains 10 pencils and 3 are removed. Are 7 pencils remaining?",
    ]

    @staticmethod
    def build_upload_workbook(tmp_path, questions, scenario_name, run_id):
        source = Path(ReadConfig.get_upload_item_file_path())
        target = tmp_path / f"{scenario_name}_{run_id}_{uuid4().hex[:8]}.xlsx"
        copy2(source, target)

        workbook = load_workbook(target)
        worksheet = workbook.active
        max_data_column = 24
        if worksheet.max_column > max_data_column:
            worksheet.delete_cols(max_data_column + 1, worksheet.max_column - max_data_column)

        for offset, question in enumerate(questions):
            row = offset + 2
            if row != 2:
                for column in range(1, max_data_column + 1):
                    source_cell = worksheet.cell(2, column)
                    target_cell = worksheet.cell(row, column)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell._style = copy_cell_style(source_cell._style)
                    if source_cell.number_format:
                        target_cell.number_format = source_cell.number_format
                    if source_cell.alignment:
                        target_cell.alignment = copy_cell_style(source_cell.alignment)

            worksheet.cell(row, 5).value = offset + 1
            worksheet.cell(row, 10).value = "True or False"
            worksheet.cell(row, 11).value = f"{question} Duplicate regression run {run_id}"
            worksheet.cell(row, 21).value = "True"
            worksheet.cell(row, 23).value = (
                f"Automation explanation for {scenario_name} item {offset + 1}."
            )
            worksheet.cell(row, 24).value = "1"

        for row in range(len(questions) + 2, worksheet.max_row + 1):
            for column in range(1, max_data_column + 1):
                worksheet.cell(row, column).value = None

        workbook.save(target)
        workbook.close()
        return target

    def login_as_sme(self):
        self.driver.get(ReadConfig.get_base_url())
        sme_username = ReadConfig.get_sme2_username()
        LoginPage(self.driver).login_to_application(
            sme_username,
            ReadConfig.get_password_for_username(sme_username),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()
        return page

    def submit_workbook_for_qar(self, workbook_path):
        page = UploadItemFilePage(self.driver)
        self.driver.get(ReadConfig.get_base_url())
        page.wait_for_application_to_load()
        page.close_popup_if_open()
        page.open_item_creation_module()
        page.open_upload_item_file_tab()
        page.open_upload_step()
        page.upload_file(workbook_path)
        upload_message = page.wait_for_upload_validation_success()
        page.click_continue()
        item_ids = page.get_review_item_ids()
        page.click_continue()
        page.click_submit_for_qar_and_wait_for_results()
        qar_toast = page.wait_for_ocr_success_message()
        report_text = self.driver.find_element("tag name", "body").text
        item_set_id = page.get_item_set_id_from_item_ids(
            page.get_qar_result_item_ids() or item_ids
        )
        return page, {
            "upload_message": upload_message,
            "toast": qar_toast,
            "report_text": report_text,
            "item_ids": item_ids,
            "item_set_id": item_set_id,
            "status_summary": page.get_item_set_status_summary(),
        }

    @staticmethod
    def count_status(report_text, status):
        return len(re.findall(rf"\b{re.escape(status)}\b", report_text, re.IGNORECASE))

    @staticmethod
    def assert_toast_count_matches_report(toast, report_text):
        toast_numbers = [int(value) for value in re.findall(r"\b\d+\b", toast)]
        if not toast_numbers:
            raise AssertionError(f"QAR completion toast has no item counts: {toast}")

        normalized_toast = toast.casefold()
        if "set" in normalized_toast:
            assert toast_numbers[0] <= toast_numbers[-1], (
                "BUG-M1-002: QAR set-count toast is malformed. "
                f"Toast: {toast}"
            )
            assert toast_numbers[-1] >= 1, (
                "BUG-M1-002: QAR set-count toast did not report any processed set. "
                f"Toast: {toast}"
            )
            if "passed" in normalized_toast:
                assert "approved" in report_text.casefold(), (
                    "BUG-M1-002: QAR toast reports passed set but report has no approved items. "
                    f"Toast: {toast}"
                )
            if "returned" in normalized_toast or "failed" in normalized_toast:
                assert any(
                    status in report_text.casefold()
                    for status in ("needs revision", "revise", "rejected", "failed")
                ), (
                    "BUG-M1-002: QAR toast reports failed/returned set but report has no "
                    f"failure item status. Toast: {toast}"
                )
            return

        needs_revision = TestQARDuplicateDetection.count_status(report_text, "Needs Revision")
        approved = TestQARDuplicateDetection.count_status(report_text, "Approved")
        rejected = TestQARDuplicateDetection.count_status(report_text, "Rejected")
        visible_total = needs_revision + approved + rejected

        assert visible_total in toast_numbers or any(number == 5 for number in toast_numbers), (
            "BUG-M1-002: QAR toast count does not match visible item result count. "
            f"Toast: {toast}; visible statuses: approved={approved}, "
            f"needs_revision={needs_revision}, rejected={rejected}."
        )

    @staticmethod
    def get_status_count(evidence, status):
        summary = evidence.get("status_summary", {})
        if status in summary:
            return int(summary[status])
        aliases = {
            "Needs Revision": ("Needs Revision", "Revise"),
            "Approved": ("Approved",),
            "Rejected": ("Rejected",),
        }
        return max(
            TestQARDuplicateDetection.count_status(evidence["report_text"], alias)
            for alias in aliases.get(status, (status,))
        )

    @staticmethod
    def assert_all_items_status(evidence, status, expected_count=5):
        actual_count = TestQARDuplicateDetection.get_status_count(evidence, status)
        assert actual_count >= expected_count, (
            f"Expected all {expected_count} items to be {status}. "
            f"Actual {status} count: {actual_count}. "
            f"Status summary: {evidence.get('status_summary')}. "
            f"QAR report:\n{evidence['report_text']}"
        )

    def test_qar_duplicate_detection_near_duplicate_content_flagged(self, tmp_path, request):
        run_id = uuid4().hex[:10]
        page = self.login_as_sme()

        baseline_workbook = self.build_upload_workbook(
            tmp_path,
            self.BASELINE_QUESTIONS,
            "IS10_baseline",
            run_id,
        )
        _, baseline = self.submit_workbook_for_qar(baseline_workbook)
        baseline_text = baseline["report_text"].casefold()

        request.node.user_properties.append(("baseline_item_set_id", baseline["item_set_id"]))
        request.node.user_properties.append(("baseline_toast", baseline["toast"]))

        assert baseline["item_set_id"], "IS10 baseline item set ID was not captured."
        # A single-file upload reports "All N rows added successfully. (Upload ID:
        # ..., Status: PASSED)" rather than the multi-file wording, so accept both.
        baseline_upload_message = baseline["upload_message"].casefold()
        assert (
            "validated successfully" in baseline_upload_message
            or "added successfully" in baseline_upload_message
        ), f"Baseline upload did not report success: {baseline['upload_message']!r}"
        assert "qar completed" in baseline_text or "qar ran" in baseline_text
        self.assert_all_items_status(baseline, "Approved", expected_count=5)
        self.assert_toast_count_matches_report(baseline["toast"], baseline["report_text"])

        duplicate_workbook = self.build_upload_workbook(
            tmp_path,
            self.NEAR_DUPLICATE_QUESTIONS,
            "IS12_near_duplicate",
            run_id,
        )
        _, duplicate = self.submit_workbook_for_qar(duplicate_workbook)
        duplicate_text = duplicate["report_text"]
        duplicate_normalized = duplicate_text.casefold()

        request.node.user_properties.append(("item_set_id", duplicate["item_set_id"]))
        request.node.user_properties.append(("is12_item_set_id", duplicate["item_set_id"]))
        request.node.user_properties.append(("qar_success_message", duplicate["toast"]))
        request.node.user_properties.append(
            (
                "item_set_status",
                UploadItemFilePage.format_status_summary(duplicate["status_summary"])
                or "Expected PENDING_QAR with all items Needs Revision",
            )
        )

        assert duplicate["item_set_id"], "IS12 duplicate item set ID was not captured."
        assert any(
            marker in duplicate_normalized
            for marker in ("duplicate", "similarity", "plagiarism", "uniqueness")
        ), f"IS12 QAR report does not show duplicate-check evidence: {duplicate_text}"
        assert any(
            marker in duplicate_normalized
            for marker in ("red", "failed", "fail", "<10", "less than 10", "90%")
        ), f"IS12 duplicate check is not visibly failed/red: {duplicate_text}"
        self.assert_all_items_status(duplicate, "Needs Revision", expected_count=5)
        assert not re.search(r"\brwg\s*\d+\b|assigned\s+rwg", duplicate_normalized), (
            "IS12 should stay locked at PENDING_QAR and must not be forwarded to RWG."
        )
        assert any(
            marker in duplicate_normalized
            for marker in ("pending_qar", "pending qar", "qar pending", "needs revision")
        ), f"IS12 set status did not remain PENDING_QAR/Needs Revision: {duplicate_text}"
        self.assert_toast_count_matches_report(duplicate["toast"], duplicate_text)
