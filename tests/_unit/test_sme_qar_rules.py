from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from time import monotonic
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from pages.common.login_page import LoginPage
from pages.rwg.review_queue_page import RWGReviewQueuePage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.read_config import ReadConfig


@pytest.mark.rtm
@pytest.mark.usefixtures("setup")
class TestSMEQARRules:
    @staticmethod
    def build_scenario_workbook(tmp_path, questions):
        source = Path(ReadConfig.get_upload_item_file_path())
        target = tmp_path / f"qar_scenario_{uuid4().hex[:10]}.xlsx"
        copy2(source, target)
        workbook = load_workbook(target)
        worksheet = workbook.active
        run_id = uuid4().hex[:10]
        for offset, question in enumerate(questions):
            row = offset + 2
            if row != 2:
                for column in range(1, 25):
                    source_cell = worksheet.cell(2, column)
                    target_cell = worksheet.cell(row, column)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell._style = copy_cell_style(source_cell._style)
            worksheet.cell(row, 5).value = offset + 1
            worksheet.cell(row, 10).value = "True or False"
            worksheet.cell(row, 11).value = f"{question} QAR run {run_id}-{offset + 1}"
            worksheet.cell(row, 21).value = "True"
            worksheet.cell(row, 23).value = "Automation QAR scenario explanation."
            worksheet.cell(row, 24).value = "1"
        for row in range(len(questions) + 2, worksheet.max_row + 1):
            for column in range(1, 25):
                worksheet.cell(row, column).value = None
        workbook.save(target)
        workbook.close()
        return target

    def submit_scenario(self, tmp_path, questions):
        self.driver.get(ReadConfig.get_base_url())
        LoginPage(self.driver).login_to_application(
            ReadConfig.get_sme2_username(),
            ReadConfig.get_all_users_password(),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()
        workbook = self.build_scenario_workbook(tmp_path, questions)
        page.open_item_creation_module()
        page.open_upload_item_file_tab()
        page.open_upload_step()
        page.upload_file(workbook)
        page.wait_for_upload_validation_success()
        page.click_continue()
        item_ids = page.get_review_item_ids()
        page.click_continue()
        started = monotonic()
        page.click_submit_for_qar_and_wait_for_results()
        elapsed = monotonic() - started
        assert len(item_ids) == len(questions), (
            f"Expected {len(questions)} uploaded items, received {len(item_ids)}: {item_ids}"
        )
        item_set_id = page.get_item_set_id_from_item_ids(item_ids)
        return (
            page,
            item_ids,
            item_set_id,
            page.driver.find_element("tag name", "body").text,
            elapsed,
        )

    def assert_item_set_blocked_from_rwg(self, item_set_id):
        UploadItemFilePage(self.driver).reset_browser_session_to_login()
        LoginPage(self.driver).login_to_application(
            ReadConfig.get_role_usernames("rwg")[0],
            ReadConfig.get_all_users_password(),
        )
        queue_text = RWGReviewQueuePage(self.driver).get_queue_body_text()
        assert item_set_id not in queue_text, (
            f"QAR-locked set {item_set_id} incorrectly reached the RWG queue."
        )

    def test_qar_report_sections_statuses_and_sla(self, tmp_path):
        questions = [f"Is {80 + i} greater than {10 + i}?" for i in range(10)]
        _, _, _, report, elapsed = self.submit_scenario(tmp_path, questions)
        normalized = report.casefold()

        assert elapsed <= 120, f"QAR exceeded SLA: {elapsed:.2f}s"
        assert "plagiarism" in normalized
        assert "bias" in normalized
        assert "clarity" in normalized or "ambiguous" in normalized
        assert any(status in normalized for status in ("approved", "needs revision", "rejected"))

    def test_qar_flags_paraphrase_bias_and_ambiguity(self, tmp_path):
        questions = [
            "A boy is naturally better at mathematics than a girl.",
            "Only children who celebrate a regional harvest festival can solve this.",
            "What is it?",
            "Reworded: determine whether ninety is larger than nine.",
        ]
        _, _, _, report, _ = self.submit_scenario(tmp_path, questions)
        normalized = report.casefold()

        assert "gender bias" in normalized and ("neutral" in normalized or "remediation" in normalized)
        assert "cultural bias" in normalized and "flag" in normalized
        assert "ambiguous" in normalized or "multiple valid interpretations" in normalized
        assert "paraphrase" in normalized or "similarity" in normalized

    def test_qar_40_percent_failure_does_not_lock_set(self, tmp_path):
        questions = [f"Is {70+i} greater than {20+i}?" for i in range(6)] + ["What is it?"] * 4
        _, _, _, report, _ = self.submit_scenario(tmp_path, questions)
        normalized = report.casefold()

        assert "locked" not in normalized
        assert "partial revision" in normalized or "needs revision" in normalized

    def test_qar_exactly_60_percent_failure_locks_and_blocks_forwarding(self, tmp_path):
        questions = [f"Is {60+i} greater than {10+i}?" for i in range(4)] + ["What is it?"] * 6
        _, _, item_set_id, report, _ = self.submit_scenario(tmp_path, questions)
        normalized = report.casefold()

        assert "locked" in normalized or "set rejected" in normalized
        assert "60%" in normalized or "60 %" in normalized
        self.assert_item_set_blocked_from_rwg(item_set_id)

    def test_qar_70_percent_failure_provides_exception_report(self, tmp_path):
        questions = [f"Is {50+i} greater than {5+i}?" for i in range(3)] + ["What is it?"] * 7
        _, _, item_set_id, report, _ = self.submit_scenario(tmp_path, questions)
        normalized = report.casefold()

        assert "locked" in normalized
        assert "exception report" in normalized
        assert "70%" in normalized or "70 %" in normalized
        self.assert_item_set_blocked_from_rwg(item_set_id)
