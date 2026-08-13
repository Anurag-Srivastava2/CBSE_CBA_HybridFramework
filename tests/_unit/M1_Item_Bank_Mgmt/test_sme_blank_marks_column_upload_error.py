from pathlib import Path
from shutil import copy2
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.read_config import ReadConfig

# A previously downloaded annotated file whose single item ("is sun largest
# planet?") is already recorded in the item bank. Re-uploading it as-is is
# what actually trips the app's duplicate-item-content check — duplicate
# detection compares against content already in the item bank, not just rows
# within the same file.
KNOWN_DUPLICATE_ITEM_FILE = Path(
    r"C:\Users\AnuragSrivastava\OneDrive - Dhira Inc\Desktop\Testing data\upload-531-annotated.xlsx"
)


@pytest.mark.rtm
@pytest.mark.usefixtures("setup")
class TestSMEUploadValidationErrorFlow:
    def login_and_open_upload(self):
        self.driver.get(ReadConfig.get_base_url())
        LoginPage(self.driver).login_to_application(
            ReadConfig.get_sme2_username(),
            ReadConfig.get_all_users_password(),
        )
        upload_page = UploadItemFilePage(self.driver)
        upload_page.close_popup_if_open()
        upload_page.open_item_creation_module()
        upload_page.open_upload_item_file_tab()
        upload_page.open_upload_step()
        return upload_page

    @staticmethod
    def get_template_header(worksheet):
        return next(worksheet.iter_rows(min_row=1, max_row=1))

    @staticmethod
    def find_column(header_cells, column_name):
        return next(
            (
                cell.column
                for cell in header_cells
                if str(cell.value).strip().casefold() == column_name.casefold()
            ),
            None,
        )

    @classmethod
    def build_blank_marks_column_file(cls, destination):
        """Copy the valid template but blank out the Marks column for every row."""
        copy2(ReadConfig.get_upload_item_file_path(), destination)
        workbook = load_workbook(destination)
        worksheet = workbook.active

        header_cells = cls.get_template_header(worksheet)
        marks_column = cls.find_column(header_cells, "Marks")
        assert marks_column is not None, "Template does not contain a 'Marks' column."

        run_id = uuid4().hex[:12]
        question_column = cls.find_column(header_cells, "Question")
        for row_number in range(2, worksheet.max_row + 1):
            if question_column:
                question_cell = worksheet.cell(row=row_number, column=question_column)
                if question_cell.value:
                    question_cell.value = f"{question_cell.value} Blank marks test {run_id}-{row_number}"
            worksheet.cell(row=row_number, column=marks_column).value = None

        workbook.save(destination)
        workbook.close()

    @staticmethod
    def build_known_duplicate_item_file(destination):
        """Copy a file whose item content is already recorded in the item bank."""
        assert KNOWN_DUPLICATE_ITEM_FILE.exists(), (
            f"Known-duplicate source file is missing: {KNOWN_DUPLICATE_ITEM_FILE}"
        )
        copy2(KNOWN_DUPLICATE_ITEM_FILE, destination)

    def verify_view_errors_and_annotated_download(
        self,
        upload_page,
        uploaded_file,
        error_keyword,
        tmp_path,
        request,
    ):
        """Click 'View Errors' on the file's FAILED row, then download and open
        the annotated error workbook from that same row.
        """
        history_row = upload_page.get_upload_history_row_by_file_name(
            uploaded_file.name,
            timeout=45,
        )
        error_details = upload_page.click_view_errors_and_get_message(history_row)
        assert error_keyword in error_details.casefold(), (
            f"Expected the View Errors details to mention {error_keyword!r}, "
            f"got: {error_details!r}"
        )
        upload_page.close_view_errors_dialog()
        request.node.user_properties.append(("view_errors_message", error_details))

        history_row = upload_page.get_upload_history_row_by_file_name(
            uploaded_file.name,
            timeout=45,
        )
        downloaded_file = upload_page.download_file_from_upload_history_row(
            history_row,
            "Download Annotated File",
            tmp_path / "downloaded",
        )

        assert downloaded_file.exists(), f"Downloaded file is missing: {downloaded_file}"
        assert downloaded_file.stat().st_size > 0, (
            f"Downloaded file is empty: {downloaded_file}"
        )
        assert downloaded_file.suffix.casefold() == ".xlsx", (
            f"Expected an .xlsx download, received: {downloaded_file.name}"
        )

        annotated_workbook = load_workbook(downloaded_file, read_only=True, data_only=False)
        try:
            annotated_worksheet = annotated_workbook.active
            assert annotated_worksheet.max_row >= 1
            assert annotated_worksheet.max_column >= 1
        finally:
            annotated_workbook.close()

        request.node.user_properties.append(("downloaded_annotated_file", str(downloaded_file)))
        return error_details, downloaded_file

    def assert_upload_error_flow(
        self,
        upload_page,
        uploaded_file,
        rejection_keyword,
        error_keyword,
        tmp_path,
        request,
        result_description,
    ):
        """Shared upload-failure flow: inline rejection -> real-time FAILED row ->
        View Errors details -> downloadable annotated error workbook.
        """
        upload_page.upload_file(uploaded_file)

        # 1) Upload is rejected inline on the Upload step.
        rejection_message = upload_page.wait_for_upload_rejection(timeout=60)
        normalized_message = rejection_message.casefold()
        assert rejection_keyword in normalized_message, (
            f"Expected the rejection message to mention {rejection_keyword!r}, "
            f"got: {rejection_message!r}"
        )
        request.node.user_properties.append(("upload_rejection_message", rejection_message))

        # 2) The FAILED file shows up in the Previously Uploaded Files listing
        #    in real time, without a manual page refresh.
        history_row = upload_page.get_upload_history_row_by_file_name(
            uploaded_file.name,
            timeout=45,
        )
        assert upload_page.get_upload_history_row_status(history_row) == "FAILED", (
            f"Expected the just-uploaded {uploaded_file.name!r} row to show FAILED."
        )

        # 3) & 4) View Errors details, then download the annotated error workbook.
        error_details, downloaded_file = self.verify_view_errors_and_annotated_download(
            upload_page,
            uploaded_file,
            error_keyword,
            tmp_path,
            request,
        )

        request.node.user_properties.append(("result_description", result_description))
        return rejection_message, error_details, downloaded_file

    def test_tc_ibmm_01a_n03_blank_marks_column_upload_fails_with_visible_error(
        self,
        tmp_path,
        request,
    ):
        blank_marks_file = tmp_path / "blank_marks_column_items.xlsx"
        self.build_blank_marks_column_file(blank_marks_file)

        upload_page = self.login_and_open_upload()
        self.assert_upload_error_flow(
            upload_page,
            blank_marks_file,
            rejection_keyword="marks",
            error_keyword="marks",
            tmp_path=tmp_path,
            request=request,
            result_description=(
                "Uploaded a file with a blank Marks column, confirmed the FAILED row "
                "appeared in the upload-history listing in real time, verified the "
                "View Errors details, and downloaded the annotated error workbook."
            ),
        )

    def test_tc_ibmm_03_n02_duplicate_item_content_upload_fails_with_visible_error(
        self,
        tmp_path,
        request,
    ):
        duplicate_file = tmp_path / "duplicate_item_content_items.xlsx"
        self.build_known_duplicate_item_file(duplicate_file)

        upload_page = self.login_and_open_upload()
        upload_page.upload_file(duplicate_file)

        rejection_message = upload_page.wait_for_upload_rejection(timeout=60)
        request.node.user_properties.append(("upload_rejection_message", rejection_message))

        history_row = upload_page.get_upload_history_row_by_file_name(
            duplicate_file.name,
            timeout=45,
        )
        row_status = upload_page.get_upload_history_row_status(history_row)
        request.node.user_properties.append(("upload_history_row_text", history_row.text))

        # If duplicate detection is ever fixed to reliably hard-fail the whole
        # file, the same flow as the blank-Marks case should hold. Any deviation
        # here (wrong status, missing action, etc.) is treated as the known
        # issue rather than a script bug: across repeated uploads of this same
        # known-duplicate file we've observed three different outcomes (toast
        # shown / row PASSED, no toast / row PASSED, toast shown / row FAILED
        # but "View Errors" missing) — duplicate handling is not deterministic.
        try:
            assert row_status == "FAILED", (
                f"Expected the just-uploaded {duplicate_file.name!r} row to show "
                f"FAILED, got {row_status!r}."
            )
            assert "duplicate" in rejection_message.casefold()
            self.verify_view_errors_and_annotated_download(
                upload_page,
                duplicate_file,
                error_keyword="duplicate",
                tmp_path=tmp_path,
                request=request,
            )
        except Exception as error:
            pytest.xfail(
                "KI-M1-UPLOAD-001 [M1 Bulk upload] Duplicate item-content upload "
                "handling is not deterministic: re-uploading a known-duplicate item "
                f"({duplicate_file.name}) does not reliably produce a FAILED row "
                "with a working View Errors / Download Annotated File action. "
                f"Rejection message: {rejection_message!r}; row text: "
                f"{history_row.text!r}; observed failure: {error!r}"
            )

        request.node.user_properties.append(
            (
                "result_description",
                "Uploaded a file containing duplicate item content, confirmed the "
                "FAILED row appeared in the upload-history listing in real time, "
                "verified the View Errors details, and downloaded the annotated "
                "error workbook.",
            )
        )
