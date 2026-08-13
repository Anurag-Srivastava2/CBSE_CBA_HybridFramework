from pathlib import Path

import pytest
from openpyxl import Workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By

from pages.sme.bulk_upload_page import BulkUploadPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.qar_recovery import (
    build_workbook_qar_correction_factory,
    recover_qar_need_improvement_items,
)


class BodyElement:
    def __init__(self, text):
        self.text = text


class QarRedirectDriver:
    def __init__(self):
        self.body = BodyElement(
            "My Item Set\nItem Set Status\nItem Set Review Stage\nUnder Review"
        )

    def find_element(self, by, value):
        assert (by, value) == (By.TAG_NAME, "body")
        return self.body

    def find_elements(self, *_locator):
        return []


class MissingSubmitWait:
    def until_present(self, _locator, timeout):
        assert timeout == 30
        raise TimeoutException()


class Cell:
    def __init__(self, text):
        self.text = text


class OriginalSubmitButton:
    text = "Submit Set for QAR"

    def __init__(self):
        self.clicked = False

    def click(self):
        self.clicked = True


class ReturnedConditionWait:
    def __init__(self, result):
        self.result = result

    def until_condition(self, _condition, timeout):
        assert timeout == 5
        return self.result


class AlwaysTimeoutWait:
    def until_condition(self, _condition, timeout):
        assert timeout == 90
        raise TimeoutException()


def build_redirected_page():
    page = UploadItemFilePage.__new__(UploadItemFilePage)
    page.driver = QarRedirectDriver()
    return page


def test_upload_file_uses_only_the_native_file_input_event(tmp_path):
    workbook_path = tmp_path / "single-trigger.xlsx"
    workbook_path.write_bytes(b"test workbook placeholder")

    class FileInput:
        def __init__(self):
            self.sent_paths = []

        def send_keys(self, value):
            self.sent_paths.append(value)

    class PresentWait:
        def __init__(self, file_input):
            self.file_input = file_input

        def until_present(self, _locator, timeout):
            assert timeout == 20
            return self.file_input

    class UploadDriver:
        def __init__(self):
            self.scripts = []

        def execute_script(self, script, *arguments):
            self.scripts.append((script, arguments))

    file_input = FileInput()
    page = UploadItemFilePage.__new__(UploadItemFilePage)
    page.driver = UploadDriver()
    page.wait_utils = PresentWait(file_input)

    assert page.upload_file(workbook_path) == workbook_path.resolve()
    assert file_input.sent_paths == [str(workbook_path.resolve())]
    assert len(page.driver.scripts) == 1
    assert "dispatchEvent" not in page.driver.scripts[0][0]


def test_image_upload_stages_excel_then_zip_as_one_workflow(tmp_path):
    workbook_path = tmp_path / "image-items.xlsx"
    images_zip_path = tmp_path / "test-images.zip"
    workbook_path.write_bytes(b"test workbook placeholder")
    images_zip_path.write_bytes(b"test images placeholder")

    events = []

    class UploadPairDriver:
        @staticmethod
        def find_elements(*_locator):
            return []

    class UploadPairWait:
        @staticmethod
        def until_visible(locator, timeout):
            if locator == BulkUploadPage.UPLOAD_IMAGES_ZIP_PANEL:
                assert timeout == 30
                events.append("zip-panel-ready")
                return object()
            if locator == BulkUploadPage.IMAGE_ZIP_UPLOAD_SUCCESS:
                assert timeout == 5
                raise TimeoutException()
            assert locator == BulkUploadPage.CONTINUE_BUTTON
            assert timeout == 20
            events.append("continue-ready")

            class ContinueButton:
                @staticmethod
                def is_enabled():
                    return True

            return ContinueButton()

        @staticmethod
        def until_condition(condition, timeout):
            assert timeout == 90
            outcome = condition(UploadPairDriver())
            assert outcome == {"accepted": True, "message": ""}
            events.append("zip-attached")
            return outcome

    page = BulkUploadPage.__new__(BulkUploadPage)
    page.driver = UploadPairDriver()
    page.wait_utils = UploadPairWait()
    page.open_item_creation_module = lambda: events.append("open-item-creation")
    page.open_upload_item_file_tab = lambda: events.append("open-upload-tab")
    page.open_upload_step = lambda: events.append("open-upload-step")
    page.discard_active_upload_if_present = lambda: events.append("discard-active")
    page.discard_staged_upload_files = lambda: events.append("discard-staged")
    page.pause_before_action = lambda: events.append("pause-for-zip-input")
    page.upload_file = lambda path: events.append(("upload", Path(path)))
    result = page.upload_excel_and_images_zip_for_validation(
        workbook_path, images_zip_path
    )

    assert result == {
        "accepted": True,
        "message": "Images ZIP accepted: the ZIP panel closed and Continue is enabled.",
    }
    assert events == [
        "open-item-creation",
        "open-upload-tab",
        "open-upload-step",
        "discard-active",
        "discard-staged",
        ("upload", workbook_path.resolve()),
        "zip-panel-ready",
        "pause-for-zip-input",
        ("upload", images_zip_path.resolve()),
        "zip-attached",
        "continue-ready",
    ]


def test_image_upload_validates_zip_before_staging_excel(tmp_path):
    workbook_path = tmp_path / "image-items.xlsx"
    workbook_path.write_bytes(b"test workbook placeholder")
    page = BulkUploadPage.__new__(BulkUploadPage)
    page.open_item_creation_module = lambda: pytest.fail(
        "Browser navigation started before the complete upload pair was validated."
    )

    with pytest.raises(FileNotFoundError, match="images ZIP not found"):
        page.upload_excel_and_images_zip_for_validation(
            workbook_path, tmp_path / "missing-test-images.zip"
        )


def test_submit_button_race_accepts_redirect_to_item_sets_as_success():
    page = build_redirected_page()
    page.wait_utils = MissingSubmitWait()

    assert page.click_submit_for_qar()
    assert page.has_qar_results_or_progress(page.driver)
    assert page.is_qar_analysis_complete(page.driver)
    assert "under review" in page.wait_for_ocr_success_message().casefold()


def test_qar_result_ids_ignore_item_set_rows_after_redirect():
    page = build_redirected_page()
    page.driver.find_elements = lambda *_locator: [
        Cell("IS431"),
        Cell("IS431-G1-Mathematics-Ch38-i1"),
        Cell("IS431-G1-Mathematics-Ch38-i2"),
    ]

    assert page.get_qar_result_item_ids() == [
        "IS431-G1-Mathematics-Ch38-i1",
        "IS431-G1-Mathematics-Ch38-i2",
    ]


def test_confirmation_guard_never_clicks_original_qar_submit_again():
    button = OriginalSubmitButton()
    page = build_redirected_page()
    page.wait_utils = ReturnedConditionWait(button)

    assert not page.confirm_submit_if_prompted()
    assert not button.clicked


def test_accepted_submit_is_not_retried_when_result_transition_times_out():
    page = build_redirected_page()
    page.wait_utils = AlwaysTimeoutWait()
    page.submit_calls = 0

    def accepted_submit():
        page.submit_calls += 1
        return True

    page.click_submit_for_qar = accepted_submit
    page.has_qar_results_or_progress = lambda _driver: False
    page.get_visible_qar_submit_blocker = lambda: ""

    with pytest.raises(TimeoutException):
        page.click_submit_for_qar_and_wait_for_results()

    assert page.submit_calls == 1


def test_initial_qar_needs_revision_items_are_corrected_and_rerun(tmp_path):
    workbook_path = tmp_path / "teacher-items.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    original_questions = (
        "Is the reading duration one hour?",
        "How much time passes between 8:00 and 9:00?",
        "The journey lasts ___ hour.",
        "What time is shown on the clock?",
    )
    for row, question in enumerate(original_questions, start=2):
        worksheet.cell(row=row, column=11).value = question
    workbook.save(workbook_path)
    workbook.close()

    class RetryPage:
        QAR_RETRY_STATUS_LABELS = (
            "Need Improvement",
            "Needs Improvement",
            "Needs Revision",
            "Revise",
        )

        def __init__(self):
            self.statuses = [
                {
                    "IS500-G1-Mathematics-Ch38-i1": "Need Improvement",
                    "IS500-G1-Mathematics-Ch38-i2": "Needs Revision",
                },
                {
                    "IS500-G1-Mathematics-Ch38-i1": "Approved",
                    "IS500-G1-Mathematics-Ch38-i2": "Approved",
                },
            ]
            self.corrections = []
            self.reruns = 0
            self.verifications = 0

        def get_qar_item_statuses(self, _item_set_id=None):
            return self.statuses.pop(0)

        def revise_qar_need_improvement_items(
            self,
            _item_set_id,
            item_ids,
            correction_factory,
            retry_number,
        ):
            self.corrections = [
                correction_factory(
                    item_id,
                    retry_number,
                    {
                        "status": "Need Improvement",
                        "failure_reasons": {"Clarity": "flagged"},
                    },
                )
                for item_id in item_ids
            ]
            return item_ids

        @staticmethod
        def compact_item_id(item_id):
            return item_id.replace("-", "")

        def rerun_qar_if_enabled(self):
            self.reruns += 1
            return "QAR rerun completed"

        def verify_item_set_from_sets_module(self, _item_set_id, _item_ids):
            self.verifications += 1

    page = RetryPage()
    result = recover_qar_need_improvement_items(
        page=page,
        item_set_id="IS500-G1-Mathematics-Ch38",
        item_ids=[
            "IS500-G1-Mathematics-Ch38-i1",
            "IS500-G1-Mathematics-Ch38-i2",
        ],
        workbook_path=workbook_path,
        max_retries=3,
        run_label="UNIT",
    )

    assert result.rerun_messages == ("QAR rerun completed",)
    assert result.retry_count == 1
    assert page.reruns == 1
    assert page.verifications == 2
    assert "Clear-language practice card" in page.corrections[0]["question"]
    assert "Clear-language practice card" in page.corrections[1]["question"]
    assert page.corrections[0]["question"] != page.corrections[1]["question"]


def test_workbook_corrections_preserve_typology_answers_and_use_report_reason(
    tmp_path, monkeypatch,
):
    # No question-bank match exists for this env, so the correction factory
    # falls back to its typology-preserving templated rewrite - see
    # test_qar_recovery_question_bank_replacement.py for the bank-backed path.
    empty_bank_path = tmp_path / "empty_bank.json"
    empty_bank_path.write_text("[]", encoding="utf-8")
    monkeypatch.setenv("CBSE_QUESTION_BANK_PATH", str(empty_bank_path))
    monkeypatch.setenv("CBSE_ENV", "unit-test-env")

    workbook_path = tmp_path / "mixed-items.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    rows = (
        ("True or False", "Original false statement?", "False"),
        ("Fill in the Blank", "The duration is ___.", "2"),
        ("Short Answer Question", "State the duration.", "2 hours"),
    )
    for row, (typology, question, answer) in enumerate(rows, start=2):
        worksheet.cell(row=row, column=10).value = typology
        worksheet.cell(row=row, column=11).value = question
        worksheet.cell(row=row, column=21).value = answer
    workbook.save(workbook_path)
    workbook.close()

    correction = build_workbook_qar_correction_factory(
        workbook_path,
        run_label="UNIT",
    )
    feedback = {
        "status": "Need Improvement",
        "failure_reasons": {"Duplicate Detection": "95% similar"},
    }

    true_false = correction("IS600-i1", 1, feedback)
    fill_blank = correction("IS600-i2", 1, feedback)
    short_answer = correction("IS600-i3", 1, feedback)

    assert "Independent duplicate-safe practice card" in true_false["question"]
    assert "4 counters" in true_false["question"]
    assert "9 counters" in true_false["question"]
    assert "recorded answer is 2" in fill_blank["question"]
    assert "answer of 2 hours" in short_answer["question"]
    assert "Duplicate Detection" in true_false["revision_note"]


def test_positive_recovery_rejects_terminal_non_editable_qar_failures(tmp_path):
    workbook_path = tmp_path / "terminal.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=2, column=10).value = "True or False"
    worksheet.cell(row=2, column=11).value = "Is nine greater than four?"
    worksheet.cell(row=2, column=21).value = "True"
    workbook.save(workbook_path)
    workbook.close()

    class TerminalPage:
        QAR_RETRY_STATUS_LABELS = UploadItemFilePage.QAR_RETRY_STATUS_LABELS

        @staticmethod
        def verify_item_set_from_sets_module(_item_set_id, _item_ids):
            return True

        @staticmethod
        def get_qar_item_statuses(_item_set_id=None):
            return {"IS700-i1": "Rejected"}

    with pytest.raises(AssertionError, match="terminal, non-editable failures"):
        recover_qar_need_improvement_items(
            page=TerminalPage(),
            item_set_id="IS700",
            item_ids=["IS700-i1"],
            workbook_path=workbook_path,
        )
