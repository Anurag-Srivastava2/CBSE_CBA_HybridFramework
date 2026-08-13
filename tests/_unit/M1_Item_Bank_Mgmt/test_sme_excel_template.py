import re
from shutil import copy2

from openpyxl import load_workbook
import pytest

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.read_config import ReadConfig


@pytest.mark.rtm
@pytest.mark.usefixtures("setup")
class TestSMEExcelTemplate:
    def login_and_open_upload_tab(self):
        self.driver.get(ReadConfig.get_base_url())
        LoginPage(self.driver).login_to_application(
            ReadConfig.get_sme2_username(),
            ReadConfig.get_all_users_password(),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()
        page.wait_for_application_to_load()
        page.open_item_creation_module()
        page.open_upload_item_file_tab()
        return page

    def download_template(self, tmp_path):
        return self.login_and_open_upload_tab().download_latest_template(tmp_path)

    def test_tc_ibmm_04_p01_template_downloads_with_headers_and_version(self, tmp_path):
        template = self.download_template(tmp_path)
        workbook = load_workbook(template, read_only=True, data_only=False)
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1] if cell.value]
        workbook.close()

        assert template.suffix.casefold() == ".xlsx"
        assert len(headers) >= 15, f"Expected at least 15 required headers, received: {headers}"
        version_text = " ".join([template.name] + [str(header) for header in headers])
        assert re.search(r"v(?:ersion)?[_ -]?\d+", version_text, re.IGNORECASE), (
            f"Downloaded template has no version marker: {version_text}"
        )

    def test_tc_ibmm_04_p02_template_has_controlled_field_validations(self, tmp_path):
        template = self.download_template(tmp_path)
        workbook = load_workbook(template, read_only=False, data_only=False)
        worksheet = workbook.active
        headers = {
            str(cell.value).strip().casefold(): cell.column_letter
            for cell in worksheet[1] if cell.value
        }
        validations = list(worksheet.data_validations.dataValidation)
        workbook.close()

        required = {
            "question_typology": ("question_typology", "typology"),
            "blooms_level": ("blooms_level", "bloom's level", "blooms level", "blooms taxonomy"),
            "difficulty_level": ("difficulty_level", "difficulty level"),
        }
        missing_headers = []
        missing_validations = []
        for name, aliases in required.items():
            column = next((headers[alias] for alias in aliases if alias in headers), None)
            if not column:
                missing_headers.append(name)
                continue
            if not any(
                validation.type == "list" and f"{column}2" in str(validation.sqref)
                for validation in validations
            ):
                missing_validations.append(name)

        assert not missing_headers, f"Controlled-field headers missing: {missing_headers}"
        assert not missing_validations, (
            f"List validation missing from controlled fields: {missing_validations}"
        )

    def test_tc_ibmm_04_n02_old_template_version_shows_upgrade_notice(self, tmp_path):
        page = self.login_and_open_upload_tab()
        current_template = page.download_latest_template(tmp_path / "current")
        old_template = tmp_path / "Item_Upload_Template_v0.xlsx"
        copy2(current_template, old_template)

        page.open_upload_step()
        page.upload_file(old_template)
        message = page.wait_for_upload_rejection(timeout=60)
        normalized = message.casefold()

        assert "version" in normalized
        assert "latest" in normalized or "upgrade" in normalized or "mismatch" in normalized
