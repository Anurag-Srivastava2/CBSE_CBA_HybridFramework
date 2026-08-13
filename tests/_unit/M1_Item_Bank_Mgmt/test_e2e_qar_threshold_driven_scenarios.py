"""Admin-configured QAR thresholds drive every downstream SME assertion."""

import json
from copy import copy as copy_cell_style
from datetime import datetime, timezone
from pathlib import Path
from shutil import copy2
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from pages.admin.qar_config_page import QARConfigPage
from pages.common.login_page import LoginPage
from pages.qar.qar_report_page import QARReportPage
from pages.rwg.review_queue_page import RWGReviewQueuePage
from pages.sme.bulk_upload_page import BulkUploadPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.item_bank_workbook_builder import build_item_workbook
from utilities.qar_artifact_fixture import QARArtifactFixtureBuilder
from utilities.qar_retry_flow import run_qar_need_improvement_retry_loop
from utilities.read_config import ReadConfig


PASS_STATUSES = {"approved", "passed"}
FAIL_STATUSES = {
    "need improvement",
    "needs improvement",
    "needs revision",
    "revise",
    "rejected",
    "failed",
    "blocked",
}
BLOCKER_CHECKS = {
    "Hard Validation",
    "Metadata Alignment",
    "Duplicate Detection",
    "Plagiarism Detection",
}


@pytest.mark.e2e
@pytest.mark.usefixtures("setup")
class TestE2EQARThresholdDrivenScenarios:
    """Fail-fast Admin validation followed by threshold-driven SME scenarios."""

    @staticmethod
    def make_prefix(scenario):
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
        return f"QAR_AUTO_CROSS_{timestamp}_{scenario.upper()}_{uuid4().hex[:6]}"

    def reset_and_login(self, username, initial=False):
        if initial:
            self.driver.get(ReadConfig.get_base_url())
        else:
            UploadItemFilePage(self.driver).reset_browser_session_to_login()
        LoginPage(self.driver).login_to_application(
            username,
            ReadConfig.get_all_users_password(),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()
        return page

    def read_and_validate_admin_thresholds(self):
        admin_username = ReadConfig.get_role_usernames("admin")[0]
        self.reset_and_login(admin_username, initial=True)
        return QARConfigPage(self.driver).open().get_threshold_snapshot()

    def submit_workbook(self, workbook_path, prefix):
        self.reset_and_login(ReadConfig.get_sme2_username())
        page = BulkUploadPage(self.driver)
        validation = page.upload_excel_for_validation(workbook_path)
        assert validation["accepted"], (
            f"Threshold scenario {prefix} was rejected before QAR: {validation['message']}"
        )
        submission = page.submit_for_qar(prefix)
        report = QARReportPage(self.driver)
        report_text = report.wait_until_processed(submission["item_set_id"], timeout=180)
        assert submission["item_ids"], f"QAR returned no item IDs for {prefix}."
        return {
            **submission,
            "page": page,
            "report": report,
            "report_text": report_text,
            "prefix": prefix,
        }

    def submit_artifact_scenario(self, tmp_path, scenario_name):
        prefix = self.make_prefix(scenario_name)
        fixture = QARArtifactFixtureBuilder.build(
            ReadConfig.get_upload_item_file_path(),
            tmp_path / scenario_name,
            prefix,
            scenario_name,
        )
        result = self.submit_workbook(fixture["outputPath"], prefix)
        result["typologies"] = {
            item_id: "True or False" for item_id in result["item_ids"]
        }
        return result

    @staticmethod
    def effective_threshold(thresholds, typology):
        normalized_typology = typology.casefold()
        for configured_typology, threshold in thresholds["typology_overrides"].items():
            configured = configured_typology.casefold()
            if configured in normalized_typology or normalized_typology in configured:
                return float(threshold)
        return float(thresholds["global_pass_threshold"])

    def assert_result_matches_admin_thresholds(self, result, thresholds):
        evidence = {}
        for item_id in result["item_ids"]:
            typology = result["typologies"][item_id]
            required_score = self.effective_threshold(thresholds, typology)
            score = result["report"].get_item_score(item_id)
            status = result["report"].get_item_status(item_id)
            reasons = result["report"].get_item_failure_reasons(item_id)
            assert score is not None, (
                f"QAR report did not expose an overall score for {item_id}; cannot "
                f"cross-check live threshold {required_score}%."
            )
            assert status, f"QAR report did not expose a status for {item_id}."
            blocker_failure = bool(BLOCKER_CHECKS.intersection(reasons))
            should_pass = score >= required_score and not blocker_failure
            if should_pass:
                assert status.casefold() in PASS_STATUSES, (
                    f"{item_id} scored {score}% against {required_score}% with no blocker, "
                    f"but status was {status!r}."
                )
            else:
                assert status.casefold() in FAIL_STATUSES, (
                    f"{item_id} scored {score}% against {required_score}% or failed a "
                    f"blocker, but status was {status!r}."
                )
                assert reasons, (
                    f"Failed item {item_id} showed no criterion-specific failure reason."
                )
                assert all(reason.strip() for reason in reasons.values())
            evidence[item_id] = {
                "typology": typology,
                "score": score,
                "threshold": required_score,
                "status": status,
                "failure_checks": sorted(reasons),
            }
        return evidence

    def assert_rwg_routing(self, item_set_id, should_route):
        visible_to = []
        for rwg_username in dict.fromkeys(ReadConfig.get_role_usernames("rwg")):
            self.reset_and_login(rwg_username)
            queue_text = RWGReviewQueuePage(self.driver).get_queue_body_text()
            if item_set_id in queue_text:
                visible_to.append(rwg_username)
        if should_route:
            assert visible_to, f"Passing set {item_set_id} did not reach any RWG queue."
        else:
            assert not visible_to, (
                f"Failing set {item_set_id} incorrectly reached RWG users {visible_to}."
            )

    @staticmethod
    def build_lock_boundary_workbook(tmp_path, prefix, failed_count, total_count):
        assert 3 <= total_count <= 5
        source = Path(ReadConfig.get_upload_item_file_path())
        target = tmp_path / f"{prefix}.xlsx"
        copy2(source, target)
        workbook = load_workbook(target)
        worksheet = workbook.active
        for offset in range(total_count):
            row = offset + 2
            if row > 2:
                for column in range(1, 25):
                    source_cell = worksheet.cell(2, column)
                    target_cell = worksheet.cell(row, column)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell._style = copy_cell_style(source_cell._style)
            worksheet.cell(row, 5).value = offset + 1
            worksheet.cell(row, 10).value = "True or False"
            if offset < failed_count:
                question = f"{prefix} ambiguous {offset + 1}: What does it mean there?"
            else:
                question = f"{prefix} valid {offset + 1}: Is {20 + offset} greater than 2?"
            worksheet.cell(row, 11).value = question
            worksheet.cell(row, 21).value = "True"
            worksheet.cell(row, 23).value = f"{prefix} explanation {offset + 1}."
            worksheet.cell(row, 24).value = "1"
        for row in range(total_count + 2, worksheet.max_row + 1):
            for column in range(1, 25):
                worksheet.cell(row, column).value = None
        workbook.save(target)
        workbook.close()
        return target

    def build_mixed_typology_workbook(self, tmp_path, prefix):
        path = tmp_path / f"{prefix}.xlsx"
        _, summaries = build_item_workbook(
            ReadConfig.get_upload_item_file_path(),
            path,
            count=5,
            seed=prefix,
        )
        typologies = [
            item["typology"]
            for items in summaries.values()
            for item in items
        ]
        assert len(typologies) == 5 and len(set(typologies)) >= 3
        return path, typologies

    def open_sme_set(self, item_set_id):
        page = self.reset_and_login(ReadConfig.get_sme2_username())
        page.open_sets_module()
        page.open_item_set_from_sets_list(item_set_id)
        return page

    @staticmethod
    def valid_correction(prefix):
        def correction(item_id, retry_number):
            return {
                "question": (
                    f"{prefix} corrected {item_id} retry {retry_number}: "
                    "Is nine greater than four?"
                ),
                "revision_note": f"Corrected threshold-driven retry {retry_number}.",
            }
        return correction

    @staticmethod
    def persistent_failure(prefix):
        def correction(item_id, retry_number):
            return {
                "question": f"{prefix} persistent duplicate: Does one hour contain sixty minutes?",
                "revision_note": f"Intentional persistent failure retry {retry_number}.",
            }
        return correction

    def run_retry_flow(self, submission, correction_factory, max_retries):
        item_set_id = submission["item_set_id"]
        active_page = {"page": None}

        def get_failed_item_ids():
            page = self.open_sme_set(item_set_id)
            active_page["page"] = page
            statuses = page.get_qar_item_statuses()
            return [
                item_id
                for item_id, status in statuses.items()
                if status.casefold() in FAIL_STATUSES
            ]

        def correct_items(item_ids, retry_number):
            page = self.open_sme_set(item_set_id)
            active_page["page"] = page
            return page.revise_qar_need_improvement_items(
                item_set_id,
                item_ids,
                correction_factory,
                retry_number,
            )

        def rerun_qar(retry_number):
            page = active_page["page"]
            assert page and page.is_rerun_qar_enabled(), (
                f"Re-run QAR disabled before configured retry {retry_number}/{max_retries}."
            )
            return page.rerun_qar_if_enabled()

        return run_qar_need_improvement_retry_loop(
            item_set_id=item_set_id,
            get_need_improvement_item_ids=get_failed_item_ids,
            correct_items=correct_items,
            rerun_qar=rerun_qar,
            assert_blocked_from_rwg=lambda set_id, failed: self.assert_rwg_routing(
                set_id, False
            ),
            assert_routed_to_rwg=lambda set_id: self.assert_rwg_routing(set_id, True),
            max_retries=max_retries,
        )

    def test_e2e_admin_thresholds_drive_qar_scenarios(self, request, tmp_path):
        # Phase 1: fail fast on missing/disabled/malformed Admin configuration.
        thresholds = self.read_and_validate_admin_thresholds()
        request.node.user_properties.append(
            ("qar_threshold_snapshot", json.dumps(thresholds, sort_keys=True))
        )

        # Phase 2a: all-pass and mixed pass/fail outcomes.
        happy = self.submit_artifact_scenario(tmp_path, "positive")
        happy_evidence = self.assert_result_matches_admin_thresholds(happy, thresholds)
        assert all(row["status"].casefold() in PASS_STATUSES for row in happy_evidence.values())
        self.assert_rwg_routing(happy["item_set_id"], True)

        mixed = self.submit_artifact_scenario(tmp_path, "duplicate")
        mixed_evidence = self.assert_result_matches_admin_thresholds(mixed, thresholds)
        mixed_statuses = {row["status"].casefold() for row in mixed_evidence.values()}
        assert mixed_statuses.intersection(PASS_STATUSES), "Mixed fixture produced no passing item."
        assert mixed_statuses.intersection(FAIL_STATUSES), "Mixed fixture produced no failing item."
        self.assert_rwg_routing(mixed["item_set_id"], False)

        # Phase 2b: edit/rerun uses the live Admin retry ceiling, not a constant.
        retry_result = self.run_retry_flow(
            mixed,
            self.valid_correction(mixed["prefix"]),
            thresholds["max_retry_limit"],
        )
        assert 1 <= retry_result.retry_count <= thresholds["max_retry_limit"]

        persistent = self.submit_artifact_scenario(tmp_path, "duplicate")
        with pytest.raises(
            AssertionError,
            match=rf"after {thresholds['max_retry_limit']} retries",
        ):
            self.run_retry_flow(
                persistent,
                self.persistent_failure(persistent["prefix"]),
                thresholds["max_retry_limit"],
            )
        self.assert_rwg_routing(persistent["item_set_id"], False)

        # Phase 2c: exact set-lock boundary and one-item-below boundary.
        lock_threshold = float(thresholds["set_lock_threshold"])
        boundary_shape = next(
            (
                (total, int(total * lock_threshold / 100))
                for total in range(3, 6)
                if (total * lock_threshold / 100).is_integer()
            ),
            None,
        )
        assert boundary_shape, (
            f"Admin set-lock threshold {lock_threshold}% cannot be represented with "
            "the required 3–5 questions per sheet."
        )
        total_count, boundary_failures = boundary_shape
        assert boundary_failures >= 1
        for label, intended_failures, should_lock in (
            ("boundary", boundary_failures, True),
            ("below_boundary", boundary_failures - 1, False),
        ):
            prefix = self.make_prefix(label)
            workbook = self.build_lock_boundary_workbook(
                tmp_path,
                prefix,
                intended_failures,
                total_count,
            )
            result = self.submit_workbook(workbook, prefix)
            result["typologies"] = {
                item_id: "True or False" for item_id in result["item_ids"]
            }
            evidence = self.assert_result_matches_admin_thresholds(result, thresholds)
            actual_failures = sum(
                row["status"].casefold() in FAIL_STATUSES for row in evidence.values()
            )
            assert actual_failures == intended_failures, (
                f"{label} fixture expected {intended_failures}/{total_count} failures, "
                f"but QAR produced {actual_failures}: {evidence}"
            )
            normalized_report = result["report_text"].casefold()
            lock_visible = "locked" in normalized_report or "set rejected" in normalized_report
            assert lock_visible is should_lock, (
                f"Set-lock boundary mismatch at {actual_failures}/{total_count}; "
                f"Admin threshold={lock_threshold}%, report={result['report_text'][:600]}"
            )
            self.assert_rwg_routing(result["item_set_id"], False)

        # Phase 2d: different typologies use their override or the global score.
        typology_prefix = self.make_prefix("typologies")
        typology_workbook, typologies = self.build_mixed_typology_workbook(
            tmp_path,
            typology_prefix,
        )
        typology_result = self.submit_workbook(typology_workbook, typology_prefix)
        assert len(typology_result["item_ids"]) == len(typologies)
        typology_result["typologies"] = dict(zip(typology_result["item_ids"], typologies))
        typology_evidence = self.assert_result_matches_admin_thresholds(
            typology_result,
            thresholds,
        )
        self.assert_rwg_routing(
            typology_result["item_set_id"],
            all(row["status"].casefold() in PASS_STATUSES for row in typology_evidence.values()),
        )

        request.node.user_properties.append(
            (
                "result_description",
                "Admin QAR thresholds drove happy, mixed, retry-ceiling, exact lock-boundary, "
                "below-boundary, and mixed-typology SME assertions.",
            )
        )
