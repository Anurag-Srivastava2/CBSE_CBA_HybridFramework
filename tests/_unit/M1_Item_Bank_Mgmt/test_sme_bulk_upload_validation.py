from shutil import copy2
from uuid import uuid4

from openpyxl import load_workbook
import pytest

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.read_config import ReadConfig


@pytest.mark.rtm
@pytest.mark.usefixtures("setup")
class TestSMEBulkUploadValidation:
    def login_and_open_upload(self):
        self.driver.get(ReadConfig.get_base_url())
        LoginPage(self.driver).login_to_application(
            ReadConfig.get_sme2_username(),
            ReadConfig.get_all_users_password(),
        )
        upload_page = UploadItemFilePage(self.driver)
        upload_page.close_popup_if_open()
        upload_page.open_item_creation_module()
        upload_page.open_upload_item_file_tab()
        upload_page.open_upload_step()
        return upload_page

    def test_tc_ibmm_01a_n01_non_xlsx_file_is_rejected(self, tmp_path):
        invalid_file = tmp_path / "sample_items.pdf"
        invalid_file.write_bytes(b"%PDF-1.4\n% invalid item-upload test file\n")

        upload_page = self.login_and_open_upload()
        upload_page.upload_file(invalid_file)
        rejection_message = upload_page.wait_for_upload_rejection()

        normalized_message = rejection_message.casefold()
        assert "xlsx" in normalized_message
        assert "invalid" in normalized_message or "only" in normalized_message

    def test_tc_ibmm_01a_n02_modified_header_upload_fails(self, tmp_path):
        modified_file = tmp_path / "modified_header_items.xlsx"
        copy2(ReadConfig.get_upload_item_file_path(), modified_file)
        workbook = load_workbook(modified_file)
        worksheet = workbook.active
        worksheet.cell(row=1, column=1).value = "Modified Grade Header"
        run_id = uuid4().hex[:12]
        for row_number in range(2, worksheet.max_row + 1):
            question_cell = worksheet.cell(row=row_number, column=11)
            if question_cell.value:
                question_cell.value = f"{question_cell.value} Header test {run_id}-{row_number}"
        workbook.save(modified_file)
        workbook.close()

        upload_page = self.login_and_open_upload()
        upload_page.upload_file(modified_file)
        rejection_message = upload_page.wait_for_upload_rejection(timeout=60)

        normalized_message = rejection_message.casefold()
        assert "failed" in normalized_message or "error" in normalized_message
        assert "header" in normalized_message or "validation" in normalized_message

    def test_tc_ibmm_03_n01_duplicate_item_content_is_rejected(self, tmp_path):
        duplicate_file = tmp_path / "duplicate_item_content.xlsx"
        copy2(ReadConfig.get_upload_item_file_path(), duplicate_file)
        workbook = load_workbook(duplicate_file)
        worksheet = workbook.active
        run_id = uuid4().hex[:12]
        for row_number in range(2, worksheet.max_row + 1):
            question_cell = worksheet.cell(row=row_number, column=11)
            if question_cell.value:
                question_cell.value = f"{question_cell.value} Duplicate test {run_id}-{row_number}"
        duplicate_values = [
            worksheet.cell(row=2, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]
        worksheet.append(duplicate_values)
        workbook.save(duplicate_file)
        workbook.close()

        upload_page = self.login_and_open_upload()
        upload_page.upload_file(duplicate_file)
        rejection_message = upload_page.wait_for_upload_rejection(timeout=60)
        normalized_message = rejection_message.casefold()

        assert "duplicate" in normalized_message
        assert "row" in normalized_message or "item" in normalized_message
