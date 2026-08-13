from openpyxl import load_workbook

from utilities.qar_plagiarism_fixture import (
    EXPECTED_PLAGIARISM_THRESHOLD,
    PUBLISHED_SOURCE_ITEMS,
    build_qar_plagiarism_workbook,
    source_similarity,
)
from utilities.read_config import ReadConfig


def test_pdf_plagiarism_fixture_preserves_six_verbatim_sources(tmp_path):
    output, evidence = build_qar_plagiarism_workbook(
        ReadConfig.get_upload_item_file_path(),
        tmp_path / "pdf-plagiarism.xlsx",
        "QAR_AUTO_PDF_PLAG_UNIT",
    )
    assert len(evidence) == 6
    assert len({row["source_item_id"] for row in evidence}) == 6
    assert {row["source_pdf_page"] for row in evidence} == {1, 2, 3}
    assert all(
        row["source_similarity"] >= EXPECTED_PLAGIARISM_THRESHOLD
        for row in evidence
    )

    workbook = load_workbook(output, data_only=True)
    worksheet = workbook.active
    for row, source in enumerate(PUBLISHED_SOURCE_ITEMS, start=2):
        assert worksheet.cell(row, 10).value == "True or False"
        assert worksheet.cell(row, 11).value == source.question
        assert source_similarity(worksheet.cell(row, 11).value, source.question) == 100
        assert worksheet.cell(row, 21).value == source.answer
    workbook.close()


def test_similarity_contract_rejects_content_below_97_percent():
    source = PUBLISHED_SOURCE_ITEMS[0].question
    assert source_similarity(source, source) == 100
    assert source_similarity("An unrelated new question.", source) < 97
