import json

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from utilities.item_bank_workbook_builder import build_item_workbook
from utilities.question_bank_manager import content_hash


def _seed_isolated_question_bank(monkeypatch, tmp_path):
    """Point the builder at a throwaway bank/env so this offline unit test never
    consumes real per-environment usage from the shared data/question_bank file."""
    typologies = (
        "Multiple Choice Question", "Fill in the Blank", "Match the Following", "True or False",
    )
    records = []
    index = 0
    for repeat in range(2):
        for typology in typologies:
            index += 1
            question = f"Fixture question {index} for {typology}"
            records.append({
                "id": f"FIXTURE-{index:03d}",
                "typology": typology,
                "grade": "5",
                "subject": "Mathematics",
                "chapter": "Fixture Chapter",
                "competency": "Fixture Competency",
                "blooms_level": "Understanding",
                "question": question,
                "options": ["A", "B", "C", "D"] if typology == "Multiple Choice Question" else [],
                "answer": (
                    "A" if typology == "Multiple Choice Question"
                    else "True" if typology == "True or False"
                    else "Fixture answer"
                ),
                "explanation": "Fixture explanation.",
                "marks": 1,
                "content_hash": content_hash(question),
                "usage": {},
            })
    bank_path = tmp_path / "fixture_questions.json"
    bank_path.write_text(json.dumps(records), encoding="utf-8")
    monkeypatch.setenv("CBSE_QUESTION_BANK_PATH", str(bank_path))
    monkeypatch.setenv("CBSE_ENV", "offline-fixture-tests")


HEADERS = (
    "Grade",
    "Subject",
    "Unit/Theme",
    "Chapter No._Name",
    "S.No.",
    "Competency",
    "Learning Outcome",
    "Blooms Taxonomy",
    "Explanation of Blooms Taxonomy",
    "Typology",
    "Question",
    "Question Image",
    "Option 1",
    "Option 2",
    "Option 3",
    "Option 4",
    "Image 1",
    "Image 2",
    "Image 3",
    "Image 4",
    "Answer",
    "Answer Image",
    "Explanation/Remarks",
    "Marks",
)


def _add_item_sheet(workbook, title):
    worksheet = workbook.create_sheet(title)
    for column, header in enumerate(HEADERS, start=1):
        worksheet.cell(row=1, column=column).value = header
    metadata = (
        "Grade 1",
        "Mathematics",
        1,
        "How do I Spend my Day? (Time)",
        1,
        "Uses time in daily life",
        "Reads time to the hour",
        "Applying",
        "Uses a clock to reason about time",
    )
    for column, value in enumerate(metadata, start=1):
        worksheet.cell(row=2, column=column).value = value
    worksheet["K6"] = "stale item that must be cleared"
    worksheet["CA2"] = "True or False"
    validation = DataValidation(type="list", formula1='"True or False,Fill in the Blank"')
    worksheet.add_data_validation(validation)
    validation.add("J2:J500")
    return worksheet


def _create_template(path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_item_sheet(workbook, "Items")
    _add_item_sheet(workbook, "Items Hindi")
    lookup = workbook.create_sheet("Lookups")
    lookup["A1"] = "must remain unchanged"
    workbook.save(path)
    workbook.close()


def test_builder_creates_four_mixed_non_image_items_per_sheet(tmp_path, monkeypatch):
    _seed_isolated_question_bank(monkeypatch, tmp_path)
    template = tmp_path / "template.xlsx"
    output = tmp_path / "offline-items.xlsx"
    _create_template(template)

    _, summaries = build_item_workbook(
        template,
        output,
        count=4,
        seed="offline-verification",
    )

    assert {name: len(items) for name, items in summaries.items()} == {
        "Items": 4,
        "Items Hindi": 4,
    }

    workbook = load_workbook(output, read_only=False, data_only=False)
    expected_typologies = [
        "Multiple Choice Question",
        "Fill in the Blank",
        "Match the Following",
        "True or False",
    ]
    all_questions = []
    for sheet_name in ("Items", "Items Hindi"):
        worksheet = workbook[sheet_name]
        questions = [worksheet.cell(row=row, column=11).value for row in range(2, 6)]
        all_questions.extend(questions)
        assert all(questions)
        assert worksheet["K6"].value is None
        assert [worksheet.cell(row=row, column=10).value for row in range(2, 6)] == (
            expected_typologies
        )
        expected_options = [
            *summaries[sheet_name][1]["options"],
            *([None] * (4 - len(summaries[sheet_name][1]["options"]))),
        ]
        assert [
            worksheet.cell(row=3, column=column).value for column in range(13, 17)
        ] == expected_options
        assert worksheet["CA2"].value == "True or False"
        assert len(worksheet.data_validations.dataValidation) == 1
        assert len(worksheet._images) == 0
        assert worksheet["L5"].value is None

    assert len(all_questions) == len(set(all_questions))
    assert workbook["Lookups"]["A1"].value == "must remain unchanged"
    workbook.close()
