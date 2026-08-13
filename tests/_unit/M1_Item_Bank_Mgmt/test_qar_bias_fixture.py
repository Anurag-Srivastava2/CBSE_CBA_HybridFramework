from openpyxl import load_workbook
import pytest

from utilities.qar_bias_fixture import (
    BIAS_QUESTION_SPECS,
    assert_bias_score_contract,
    build_qar_bias_workbook,
)
from utilities.read_config import ReadConfig


def test_bias_fixture_builds_six_unique_ordered_items(tmp_path):
    prefix = "QAR_AUTO_BIAS_UNIT_123456"
    output, specs = build_qar_bias_workbook(
        ReadConfig.get_upload_item_file_path(),
        tmp_path / "bias.xlsx",
        prefix,
    )
    assert len(specs) == 6
    assert [spec["severity_rank"] for spec in specs] == list(range(1, 7))
    assert len({spec["topic"] for spec in specs}) == 6

    workbook = load_workbook(output, data_only=True)
    worksheet = workbook.active
    assert worksheet.max_row >= 7
    for row, spec in enumerate(BIAS_QUESTION_SPECS, start=2):
        assert worksheet.cell(row, 10).value == "True or False"
        assert worksheet.cell(row, 11).value == spec.question
        assert worksheet.cell(row, 21).value == "False"
        assert worksheet.cell(row, 23).value.startswith(prefix)
        assert spec.explanation in worksheet.cell(row, 23).value
        assert worksheet.cell(row, 24).value in (1, "1")
    workbook.close()


def test_bias_score_contract_allows_equal_flagged_scores():
    evidence = [
        {"severity_rank": 1, "score": 70, "threshold": 60},
        {"severity_rank": 2, "score": 55, "threshold": 60},
        {"severity_rank": 3, "score": 40, "threshold": 60},
        {"severity_rank": 4, "score": 20, "threshold": 60},
        {"severity_rank": 5, "score": 0, "threshold": 60},
        {"severity_rank": 6, "score": 0, "threshold": 60},
    ]
    assert assert_bias_score_contract(evidence) == (70, 55, 40, 20, 0, 0)


@pytest.mark.parametrize(
    "scores,thresholds",
    (
        ((50, 50, 50, 50, 70, 70), (60, 60, 60, 60, 60, 60)),
        ((10, 20, 30, 40, 50, 60), (60, 60, 60, 60, 60, 60)),
        ((-1, 20, 30, 40, 50, 50), (60, 60, 60, 60, 60, 60)),
        ((10, 20, 30, 40, 50, 50), (60, 60, 60, 60, 101, 60)),
    ),
)
def test_bias_score_contract_rejects_invalid_or_missed_results(scores, thresholds):
    evidence = [
        {"severity_rank": rank, "score": score, "threshold": threshold}
        for rank, (score, threshold) in enumerate(zip(scores, thresholds), start=1)
    ]
    with pytest.raises(AssertionError):
        assert_bias_score_contract(evidence)
