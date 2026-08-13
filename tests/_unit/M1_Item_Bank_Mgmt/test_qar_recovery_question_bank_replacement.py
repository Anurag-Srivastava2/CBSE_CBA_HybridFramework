import json

from openpyxl import Workbook

from utilities.qar_recovery import build_workbook_qar_correction_factory
from utilities.question_bank_manager import content_hash


def _write_uploaded_workbook(path, typology, question, options, answer):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=2, column=10).value = typology
    worksheet.cell(row=2, column=11).value = question
    for offset, option in enumerate(options):
        worksheet.cell(row=2, column=13 + offset).value = option
    worksheet.cell(row=2, column=21).value = answer
    workbook.save(path)
    workbook.close()


def _seed_bank(path, typology, replacement_question):
    record = {
        "id": "QB-REPLACEMENT-001",
        "typology": typology,
        "grade": "5",
        "subject": "Mathematics",
        "chapter": "Fixture Chapter",
        "competency": "Fixture Competency",
        "blooms_level": "Understanding",
        "question": replacement_question,
        "options": [],
        "answer": "Fixture answer",
        "explanation": "Fixture explanation.",
        "marks": 1,
        "content_hash": content_hash(replacement_question),
        "usage": {},
    }
    path.write_text(json.dumps([record]), encoding="utf-8")


def test_duplicate_failure_is_corrected_with_a_bank_backed_replacement(tmp_path, monkeypatch):
    typology = "Short Answer Question"
    original_question = "What is the capital of a landlocked country in this practice set?"
    replacement_question = "Explain briefly why the sky appears blue during the day."

    workbook_path = tmp_path / "uploaded.xlsx"
    _write_uploaded_workbook(
        workbook_path,
        typology=typology,
        question=original_question,
        options=[],
        answer="Because of Rayleigh scattering.",
    )

    bank_path = tmp_path / "bank.json"
    _seed_bank(bank_path, typology, replacement_question)
    monkeypatch.setenv("CBSE_QUESTION_BANK_PATH", str(bank_path))
    monkeypatch.setenv("CBSE_ENV", "recovery-test-env")

    correction = build_workbook_qar_correction_factory(workbook_path, run_label="RECOVERY-TEST")
    result = correction(
        "SET123-i1",
        retry_number=1,
        qar_feedback={"failure_reasons": {"Duplicate Detection": "Matches an existing item."}},
    )

    assert result["question"] == replacement_question
    assert "question-bank" in result["revision_note"]

    persisted = json.loads(bank_path.read_text(encoding="utf-8"))
    assert persisted[0]["usage"]["recovery-test-env"]["status"] == "used"


def test_duplicate_failure_falls_back_to_templated_rewrite_when_bank_exhausted(tmp_path, monkeypatch):
    typology = "Fill in the Blank"
    original_question = "Fill in the blank about a duplicated fact."

    workbook_path = tmp_path / "uploaded.xlsx"
    _write_uploaded_workbook(
        workbook_path,
        typology=typology,
        question=original_question,
        options=[],
        answer="42",
    )

    bank_path = tmp_path / "bank.json"
    bank_path.write_text(json.dumps([]), encoding="utf-8")
    monkeypatch.setenv("CBSE_QUESTION_BANK_PATH", str(bank_path))
    monkeypatch.setenv("CBSE_ENV", "recovery-test-env")

    correction = build_workbook_qar_correction_factory(workbook_path, run_label="RECOVERY-TEST")
    result = correction(
        "SET123-i1",
        retry_number=1,
        qar_feedback={"failure_reasons": {"Duplicate Detection": "Matches an existing item."}},
    )

    assert "recorded answer" in result["question"]
    assert "question-bank" not in result["revision_note"]
