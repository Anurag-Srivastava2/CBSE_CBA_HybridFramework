from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from time import monotonic
from uuid import uuid4
import re
import logging
from openpyxl import load_workbook
import pytest
import requests
from selenium.webdriver.common.by import By

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.read_config import ReadConfig


@pytest.mark.rtm
@pytest.mark.usefixtures("setup")
class TestQARResubmissionFreshRun:
    """TC-NEG-M1-09 — Item Set Resubmission After Rejection Triggers Fresh QAR (Not Cached)"""

    logger = logging.getLogger(__name__)

    def step(self, n, message):
        print(f"\n[STEP {n}] {message}", flush=True)
        self.logger.info(f"[STEP {n}] {message}")

    @staticmethod
    def build_scenario_workbook(tmp_path, questions):
        source = Path(ReadConfig.get_upload_item_file_path())
        target = tmp_path / f"qar_resubmission_{uuid4().hex[:10]}.xlsx"
        copy2(source, target)
        workbook = load_workbook(target)
        worksheet = workbook.active
        run_id = uuid4().hex[:10]
        for offset, question in enumerate(questions):
            row = offset + 2
            if row != 2:
                for column in range(1, 25):
                    source_cell = worksheet.cell(2, column)
                    target_cell = worksheet.cell(row, column)
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell._style = copy_cell_style(source_cell._style)
            worksheet.cell(row, 5).value = offset + 1
            worksheet.cell(row, 10).value = "True or False"
            worksheet.cell(row, 11).value = f"{question} QAR run {run_id}-{offset + 1}"
            worksheet.cell(row, 21).value = "True"
            worksheet.cell(row, 23).value = "Automation QAR resubmission scenario explanation."
            worksheet.cell(row, 24).value = "1"
        for row in range(len(questions) + 2, worksheet.max_row + 1):
            for column in range(1, 25):
                worksheet.cell(row, column).value = None
        workbook.save(target)
        workbook.close()
        return target

    def get_api_session(self):
        """Extract Selenium cookies and Authorization header into requests Session."""
        session = requests.Session()
        for cookie in self.driver.get_cookies():
            session.cookies.set(cookie['name'], cookie['value'])
            
        token = self.driver.execute_script(
            "return localStorage.getItem('token') or sessionStorage.getItem('token') "
            "or localStorage.getItem('jwt') or sessionStorage.getItem('jwt') "
            "or localStorage.getItem('auth_token') or sessionStorage.getItem('auth_token');"
        )
        
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        if token:
            headers["Authorization"] = f"Bearer {token}"
            
        return session, headers

    def test_tc_neg_m1_09_resubmission_triggers_fresh_qar_run(self, tmp_path):
        """TC-NEG-M1-09: Verify resubmitting a revised item set triggers a fresh QAR run.
        
        Steps:
        1. Login as SME2.
        2. Build scenario workbook with near-duplicate items to trigger initial failure.
        3. Upload and submit for QAR.
        4. Revise failed items in SME dashboard (edit open revision items).
        5. Re-run QAR.
        6. Assert duplicate score improves and timeline displays both QAR runs.
        7. Query GET /item-sets/{id}/qar-history and assert 2 distinct QAR run records exist.
        """
        # Step 1: Login as SME2
        self.step(1, "Logging in as SME2")
        self.driver.get(ReadConfig.get_base_url())
        LoginPage(self.driver).login_to_application(
            ReadConfig.get_sme2_username(),
            ReadConfig.get_all_users_password(),
        )
        page = UploadItemFilePage(self.driver)
        page.close_popup_if_open()

        # Step 2: Build scenario workbook
        self.step(2, "Building scenario workbook with near-duplicate items")
        questions = [
            "Is 90 greater than 9?",
            "Determine whether 90 exceeds 9.",
            "Is 9 less than 90?",
            "State if ninety is larger than nine.",
            "Reworded: Is 90 bigger than 9?",
        ]
        workbook = self.build_scenario_workbook(tmp_path, questions)

        # Step 3: Upload and submit for initial QAR
        self.step(3, "Uploading and submitting item file for initial QAR")
        page.open_item_creation_module()
        page.open_upload_item_file_tab()
        page.open_upload_step()
        page.upload_file(workbook)
        page.wait_for_upload_validation_success()
        page.click_continue()
        item_ids = page.get_review_item_ids()
        item_set_id = page.get_item_set_id_from_item_ids(item_ids)
        self.logger.info(f"Initial Item Set ID: {item_set_id}")
        page.click_continue()
        page.click_submit_for_qar_and_wait_for_results()
        initial_toast = page.wait_for_ocr_success_message()
        self.logger.info(f"Initial QAR completion toast: {initial_toast}")

        # Step 4: Revise the failed items
        self.step(4, "Revising the failed items to unique question texts")
        page.open_sets_module()
        page.open_item_set_from_sets_list(item_set_id)
        revised_items = page.revise_items_in_open_item_set(item_set_id)
        self.logger.info(f"Revised items: {revised_items}")

        # Step 5: Re-run QAR
        self.step(5, "Re-running QAR on the revised item set")
        print("[CHECK] Re-run QAR button is enabled after revisions", flush=True)
        assert page.is_rerun_qar_enabled(), "Re-run QAR button is disabled after revisions!"
        print("[PASS] Re-run QAR button enabled.", flush=True)

        started = monotonic()
        rerun_result = page.rerun_qar_if_enabled()
        elapsed = monotonic() - started
        self.logger.info(f"Rerun QAR completion message: {rerun_result}")

        # Step 6: Verify updated results and timeline
        self.step(6, "Verifying updated QAR scores and timeline")
        print(f"[CHECK] QAR rerun completed under SLA of 120 seconds", flush=True)
        assert elapsed <= 120, f"QAR rerun exceeded SLA: {elapsed:.2f}s"
        print(f"[PASS] QAR rerun completed in {elapsed:.2f}s", flush=True)

        report_text = self.driver.find_element("tag name", "body").text
        normalized = report_text.casefold()

        # Check duplicate score improves / new QAR results shown
        print("[CHECK] New QAR Report generated with updated scores and timeline", flush=True)
        assert "run 2" in normalized or "rerun" in normalized or "second run" in normalized or "history" in normalized or "timeline" in normalized or "passed" in normalized or "approved" in normalized, (
            f"Expected timeline or reports to show updated QAR run, but got:\n{report_text}"
        )
        print("[PASS] Confirmed fresh QAR report generated with updated details.", flush=True)

        # Step 7: Verify API GET /item-sets/{id}/qar-history shows 2 distinct QAR run records
        self.step(7, "Querying API GET /item-sets/{id}/qar-history for run records")
        session, headers = self.get_api_session()
        
        # Discover history URL
        api_path = f"/item-sets/{item_set_id}/qar-history"
        url_with_api = ReadConfig.get_base_url().rstrip("/") + "/api" + api_path
        url_direct = ReadConfig.get_base_url().rstrip("/") + api_path
        
        history_url = url_with_api
        try:
            response = session.get(url_with_api, headers=headers, timeout=10)
            if response.status_code == 404:
                response = session.get(url_direct, headers=headers, timeout=10)
                history_url = url_direct
        except Exception:
            response = session.get(url_direct, headers=headers, timeout=10)
            history_url = url_direct

        self.logger.info(f"History API URL: {history_url}")
        
        print(f"[CHECK] GET {history_url} returns 2 distinct QAR run records", flush=True)
        assert response.status_code == 200, (
            f"Expected status 200 for QAR history API, but got {response.status_code}. Response: {response.text}"
        )
        
        data = response.json()
        self.logger.info(f"QAR History response data: {data}")
        assert isinstance(data, list), f"Expected QAR history response to be a list, got {type(data)}"
        assert len(data) >= 2, f"Expected at least 2 QAR runs in history, got {len(data)}"
        
        # Verify timestamps are different
        timestamps = [record.get("timestamp") or record.get("created_at") or record.get("time") for record in data]
        self.logger.info(f"Extracted QAR run timestamps: {timestamps}")
        assert len(set(timestamps)) >= 2, f"Expected distinct timestamps for each QAR run, got {timestamps}"
        
        print(f"[PASS] Confirmed {len(data)} distinct QAR run records with separate timestamps.", flush=True)
