import json

from openpyxl import load_workbook
import pytest

from pages.common.login_page import LoginPage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.element_checks import ElementChecks
from utilities.read_config import ReadConfig

# The upload-step survey runs ~30 checks against elements that paint together;
# every absent one costs its full timeout.
CHECK_TIMEOUT = 2


@pytest.mark.rtm
# Signs in as the shared teacher and SME accounts and reads their upload
# history, which another session on the same account adds to mid-test.
@pytest.mark.serial
@pytest.mark.usefixtures("setup")
class TestUploadHistoryExcelDownload:
    @staticmethod
    def get_secondary_contributor_users():
        """Every non-primary SME/Teacher account, most likely candidate first.

        This test needs an account whose upload history holds both a PASSED
        and a FAILED row. The history view shows only the most recent uploads,
        so a heavily exercised account can have its older FAILED rows pushed
        out of view entirely by a run's worth of passing uploads - checking
        one alternate per role then reports a data gap that other configured
        accounts do not have.
        """
        primary_users = {
            ReadConfig.get_sme2_username().casefold(),
            ReadConfig.get_teacher_username().casefold(),
        }
        secondary_users = []
        seen = set()
        for role in ("sme", "teacher"):
            for username in ReadConfig.get_role_usernames(role):
                key = username.casefold()
                if key in primary_users or key in seen:
                    continue
                seen.add(key)
                secondary_users.append((role, username))
        return secondary_users

    def open_base_url(self, attempts=2):
        """Navigate to the app, surviving a navigation that never settles.

        A driver.get() here can sit until the remote command times out when
        the app is slow to respond. Stopping the stalled load and retrying
        recovers the session instead of failing the whole flow.
        """
        last_error = None
        for attempt in range(attempts):
            try:
                self.driver.set_page_load_timeout(60)
                self.driver.get(ReadConfig.get_base_url())
                return
            except Exception as error:
                last_error = error
                try:
                    self.driver.execute_script("window.stop();")
                except Exception:
                    pass
            finally:
                try:
                    self.driver.set_page_load_timeout(300)
                except Exception:
                    pass
        raise last_error

    def login_and_open_upload_history(self, username, reset_session=False):
        upload_page = UploadItemFilePage(self.driver)
        if reset_session:
            upload_page.reset_browser_session_to_login()
        else:
            self.open_base_url()

        LoginPage(self.driver).login_to_application(
            username,
            ReadConfig.get_all_users_password(),
        )
        upload_page.close_popup_if_open()
        upload_page.wait_for_application_to_load()
        upload_page.open_item_creation_module()
        upload_page.open_upload_item_file_tab()
        upload_page.open_upload_step()
        return upload_page

    def survey(self, upload_page, record_property, scope):
        """Soft-check the upload step furniture.

        The history table is fetched after the step paints, so it is waited for
        rather than sampled — sampling it immediately records an absent table on
        an account that actually has one.
        """
        checks = ElementChecks(
            upload_page, record_property, page_name=f"Upload Item File — {scope}"
        )

        checks.check("Workspace title", upload_page.WORKSPACE_TITLE, timeout=CHECK_TIMEOUT)
        checks.check("Button — Back", upload_page.BACK_BUTTON, timeout=CHECK_TIMEOUT)
        checks.check(
            "Section — Upload Documents",
            upload_page.UPLOAD_DOCUMENTS_HEADING,
            timeout=CHECK_TIMEOUT,
        )
        checks.check("Drag and drop zone", upload_page.DROPZONE, timeout=CHECK_TIMEOUT)
        checks.check(
            "Button — Add items Individually",
            upload_page.ADD_ITEMS_INDIVIDUALLY_BTN,
            timeout=CHECK_TIMEOUT,
        )
        checks.check(
            "Link — download the template",
            upload_page.INLINE_TEMPLATE_LINK,
            timeout=CHECK_TIMEOUT,
        )
        checks.check(
            "Section — Upload Prerequisites",
            upload_page.PREREQUISITES_HEADING,
            timeout=CHECK_TIMEOUT,
        )

        # The file input is deliberately hidden behind the Browse control, so a
        # visibility check would record a gap for markup that is working as
        # designed. Presence in the DOM is the meaningful check here.
        file_inputs = checks.safe_call(
            lambda: len(self.driver.find_elements(*upload_page.FILE_INPUT)), 0
        )
        checks.check_condition(
            "File input present (visually hidden by design)",
            file_inputs,
            detail=f"{file_inputs} file inputs in the DOM",
        )

        missing_mode_tabs = checks.safe_call(upload_page.missing_mode_tabs)
        for label in upload_page.MODE_TAB_LABELS:
            checks.check_condition(f"Tab — {label}", label not in missing_mode_tabs)

        missing_steps = checks.safe_call(upload_page.missing_wizard_steps)
        for label in upload_page.WIZARD_STEP_LABELS:
            checks.check_condition(f"Wizard step — {label}", label not in missing_steps)

        missing_prereq_tabs = checks.safe_call(upload_page.missing_prerequisite_tabs)
        for label in upload_page.PREREQUISITE_TAB_LABELS:
            checks.check_condition(
                f"Prerequisite tab — {label}", label not in missing_prereq_tabs
            )

        # --- Previously Uploaded Files -------------------------------------
        checks.check(
            "Section — Previously Uploaded Files",
            upload_page.UPLOAD_HISTORY_HEADING,
            timeout=CHECK_TIMEOUT,
        )
        history_loaded = checks.safe_call(
            lambda: upload_page.wait_for_upload_history(timeout=60), False
        )
        checks.check_condition(
            "Upload history finished loading",
            history_loaded,
            detail="still showing the loading placeholder" if not history_loaded else "",
        )
        checks.check(
            "Upload history table", upload_page.UPLOAD_HISTORY_TABLE, timeout=CHECK_TIMEOUT
        )

        missing_columns = checks.safe_call(upload_page.missing_upload_history_columns)
        headers = checks.safe_call(upload_page.get_upload_history_headers)
        for column in upload_page.UPLOAD_HISTORY_COLUMNS:
            checks.check_condition(
                f"Column — {column}",
                column not in missing_columns,
                detail=f"found: {headers}" if column in missing_columns else "",
            )

        row_count = checks.safe_call(upload_page.get_upload_history_row_count, 0)
        checks.check_condition(
            "Upload history has rows", row_count, detail=f"{row_count} rows"
        )

        # --- Application chrome ---------------------------------------------
        checks.check("Header bar", upload_page.HEADER, timeout=CHECK_TIMEOUT)
        checks.check("Sidebar nav", upload_page.SIDEBAR_NAV, timeout=CHECK_TIMEOUT)
        checks.check(
            "Notification bell", upload_page.NOTIFICATION_BELL, timeout=CHECK_TIMEOUT
        )
        checks.check("Theme picker", upload_page.THEME_PICKER, timeout=CHECK_TIMEOUT)
        checks.check(
            "Screen-reader toggle",
            upload_page.SCREEN_READER_TOGGLE,
            timeout=CHECK_TIMEOUT,
        )
        checks.check("Language — EN", upload_page.LANG_EN, timeout=CHECK_TIMEOUT)
        checks.check("Language — हिंदी", upload_page.LANG_HI, timeout=CHECK_TIMEOUT)
        return checks

    @staticmethod
    def verify_excel_workbook(downloaded_file):
        assert downloaded_file.exists(), f"Downloaded file is missing: {downloaded_file}"
        assert downloaded_file.stat().st_size > 0, (
            f"Downloaded file is empty: {downloaded_file}"
        )
        assert downloaded_file.suffix.casefold() == ".xlsx", (
            f"Expected an .xlsx download, received: {downloaded_file.name}"
        )

        workbook = load_workbook(downloaded_file, read_only=True, data_only=False)
        try:
            worksheet = workbook.active
            assert worksheet.max_row >= 1
            assert worksheet.max_column >= 1
        finally:
            workbook.close()

    def test_passed_and_failed_upload_history_files_download_as_excel(
        self,
        tmp_path,
        request,
        record_property,
    ):
        """A PASSED and a FAILED upload-history row each download as Excel.

        The upload step is surveyed softly. Everything about the downloads —
        that both files arrive, open as workbooks, and are distinct files — is
        data integrity and stays hard.
        """
        contributors = self.get_secondary_contributor_users()
        assert contributors, "No secondary SME or Teacher user is configured."

        checked_accounts = []
        selected_account = None
        upload_page = None
        for index, (role, username) in enumerate(contributors):
            try:
                candidate_page = self.login_and_open_upload_history(
                    username,
                    reset_session=index > 0,
                )
            except Exception as error:
                # Not every configured account is necessarily usable in this
                # environment (a role list can name a login that does not
                # exist or does not share the common password). That is a
                # provisioning gap in one candidate, not a failure of the
                # download behaviour under test - record it and try the next.
                checked_accounts.append(
                    f"{role}:{username}=unavailable ({type(error).__name__})"
                )
                continue
            upload_page = candidate_page
            statuses = upload_page.get_upload_history_statuses()
            checked_accounts.append(f"{role}:{username}={','.join(statuses) or 'none'}")
            if {"PASSED", "FAILED"}.issubset(set(statuses)):
                selected_account = (role, username)
                break

        # Surveyed on whichever account the search settled on, so the element
        # table describes the page the downloads were actually taken from.
        survey_summary = ""
        if upload_page is not None:
            checks = self.survey(upload_page, record_property, "History")
            checks.check_interaction(
                "Prerequisite tab responds — Typology",
                lambda: upload_page.switch_tab("Typology"),
                lambda: upload_page.is_tab_active("Typology"),
            )
            checks.check_interaction(
                "Prerequisite tab responds — Validation Rules",
                lambda: upload_page.switch_tab("Validation Rules"),
                lambda: upload_page.is_tab_active("Validation Rules"),
            )
            checks.check_interaction(
                "Prerequisite tab responds — Required Columns",
                lambda: upload_page.switch_tab("Required Columns"),
                lambda: upload_page.is_tab_active("Required Columns"),
            )
            # Kept for the closing result_description rather than recorded now:
            # conftest collapses user_properties with dict(), so writing the key
            # twice would drop whichever message came first.
            survey_summary = checks.publish()

        assert selected_account is not None, (
            "A secondary SME/Teacher account must have both PASSED and FAILED "
            "upload-history rows. Checked: " + "; ".join(checked_accounts)
        )

        role, username = selected_account
        request.node.user_properties.append(("download_account", f"{role}:{username}"))
        request.node.user_properties.append(("upload_history", "; ".join(checked_accounts)))

        passed_status_screenshot = upload_page.capture_upload_history_status_screenshot(
            "PASSED",
            f"{request.node.name}_passed_file_status",
            action_text="Download File",
        )
        failed_status_screenshot = upload_page.capture_upload_history_status_screenshot(
            "FAILED",
            f"{request.node.name}_failed_file_status",
            action_text="Download Annotated File",
        )
        request.node.user_properties.append(
            (
                "evidence_screenshots",
                json.dumps(
                    [
                        {
                            "name": "PASSED file status and Download File action",
                            "path": passed_status_screenshot,
                        },
                        {
                            "name": "FAILED file status and Download Annotated File action",
                            "path": failed_status_screenshot,
                        },
                    ]
                ),
            )
        )

        passed_file = upload_page.download_upload_history_file(
            "PASSED",
            tmp_path / "passed",
        )
        self.verify_excel_workbook(passed_file)

        failed_file = upload_page.download_upload_history_file(
            "FAILED",
            tmp_path / "failed",
        )
        self.verify_excel_workbook(failed_file)

        assert passed_file.resolve() != failed_file.resolve()
        request.node.user_properties.append(("passed_download", str(passed_file)))
        request.node.user_properties.append(("failed_download", str(failed_file)))
        request.node.user_properties.append(
            (
                "result_description",
                f"Downloaded and opened PASSED and FAILED Excel files as {role} {username}."
                + (f" {survey_summary}" if survey_summary else ""),
            )
        )
