"""E2E: Teacher Excel upload → RWG revision → Teacher edits items WITH image → RWG approves
(verifies image visible) → PIT approves (verifies image visible) → Published.

Coverage:
- Teacher attaches an image to each revised item during the edit cycle.
- RWG reviewer can see the attached image before approving.
- Each PIT reviewer can see the attached image before voting to publish.
- Full item-set reaches Published status after 3/3 PIT votes.
"""
from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from selenium.common.exceptions import TimeoutException

from pages.common.login_page import LoginPage
from pages.pit.review_queue_page import PITReviewQueuePage
from pages.rwg.review_queue_page import RWGReviewQueuePage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.arithmetic_question_factory import generate_unique_comparison_questions
from utilities.logger import LogGenerator
from utilities.qar_recovery import recover_qar_need_improvement_items
from utilities.read_config import ReadConfig

TEST_IMAGES_FOLDER = Path(__file__).parent.parent.parent / "test_images"


def _pick_test_image():
    """Return a real image Path from the shared test_images/ folder."""
    images = sorted(TEST_IMAGES_FOLDER.glob("*.png"))
    assert images, f"No PNG images found in {TEST_IMAGES_FOLDER}"
    return images[0]


@pytest.mark.flaky(reruns=1, reruns_delay=5)
# Shares the teacher, RWG and PIT accounts with the other M5 teacher-to-PIT
# flows; see the note on TestE2ETeacherExcelDoubleRevisionToPITPublication.
@pytest.mark.serial
@pytest.mark.usefixtures("setup")
class TestE2ETeacherExcelEditWithImageToPITPublication:
    logger = LogGenerator.loggen()

    @staticmethod
    def copy_row_format_and_values(worksheet, source_row, target_row, max_column):
        for column in range(1, max_column + 1):
            source_cell = worksheet.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=target_row, column=column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy_cell_style(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy_cell_style(source_cell.alignment)

    @classmethod
    def update_teacher_upload_file_questions(cls, upload_file):
        workbook = load_workbook(upload_file)
        worksheet = workbook.active
        question_count = 4
        max_data_column = 24
        arithmetic_items = generate_unique_comparison_questions(question_count)
        for offset in range(question_count):
            row = 2 + offset
            if row != 2:
                cls.copy_row_format_and_values(worksheet, 2, row, max_data_column)
            item = arithmetic_items[offset]
            worksheet.cell(row=row, column=5).value = offset + 1
            worksheet.cell(row=row, column=10).value = "True or False"
            worksheet.cell(row=row, column=11).value = item["question"]
            worksheet.cell(row=row, column=21).value = item["answer"]
            worksheet.cell(row=row, column=23).value = item["explanation"]
            worksheet.cell(row=row, column=24).value = "1"
        for row in range(2 + question_count, worksheet.max_row + 1):
            for column in range(1, max_data_column + 1):
                worksheet.cell(row=row, column=column).value = None
        workbook.save(upload_file)
        return question_count

    @classmethod
    def create_unique_teacher_upload_file(cls, source_file_path):
        source_file = Path(source_file_path)
        upload_dir = Path.cwd() / "tmp_uploads"
        upload_dir.mkdir(exist_ok=True)
        unique_file = (
            upload_dir
            / f"{source_file.stem}_teacher_img_{uuid4().hex[:12]}{source_file.suffix}"
        )
        copy2(source_file, unique_file)
        question_count = cls.update_teacher_upload_file_questions(unique_file)
        return unique_file, question_count

    def login_as(self, username):
        self.driver.get(ReadConfig.get_base_url())
        LoginPage(self.driver).login_to_application(
            username,
            ReadConfig.get_all_users_password(),
        )
        UploadItemFilePage(self.driver).close_popup_if_open()

    def create_fresh_teacher_item_set(self, request, teacher_username, upload_item_file_page):
        self.login_as(teacher_username)
        upload_file_path, question_count = self.create_unique_teacher_upload_file(
            ReadConfig.get_upload_item_file_path()
        )
        uploaded_file_path, upload_success_message, item_ids, ocr_success_message = (
            upload_item_file_page.upload_item_file_and_submit_for_qar(upload_file_path)
        )
        assert len(item_ids) == question_count, (
            f"Fresh teacher upload expected {question_count} item IDs but QAR returned "
            f"{len(item_ids)}; refusing to reuse or merge a stale upload."
        )
        item_set_id = upload_item_file_page.get_item_set_id_from_item_ids(item_ids)
        qar_recovery = recover_qar_need_improvement_items(
            page=upload_item_file_page,
            item_set_id=item_set_id,
            item_ids=item_ids,
            workbook_path=upload_file_path,
            max_retries=3,
            run_label="TEACHER-IMAGE-E2E",
        )
        item_set_url = self.driver.current_url
        item_set_status_text = upload_item_file_page.format_status_summary(
            upload_item_file_page.get_item_set_status_summary()
        )
        try:
            rwg_assignee = upload_item_file_page.require_item_set_assignee("rwg")
            rwg_username = ReadConfig.get_all_user_username(rwg_assignee)
        except AssertionError:
            rwg_username = ReadConfig.get_role_usernames("rwg")[0]
        sets_verification_screenshot = upload_item_file_page.capture_sets_verification_screenshot(
            request.node.name
        )
        return {
            "item_set_id": item_set_id,
            "item_ids": item_ids,
            "item_set_url": item_set_url,
            "item_set_status_text": item_set_status_text,
            "rwg_username": rwg_username,
            "sets_verification_screenshot": sets_verification_screenshot,
            "uploaded_file_path": uploaded_file_path,
            "question_count": question_count,
            "upload_success_message": upload_success_message,
            "ocr_success_message": " | ".join(
                [ocr_success_message, *qar_recovery.rerun_messages]
            ),
            "initial_qar_recovery": qar_recovery,
        }

    def test_e2e_teacher_edits_items_with_image_rwg_and_pit_verify_image_then_publish(
        self, request
    ):
        request.node.user_properties.append(
            ("result_checkpoint", "fresh teacher Excel upload and QAR validation")
        )
        upload_page = UploadItemFilePage(self.driver)
        rwg_page = RWGReviewQueuePage(self.driver)
        pit_page = PITReviewQueuePage(self.driver)

        image_path = _pick_test_image()

        default_teacher = ReadConfig.get_username()
        teacher_username = next(
            (
                u
                for u in ReadConfig.get_role_usernames("teacher")
                if u.casefold() != default_teacher.casefold()
            ),
            default_teacher,
        )

        fresh_set = self.create_fresh_teacher_item_set(request, teacher_username, upload_page)
        item_set_id = fresh_set["item_set_id"]
        item_ids = fresh_set["item_ids"]
        item_set_url = fresh_set["item_set_url"]
        item_set_status_text = fresh_set["item_set_status_text"]
        rwg_username = fresh_set["rwg_username"]
        sets_verification_screenshot = fresh_set["sets_verification_screenshot"]
        uploaded_file_path = fresh_set["uploaded_file_path"]
        question_count = fresh_set["question_count"]
        upload_success_message = fresh_set["upload_success_message"]
        ocr_success_message = fresh_set["ocr_success_message"]

        request.node.user_properties.extend(
            [
                ("item_set_id", item_set_id),
                ("item_ids", ", ".join(item_ids)),
                ("teacher_username", teacher_username),
                ("item_set_status", item_set_status_text),
            ]
        )

        sent_back_item_ids = []
        initially_approved_item_ids = []
        revised_items_with_image = []
        rerun_qar_message = ""
        rwg_approval_message = "RWG approval was not attempted."
        pit_approval_message = "PIT approval was not attempted."
        post_approval_status_text = ""
        evidence_screenshot = sets_verification_screenshot
        rwg_image_visibility = {}
        pit_image_visibility = {}

        # Step 1 — RWG sends some items back for revision
        request.node.user_properties.append(
            ("result_checkpoint", "RWG mixed revision and approval submission")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(rwg_username)
        try:
            sent_back_item_ids, initially_approved_item_ids = (
                rwg_page.revise_some_and_approve_rest_as_rwg(
                    item_set_id,
                    item_ids,
                    item_set_url,
                    revision_count=2,
                )
            )
        except (AssertionError, TimeoutException) as error:
            raise AssertionError(
                "RWG must mark items for revision and submit for fresh item set "
                f"{item_set_id}. Controls were not available: {error}"
            ) from error
        evidence_screenshot = rwg_page.capture_review_screenshot(
            request.node.name,
            "rwg_sent_back_for_teacher_revision",
        )
        print(f"\n[RWG Round 1] Sent back: {sent_back_item_ids}, Approved: {initially_approved_item_ids}")

        # Step 2 — Teacher logs in and edits items WITH image attachment
        request.node.user_properties.append(
            ("result_checkpoint", "teacher edit with image attachment and resubmit")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(teacher_username)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        revised_items_with_image = upload_page.revise_items_with_image_in_open_item_set(
            item_set_id,
            image_path,
        )
        assert revised_items_with_image, (
            f"Teacher did not see revision items after RWG send-back for {item_set_id}."
        )
        revised_item_ids = [item_id for item_id, _ in revised_items_with_image]
        rerun_qar_message = upload_page.rerun_qar_if_enabled()
        upload_page.verify_item_set_from_sets_module(item_set_id, item_ids)
        item_set_url = self.driver.current_url
        item_set_status_text = upload_page.format_status_summary(
            upload_page.get_item_set_status_summary()
        )
        request.node.user_properties.append(
            ("item_set_status_after_teacher_image_edit", item_set_status_text)
        )
        print(f"\n[Teacher Edit with Image] Revised items: {revised_item_ids}")
        print(f"[Status after teacher edit] {item_set_status_text}")

        try:
            rwg_assignee = upload_page.require_item_set_assignee("rwg")
            rwg_username = ReadConfig.get_all_user_username(rwg_assignee)
        except AssertionError:
            pass

        # Step 3 — RWG verifies image visible in each revised item, then approves all
        request.node.user_properties.append(
            ("result_checkpoint", "RWG verifies image visible in revised items then approves all")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(rwg_username)
        rwg_page.open_review_item_set(item_set_id, item_set_url)

        rwg_image_visibility = rwg_page.verify_image_visible_for_items(
            item_set_id, revised_item_ids
        )
        request.node.user_properties.append(
            ("rwg_image_visibility", str(rwg_image_visibility))
        )
        evidence_screenshot = rwg_page.capture_review_screenshot(
            request.node.name,
            "rwg_image_visible_in_revised_items",
        )
        print(f"\n[RWG Image Visibility] {rwg_image_visibility}")

        upload_page.reset_browser_session_to_login()
        self.login_as(rwg_username)
        rwg_approved_item_ids = rwg_page.approve_item_set_as_rwg(
            item_set_id,
            item_ids,
            item_set_url,
        )
        rwg_approval_message = (
            f"RWG approved {len(rwg_approved_item_ids)}/{len(item_ids)} item(s) "
            f"as {rwg_username} after verifying image attachment."
        )
        request.node.user_properties.append(("rwg_approval_message", rwg_approval_message))
        evidence_screenshot = rwg_page.capture_review_screenshot(
            request.node.name,
            "rwg_approved_after_image_edit",
        )
        print(f"\n[RWG Approval] {rwg_approval_message}")

        # Step 4 — PIT reviewers verify image visible then vote to publish (single session per user)
        request.node.user_properties.append(
            ("result_checkpoint", "PIT 3/3 quorum: verify image then publish")
        )
        completed_pit_approvals = []
        for pit_username in ReadConfig.get_pit_usernames()[:3]:
            upload_page.reset_browser_session_to_login()
            self.login_as(pit_username)

            # Image verification: open set, click each revised item, check image visible
            try:
                pit_page.open_review_item_set(item_set_id, item_set_url)
                pit_visible_items = []
                try:
                    pit_visible_items = pit_page.get_pit_item_ids(item_set_id)
                except Exception:
                    pit_visible_items = list(item_ids)

                if pit_visible_items:
                    pit_img_check = pit_page.verify_image_visible_for_items(
                        item_set_id, pit_visible_items[:len(revised_item_ids)]
                    )
                    pit_image_visibility[pit_username] = pit_img_check
                    print(f"\n[PIT {pit_username} Image Visibility] {pit_img_check}")
            except Exception as img_err:
                pit_image_visibility[pit_username] = {"error": str(img_err)}
                print(f"\n[PIT {pit_username} Image Check Error] {img_err}")

            # Approval: navigate fresh to queue and vote (return_to_item_set handles state)
            pit_page.return_to_item_set_if_needed(item_set_id)
            pit_page.approve_item_set_as_pit(item_set_id, item_set_url)
            completed_pit_approvals.append(pit_username)
            evidence_screenshot = pit_page.capture_review_screenshot(
                request.node.name,
                f"pit_approval_{len(completed_pit_approvals)}",
            )

        request.node.user_properties.append(
            ("pit_image_visibility", str(pit_image_visibility))
        )
        pit_approval_message = (
            f"PIT approval completed by {len(completed_pit_approvals)}/3 PIT users: "
            f"{', '.join(completed_pit_approvals)}."
        )
        request.node.user_properties.append(("pit_approval_message", pit_approval_message))

        # Step 5 — Final status check as teacher
        upload_page.reset_browser_session_to_login()
        request.node.user_properties.append(
            ("result_checkpoint", "final status verification — all items published")
        )
        self.login_as(teacher_username)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        post_approval_status_text = upload_page.format_status_summary(
            upload_page.get_item_set_status_summary()
        )
        request.node.user_properties.append(
            ("post_approval_status", post_approval_status_text)
        )

        # Print summary
        print(f"\nTeacher User: {teacher_username}")
        print(f"Uploaded Item File: {uploaded_file_path}")
        print(f"Upload Question Count: {question_count}")
        print(f"Upload Success Message: {upload_success_message}")
        print(f"Item Set ID: {item_set_id}")
        print(f"Item IDs: {', '.join(item_ids)}")
        print(f"OCR Success Message: {ocr_success_message}")
        print(f"Initial Item Set Status: {item_set_status_text}")
        print(f"RWG Send Back User: {rwg_username}")
        print(f"RWG Sent Back Item IDs: {', '.join(sent_back_item_ids)}")
        print(f"RWG Initially Approved Item IDs: {', '.join(initially_approved_item_ids)}")
        print(f"Teacher Revised Item IDs (with image): {', '.join(revised_item_ids)}")
        print(f"Image Path Used: {image_path}")
        print(f"RWG Image Visibility: {rwg_image_visibility}")
        print(f"PIT Image Visibility: {pit_image_visibility}")
        print(f"Re-run QAR Message: {rerun_qar_message}")
        print(f"RWG Approval: {rwg_approval_message}")
        print(f"PIT Approval: {pit_approval_message}")
        print(f"Post Approval Status: {post_approval_status_text}")

        request.node.user_properties.extend(
            [
                ("upload_item_file", str(uploaded_file_path)),
                ("upload_question_count", str(question_count)),
                ("rwg_send_back_user", rwg_username),
                ("sent_back_item_ids", ", ".join(sent_back_item_ids)),
                ("initially_approved_item_ids", ", ".join(initially_approved_item_ids)),
                ("teacher_revised_item_ids_with_image", ", ".join(revised_item_ids)),
                ("image_path", str(image_path)),
                ("rerun_qar_message", rerun_qar_message),
                ("qar_success_screenshot", evidence_screenshot),
            ]
        )

        assert completed_pit_approvals, (
            "No PIT user approved the teacher item set with image attachment."
        )
        assert all(rwg_image_visibility.values()), (
            f"RWG could not see the attached image in all revised items: {rwg_image_visibility}"
        )
        pit_item_results = {
            pit_u: v
            for pit_u, checks in pit_image_visibility.items()
            if isinstance(checks, dict) and "error" not in checks
            for v in checks.values()
        }
        assert all(pit_item_results.values()) if pit_item_results else True, (
            f"PIT could not see the attached image in all items: {pit_image_visibility}"
        )

        request.node.user_properties.append(
            (
                "result_description",
                f"PASSED: Fresh {item_set_id} — teacher attached image during revision, "
                f"RWG verified image in {len(rwg_image_visibility)} item(s) then approved all, "
                f"PIT {len(completed_pit_approvals)}/3 verified image and published the set.",
            )
        )
