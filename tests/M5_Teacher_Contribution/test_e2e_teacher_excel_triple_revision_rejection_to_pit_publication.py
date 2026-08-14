import json
import re
from copy import copy as copy_cell_style
from datetime import datetime
from pathlib import Path
from shutil import copy2
from time import sleep
from uuid import uuid4

from openpyxl import load_workbook
import pytest
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

from pages.common.login_page import LoginPage
from pages.pit.review_queue_page import PITReviewQueuePage
from pages.rwg.review_queue_page import RWGReviewQueuePage
from pages.sme.upload_item_file_page import UploadItemFilePage
from utilities.arithmetic_question_factory import (
    MAX_MIXED_QUESTION_COUNT,
    generate_qar_ready_mixed_questions,
)
from utilities.logger import LogGenerator
from utilities.qar_recovery import recover_qar_need_improvement_items
from utilities.read_config import ReadConfig
from utilities.screenshot_utils import ScreenshotUtils


class MajorActionEvidenceMixin:
    """Capture submit/reject confirmation dialogs before they are dismissed."""

    def set_evidence_recorder(self, recorder):
        self._evidence_recorder = recorder

    def get_visible_confirmation_summary(self):
        return self.driver.execute_script(
            r"""
            const visible = element => {
                const rect = element.getBoundingClientRect();
                const style = getComputedStyle(element);
                return rect.width > 0 && rect.height > 0
                    && style.display !== 'none' && style.visibility !== 'hidden';
            };
            const selectors = [
                '[role="dialog"]', '[role="alertdialog"]', '[aria-modal="true"]',
                '[class*="modal" i]', '[class*="dialog" i]', '[class*="popup" i]'
            ];
            for (const selector of selectors) {
                for (const element of document.querySelectorAll(selector)) {
                    if (!visible(element)) continue;
                    const text = (element.innerText || element.textContent || '')
                        .replace(/\s+/g, ' ').trim();
                    if (text) return text.slice(0, 220);
                }
            }
            return '';
            """
        )

    def record_confirmation_if_visible(self, action_name):
        recorder = getattr(self, "_evidence_recorder", None)
        if not recorder:
            return ""
        try:
            summary = self.get_visible_confirmation_summary()
        except Exception:
            return ""
        if summary:
            try:
                recorder(f"{action_name} confirmation popup shown: {summary}")
            except Exception:
                pass
        return summary

    def click_required_and_confirm(self, locators, button_name, timeout=15):
        for locator in locators:
            try:
                element = self.wait_utils.until_clickable(locator, timeout=timeout)
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block: 'center'});",
                    element,
                )
                self.pause_before_action()
                try:
                    element.click()
                except Exception:
                    self.driver.execute_script("arguments[0].click();", element)
                sleep(1)
                self.record_confirmation_if_visible(button_name)
                self.confirm_if_prompted()
                return True
            except Exception:
                continue
        raise TimeoutException(
            f"{button_name} button was not available after all item actions."
        )


class TripleIterationUploadItemFilePage(MajorActionEvidenceMixin, UploadItemFilePage):
    def submit_uploaded_item_set_for_qar(self):
        self._confirmation_action_name = "Submit teacher item set for QAR"
        return super().submit_uploaded_item_set_for_qar()

    def rerun_qar_if_enabled(self):
        self._confirmation_action_name = "Teacher resubmit revised item set"
        return super().rerun_qar_if_enabled()

    def confirm_submit_if_prompted(self):
        try:
            self.wait_utils.until_condition(
                lambda driver: self.get_visible_confirmation_summary() or False,
                timeout=4,
            )
        except TimeoutException:
            pass
        self.record_confirmation_if_visible(
            getattr(self, "_confirmation_action_name", "Submit item set")
        )
        return super().confirm_submit_if_prompted()


class TripleIterationRWGReviewQueuePage(MajorActionEvidenceMixin, RWGReviewQueuePage):
    """Test-local RWG behavior for the terminal third-iteration rejection."""

    REJECT_ITEM_LOCATORS = [
        (By.XPATH, "//button[normalize-space()='Reject' and not(@disabled)]"),
        (
            By.XPATH,
            "//*[self::button or @role='button']"
            "[normalize-space()='Reject' and not(@disabled)]",
        ),
    ]
    REJECT_CONFIRM_LOCATOR = (
        By.XPATH,
        "//*[@role='dialog' or @role='alertdialog' or @aria-modal='true' "
        "or contains(translate(@class,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'modal') "
        "or contains(translate(@class,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'dialog')]"
        "//*[self::button or @role='button']"
        "[(normalize-space()='Reject' or normalize-space()='Confirm' "
        "or normalize-space()='Yes') and not(@disabled)]",
    )

    def get_reject_button(self):
        def find_open_item_reject(driver):
            return driver.execute_script(
                r"""
                const buttons = Array.from(document.querySelectorAll(
                    'button:not([disabled]), [role="button"]'
                )).filter(button => {
                    const rect = button.getBoundingClientRect();
                    const text = (button.innerText || button.textContent || '').trim();
                    return rect.width > 0
                        && rect.height > 0
                        && rect.left > window.innerWidth * 0.42
                        && /^Reject$/i.test(text)
                        && button.getAttribute('aria-disabled') !== 'true';
                });
                return buttons[0] || null;
                """
            )

        return self.wait_utils.until_condition(find_open_item_reject, timeout=30)

    def confirm_reject_if_prompted(self):
        try:
            confirm_button = self.wait_utils.until_clickable(
                self.REJECT_CONFIRM_LOCATOR,
                timeout=5,
            )
        except TimeoutException:
            return False
        self.driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});",
            confirm_button,
        )
        self.record_confirmation_if_visible("Reject item after third RWG iteration")
        try:
            confirm_button.click()
        except Exception:
            self.driver.execute_script("arguments[0].click();", confirm_button)
        return True

    def is_open_item_rejected(self, driver, starting_title):
        if not starting_title:
            return False
        return bool(
            driver.execute_script(
                r"""
                const normalize = value => (value || '')
                    .replace(/\s+/g, ' ')
                    .trim()
                    .toLowerCase();
                const title = normalize(arguments[0]);
                return Array.from(
                    document.querySelectorAll('table tbody tr, button, a, [role="button"], div')
                ).some(element => {
                    const rect = element.getBoundingClientRect();
                    const text = normalize(element.innerText || element.textContent || '');
                    const isTableRow = element.matches('table tbody tr');
                    const isLeftCard = rect.width > 80
                        && rect.width < window.innerWidth * 0.42
                        && rect.height > 30
                        && rect.height < 280
                        && rect.left < window.innerWidth * 0.42;
                    return (isTableRow || isLeftCard)
                        && text.includes(title)
                        && /\brejected\b/i.test(text);
                });
                """,
                starting_title,
            )
        )

    def reject_open_item_after_iteration_limit(self, item_id):
        starting_title = self.get_open_item_title()
        self.mark_all_criteria_no()
        reject_button = self.get_reject_button()
        self.driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});",
            reject_button,
        )
        self.pause_before_action()
        try:
            reject_button.click()
        except Exception:
            self.driver.execute_script("arguments[0].click();", reject_button)
        sleep(0.5)
        self.confirm_reject_if_prompted()
        self.wait_utils.until_condition(
            lambda driver: self.is_open_item_rejected(driver, starting_title),
            timeout=30,
        )
        return item_id

    def get_review_item_statuses(self, item_set_id):
        return self.driver.execute_script(
            r"""
            const itemSetId = arguments[0];
            const escaped = itemSetId.replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
            const itemPattern = new RegExp('(' + escaped + '-i\\d+)', 'i');
            const statusPattern = /^(Pending|Under Review|Approved|Revise|Revised|Rejected)$/i;
            const result = {};
            for (const row of document.querySelectorAll('table tbody tr')) {
                const rowText = (row.innerText || row.textContent || '').trim();
                const itemMatch = rowText.match(itemPattern);
                if (!itemMatch) continue;
                const cells = Array.from(row.querySelectorAll('td'))
                    .map(cell => (cell.innerText || cell.textContent || '').trim());
                const status = cells.find(value => statusPattern.test(value))
                    || (rowText.match(/\b(Pending|Under Review|Approved|Revise|Revised|Rejected)\b/i) || [])[1]
                    || '';
                result[itemMatch[1]] = status;
            }
            return result;
            """,
            item_set_id,
        )

    def reject_revised_and_approve_remaining_as_rwg(
        self,
        item_set_id,
        rejected_item_ids,
        approved_item_ids,
        item_set_url="",
    ):
        self.open_review_item_set(item_set_id, item_set_url)
        rejected = []
        for item_id in rejected_item_ids:
            self.return_to_item_set_if_needed(item_set_id)
            self.click_item(item_id)
            rejected.append(self.reject_open_item_after_iteration_limit(item_id))

        self.return_to_item_set_if_needed(item_set_id)
        row_statuses = self.get_review_item_statuses(item_set_id)
        pending_approved_ids = [
            item_id
            for item_id in approved_item_ids
            if row_statuses.get(item_id, "").casefold() in {"pending", "under review"}
        ]
        if pending_approved_ids:
            approved = self.approve_items_with_yes(item_set_id, pending_approved_ids)
        else:
            approved = list(approved_item_ids)
            self.click_item(approved_item_ids[0])

        self.click_required_and_confirm(
            self.SUBMIT_REVIEW_LOCATORS,
            "Submit RWG Review",
            timeout=30,
        )
        return rejected, approved


class TripleIterationPITReviewQueuePage(MajorActionEvidenceMixin, PITReviewQueuePage):
    """PIT helper that proves rejected rows never enter the actionable loop."""

    def get_pit_item_statuses(self, item_set_id):
        return self.driver.execute_script(
            r"""
            const itemSetId = arguments[0];
            const escaped = itemSetId.replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
            const itemPattern = new RegExp('(' + escaped + '-i\\d+)', 'i');
            const statusPattern = /^(Pending|Under Review|Approved|Rejected|Published)$/i;
            const result = {};
            for (const row of document.querySelectorAll('table tbody tr')) {
                const rowText = (row.innerText || row.textContent || '').trim();
                const itemMatch = rowText.match(itemPattern);
                if (!itemMatch) continue;
                const cells = Array.from(row.querySelectorAll('td'))
                    .map(cell => (cell.innerText || cell.textContent || '').trim());
                const status = cells.find(value => statusPattern.test(value))
                    || (rowText.match(/\b(Pending|Under Review|Approved|Rejected|Published)\b/i) || [])[1]
                    || '';
                result[itemMatch[1]] = status;
            }
            return result;
            """,
            item_set_id,
        )

    def approve_only_expected_items_as_pit(
        self,
        item_set_id,
        expected_approved_item_ids,
        rejected_item_ids,
        item_set_url="",
    ):
        self.open_review_item_set(item_set_id, item_set_url)
        starting_quorum_count = self.get_pit_quorum_approval_count()
        all_item_ids = self.get_pit_item_ids(item_set_id)
        pending_item_ids = self.get_pending_pit_item_ids(item_set_id)
        expected = {item_id.casefold() for item_id in expected_approved_item_ids}
        rejected = {item_id.casefold() for item_id in rejected_item_ids}
        actual_pending = {item_id.casefold() for item_id in pending_item_ids}

        assert actual_pending == expected, (
            f"PIT actionable items for {item_set_id} were {pending_item_ids}; "
            f"expected only RWG-approved items {expected_approved_item_ids}."
        )
        assert actual_pending.isdisjoint(rejected), (
            f"Rejected RWG items incorrectly became PIT-actionable: "
            f"{sorted(actual_pending.intersection(rejected))}"
        )

        row_statuses = self.get_pit_item_statuses(item_set_id)
        for rejected_item_id in rejected_item_ids:
            if rejected_item_id in all_item_ids:
                assert row_statuses.get(rejected_item_id, "").casefold() == "rejected", (
                    f"PIT displayed {rejected_item_id} with status "
                    f"{row_statuses.get(rejected_item_id)!r}, not Rejected."
                )

        for item_id in pending_item_ids:
            self.click_item(item_id)
            assert not self.is_pit_item_already_approved(item_id), (
                f"{item_id} was already actioned by this PIT user."
            )
            self.wait_for_pit_decision_panel(item_set_id)
            assert self.has_actionable_pit_vote(), (
                f"Authorise was unavailable for RWG-approved PIT item {item_id}."
            )
            self.authorise_pit_item(item_id)
            self.return_to_item_set_if_needed(item_set_id)

        self.click_item(pending_item_ids[0])
        self.submit_pit_review()
        self.wait_utils.until_condition(
            lambda driver: self.get_pit_quorum_approval_count() > starting_quorum_count
            or self.page_contains_text(driver, "Published"),
            timeout=60,
        )
        return all_item_ids, pending_item_ids, row_statuses

    def get_item_set_queue_status(self, item_set_id):
        return self.driver.execute_script(
            r"""
            const itemSetId = arguments[0].toLowerCase();
            const row = Array.from(document.querySelectorAll('table tbody tr')).find(
                candidate => (candidate.innerText || candidate.textContent || '')
                    .toLowerCase()
                    .includes(itemSetId)
            );
            if (!row) return '';
            const text = row.innerText || row.textContent || '';
            const match = text.match(/\b(Published|Pending Review|Rejected)\b/i);
            return match ? match[1] : '';
            """,
            item_set_id,
        )


@pytest.mark.flaky(reruns=1, reruns_delay=5)
# Shares the teacher, RWG and PIT accounts with the other M5 teacher-to-PIT
# flows; see the note on TestE2ETeacherExcelDoubleRevisionToPITPublication.
@pytest.mark.serial
@pytest.mark.usefixtures("setup")
class TestE2ETeacherExcelTripleRevisionRejectionToPITPublication:
    logger = LogGenerator.loggen()

    @staticmethod
    def copy_row_format_and_values(worksheet, source_row, target_row, max_column):
        for column in range(1, max_column + 1):
            source_cell = worksheet.cell(row=source_row, column=column)
            target_cell = worksheet.cell(row=target_row, column=column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy_cell_style(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy_cell_style(source_cell.alignment)

    @classmethod
    def update_teacher_upload_file_questions(cls, upload_file):
        workbook = load_workbook(upload_file)
        worksheet = workbook.active
        max_data_column = 24
        # MAX_MIXED_QUESTION_COUNT (5) pulls in every typology the factory
        # knows about - True or False, Short Answer, Fill in the Blank,
        # Very Short Answer, and MCQ - so this flow proves the triple
        # revision/rejection path is typology-agnostic, not True/False-only.
        mixed_items = generate_qar_ready_mixed_questions(MAX_MIXED_QUESTION_COUNT)
        question_count = len(mixed_items)
        for offset, item in enumerate(mixed_items):
            row = 2 + offset
            if row != 2:
                cls.copy_row_format_and_values(worksheet, 2, row, max_data_column)
            worksheet.cell(row=row, column=5).value = offset + 1
            worksheet.cell(row=row, column=10).value = item["typology"]
            worksheet.cell(row=row, column=11).value = item["question"]
            for option_offset, option_text in enumerate(item["options"]):
                worksheet.cell(row=row, column=13 + option_offset).value = option_text
            worksheet.cell(row=row, column=21).value = item["answer"]
            worksheet.cell(row=row, column=23).value = item["explanation"]
            worksheet.cell(row=row, column=24).value = item["marks"]
        for row in range(2 + question_count, worksheet.max_row + 1):
            for column in range(1, max_data_column + 1):
                worksheet.cell(row=row, column=column).value = None
        workbook.save(upload_file)
        return question_count

    @classmethod
    def create_unique_teacher_upload_file(cls, source_file_path):
        source_file = Path(source_file_path)
        upload_dir = Path.cwd() / "tmp_uploads"
        upload_dir.mkdir(exist_ok=True)
        unique_file = upload_dir / (
            f"{source_file.stem}_teacher_triple_{uuid4().hex[:12]}{source_file.suffix}"
        )
        copy2(source_file, unique_file)
        return unique_file, cls.update_teacher_upload_file_questions(unique_file)

    def login_as(self, username):
        """Log in with one clean-session retry for a stuck SPA loading shell.

        Catches broadly (not just Selenium's TimeoutException) because a
        stalled WebDriver navigation command surfaces as a raw urllib3
        transport timeout, not a Selenium exception, and would otherwise
        skip this retry entirely.
        """
        page = UploadItemFilePage(self.driver)
        last_error = None
        for attempt in range(2):
            try:
                self.driver.get(ReadConfig.get_base_url())
                LoginPage(self.driver).login_to_application(
                    username,
                    ReadConfig.get_all_users_password(),
                )
                page.wait_for_application_to_load()
                page.close_popup_if_open()
                return
            except Exception as error:
                last_error = error
                if attempt == 1:
                    break
                try:
                    self.driver.delete_all_cookies()
                    self.driver.execute_script(
                        "window.localStorage.clear(); window.sessionStorage.clear();"
                    )
                except Exception:
                    pass
        raise TimeoutException(
            f"The application did not become ready while logging in as {username}."
        ) from last_error

    def resolve_rwg_username_for_item_set(
        self, upload_page, item_set_id, preferred_username, item_set_url=""
    ):
        """Log in as whichever configured RWG user actually holds this item set.

        The teacher-side "Assigned RWG" label can go undetected before the
        detail page finishes rendering, in which case the caller falls back to
        the first configured RWG account by default. With 5 RWG accounts
        configured, that guess is frequently wrong, and the guessed-wrong
        account's review queue never receives the item set, hanging
        indefinitely. Probing each configured RWG queue via the same
        open_review_item_set navigation (which searches for the item set
        rather than reading only whatever the default queue view happens to
        render) finds the real assignee instead.
        """
        candidates = list(
            dict.fromkeys([preferred_username, *ReadConfig.get_role_usernames("rwg")])
        )
        last_error = None
        max_passes = 3
        for pass_number in range(1, max_passes + 1):
            for candidate in candidates:
                try:
                    # The login belongs inside the try: a candidate whose
                    # sign-in strands on the login form is just another
                    # candidate that did not work, so move to the next account
                    # instead of abandoning every remaining one and pass.
                    upload_page.reset_browser_session_to_login()
                    self.login_as(candidate)
                    RWGReviewQueuePage(self.driver).open_review_item_set(
                        item_set_id, item_set_url
                    )
                except TimeoutException as error:
                    last_error = error
                    continue
                return candidate
            if pass_number < max_passes:
                # None of the configured RWG accounts could see the set yet;
                # a backend QAR -> RWG handoff lag can outlast one pass over
                # every candidate, so wait and sweep again before giving up.
                sleep(10)
        raise TimeoutException(
            f"Item set {item_set_id} was not visible in any configured RWG queue "
            f"({candidates}) after {max_passes} passes."
        ) from last_error

    def create_fresh_teacher_item_set(
        self,
        request,
        teacher_username,
        upload_page,
        evidence_screenshots,
        max_upload_attempts=3,
    ):
        self.login_as(teacher_username)
        for upload_attempt in range(1, max_upload_attempts + 1):
            upload_file, question_count = self.create_unique_teacher_upload_file(
                ReadConfig.get_upload_item_file_path()
            )
            uploaded_file, upload_message = upload_page.upload_item_file_and_validate(
                upload_file
            )
            self.capture_checkpoint_evidence(
                request,
                evidence_screenshots,
                f"Teacher Excel file validation success message shown: {upload_message}",
                "teacher_file_validation_success_message",
            )
            item_ids, ocr_message = upload_page.submit_uploaded_item_set_for_qar()
            assert len(item_ids) == question_count, (
                f"Fresh upload expected {question_count} IDs but received {len(item_ids)}."
            )
            item_set_id = upload_page.get_item_set_id_from_item_ids(item_ids)
            self.capture_checkpoint_evidence(
                request,
                evidence_screenshots,
                f"QAR completion message shown for {item_set_id}: {ocr_message}",
                "teacher_qar_completion_message",
            )
            try:
                qar_recovery = recover_qar_need_improvement_items(
                    page=upload_page,
                    item_set_id=item_set_id,
                    item_ids=item_ids,
                    workbook_path=upload_file,
                    max_retries=3,
                    run_label="TEACHER-TRIPLE-E2E",
                )
            except AssertionError as error:
                if (
                    "terminal, non-editable failures" not in str(error)
                    or upload_attempt == max_upload_attempts
                ):
                    raise
                self.capture_checkpoint_evidence(
                    request,
                    evidence_screenshots,
                    f"Upload attempt {upload_attempt}/{max_upload_attempts} for "
                    f"{item_set_id} hit a terminal QAR outcome ({error}); "
                    "regenerating a fresh item set and retrying.",
                    f"teacher_qar_terminal_failure_retry_{upload_attempt}",
                )
                continue
            break
        item_set_url = self.driver.current_url
        rwg_username = ReadConfig.get_role_usernames("rwg")[0]
        try:
            rwg_assignee = upload_page.require_item_set_assignee("rwg")
            rwg_username = ReadConfig.get_all_user_username(rwg_assignee)
        except AssertionError:
            pass
        return {
            "item_set_id": item_set_id,
            "item_ids": item_ids,
            "item_set_url": item_set_url,
            "rwg_username": rwg_username,
            "uploaded_file": uploaded_file,
            "question_count": question_count,
            "upload_message": upload_message,
            "ocr_message": ocr_message,
            "initial_qar_recovery": qar_recovery,
            "screenshot": upload_page.capture_sets_verification_screenshot(
                request.node.name
            ),
        }

    @staticmethod
    def assert_exact_item_ids(actual_item_ids, expected_item_ids, label):
        actual = {item_id.casefold() for item_id in actual_item_ids}
        expected = {item_id.casefold() for item_id in expected_item_ids}
        assert actual == expected, (
            f"{label} item IDs were {actual_item_ids}; expected {expected_item_ids}."
        )

    def get_teacher_item_statuses(self, item_set_id):
        return self.driver.execute_script(
            r"""
            const itemSetId = arguments[0];
            const escaped = itemSetId.replace(/[.*+?^${}()|[\]\\]/g, '\\$&');
            const itemPattern = new RegExp('(' + escaped + '-i\\d+)', 'i');
            const statusPattern = /^(Pending|Under Review|Approved|Revise|Revised|Rejected|Published)$/i;
            const result = {};
            for (const row of document.querySelectorAll('table tbody tr')) {
                const rowText = (row.innerText || row.textContent || '').trim();
                const itemMatch = rowText.match(itemPattern);
                if (!itemMatch) continue;
                const cells = Array.from(row.querySelectorAll('td'))
                    .map(cell => (cell.innerText || cell.textContent || '').trim());
                result[itemMatch[1]] = cells.find(value => statusPattern.test(value)) || '';
            }
            return result;
            """,
            item_set_id,
        )

    def get_progress_bar_container(self):
        def find_progress_bar(driver):
            return driver.execute_script(
                r"""
                const visible = element => {
                    const rect = element.getBoundingClientRect();
                    const style = getComputedStyle(element);
                    return rect.width > 0
                        && rect.height > 0
                        && style.display !== 'none'
                        && style.visibility !== 'hidden';
                };
                const exactLabel = (root, label) => Array.from(
                    root.querySelectorAll('*')
                ).some(element => visible(element)
                    && (element.innerText || element.textContent || '').trim() === label);
                const qarLabel = Array.from(document.querySelectorAll('*')).find(
                    element => visible(element)
                        && (element.innerText || element.textContent || '').trim() === 'QAR'
                );
                if (!qarLabel) return null;
                const candidates = [];
                let container = qarLabel.parentElement;
                while (container && container !== document.body) {
                    const rect = container.getBoundingClientRect();
                    if (visible(container)
                        && rect.width >= 180
                        && rect.width <= 600
                        && rect.height >= 45
                        && rect.height <= 220
                        && exactLabel(container, 'QAR')
                        && exactLabel(container, 'RWG')
                        && exactLabel(container, 'PIT')) {
                        candidates.push(container);
                    }
                    container = container.parentElement;
                }
                candidates.sort((left, right) => {
                    const leftRect = left.getBoundingClientRect();
                    const rightRect = right.getBoundingClientRect();
                    return leftRect.width * leftRect.height
                        - rightRect.width * rightRect.height;
                });
                return candidates[0] || null;
                """
            )

        return WebDriverWait(self.driver, 30).until(find_progress_bar)

    def inspect_progress_bar(self):
        progress_bar = self.get_progress_bar_container()
        state = self.driver.execute_script(
            r"""
            const root = arguments[0];
            const visible = element => {
                const rect = element.getBoundingClientRect();
                const style = getComputedStyle(element);
                return rect.width > 0
                    && rect.height > 0
                    && style.display !== 'none'
                    && style.visibility !== 'hidden';
            };
            const parseRgb = value => {
                const match = String(value || '').match(
                    /rgba?\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)/i
                );
                return match
                    ? [Number(match[1]), Number(match[2]), Number(match[3])]
                    : null;
            };
            const classify = value => {
                const rgb = parseRgb(value);
                if (!rgb) return '';
                const [red, green, blue] = rgb;
                if (blue >= 180 && blue > green * 1.45 && blue > red * 2) {
                    return 'blue';
                }
                if (green >= 105 && green > red * 1.30 && green > blue * 1.08) {
                    return 'green';
                }
                const spread = Math.max(red, green, blue) - Math.min(red, green, blue);
                if (red >= 90 && red <= 225 && spread <= 45) {
                    return 'gray';
                }
                if (spread > 45) {
                    // Brand-accent (e.g. violet/purple) "in-progress, not yet
                    // confirmed" marker that doesn't fit the blue/green
                    // heuristics above; treated as equivalent to 'green'
                    // by the caller.
                    return 'purple';
                }
                return '';
            };
            const labelCenters = {};
            for (const label of ['QAR', 'RWG', 'PIT']) {
                const element = Array.from(root.querySelectorAll('*')).find(
                    candidate => visible(candidate)
                        && (candidate.innerText || candidate.textContent || '').trim() === label
                );
                if (!element) return {error: 'Missing ' + label + ' label'};
                const rect = element.getBoundingClientRect();
                labelCenters[label] = rect.left + rect.width / 2;
            }

            const candidates = [];
            for (const element of root.querySelectorAll('*')) {
                if (!visible(element)) continue;
                const rect = element.getBoundingClientRect();
                if (rect.width < 12 || rect.width > 42 || rect.height < 4 || rect.height > 20) {
                    continue;
                }
                const style = getComputedStyle(element);
                const values = [
                    style.backgroundColor,
                    style.fill,
                    element.getAttribute('fill'),
                    style.stroke,
                    element.getAttribute('stroke'),
                ];
                for (const value of values) {
                    const color = classify(value);
                    if (!color) continue;
                    candidates.push({
                        color,
                        code: value,
                        x: rect.left + rect.width / 2,
                        key: [
                            Math.round(rect.left),
                            Math.round(rect.top),
                            Math.round(rect.width),
                            Math.round(rect.height),
                            color,
                        ].join(':'),
                    });
                    break;
                }
            }

            const unique = Array.from(
                new Map(candidates.map(candidate => [candidate.key, candidate])).values()
            ).sort((left, right) => left.x - right.x);
            const stages = {QAR: [], RWG: [], PIT: []};
            for (const segment of unique) {
                const stage = Object.keys(labelCenters).sort(
                    (left, right) => Math.abs(segment.x - labelCenters[left])
                        - Math.abs(segment.x - labelCenters[right])
                )[0];
                stages[stage].push({color: segment.color, code: segment.code});
            }
            const checkMarks = root.querySelectorAll(
                'svg path, svg polyline, [data-testid*="check" i], [class*="check" i]'
            ).length;
            return {stages, checkMarks};
            """,
            progress_bar,
        )
        if state.get("error"):
            raise AssertionError(f"Progress bar could not be inspected: {state['error']}.")
        return progress_bar, state

    @staticmethod
    def progress_signature(state):
        return tuple(
            (stage, tuple(segment["color"] for segment in state["stages"][stage]))
            for stage in ("QAR", "RWG", "PIT")
        )

    @staticmethod
    def describe_progress(checkpoint, state):
        stage_details = []
        color_codes = {}
        for stage in ("QAR", "RWG", "PIT"):
            segments = state["stages"][stage]
            wording = ", ".join(
                f"{segment['color']} segment"
                for segment in segments
            )
            stage_details.append(f"{stage} shows {wording}")
            for segment in segments:
                color_codes.setdefault(segment["color"], segment["code"])
        codes = ", ".join(
            f"{color.title()} code {color_codes[color]}"
            for color in ("blue", "green", "gray")
            if color in color_codes
        )
        return f"{checkpoint}: {'; '.join(stage_details)}. {codes}."

    @staticmethod
    def sync_evidence_to_report(request, evidence_screenshots):
        request.node.user_properties[:] = [
            property_entry
            for property_entry in request.node.user_properties
            if property_entry[0] != "evidence_screenshots"
        ]
        request.node.user_properties.append(
            ("evidence_screenshots", json.dumps(evidence_screenshots))
        )

    def record_checkpoint_evidence(
        self,
        request,
        evidence_screenshots,
        checkpoint_detail,
        screenshot_path,
    ):
        one_line_detail = re.sub(r"\s+", " ", str(checkpoint_detail)).strip()
        checkpoint_name = (
            f"Checkpoint {len(evidence_screenshots) + 1:02d} - {one_line_detail}"
        )
        evidence_screenshots.append(
            {"name": checkpoint_name, "path": str(screenshot_path)}
        )
        self.sync_evidence_to_report(request, evidence_screenshots)
        return str(screenshot_path), checkpoint_name

    def capture_checkpoint_evidence(
        self,
        request,
        evidence_screenshots,
        checkpoint_detail,
        screenshot_suffix,
    ):
        screenshot_path = ScreenshotUtils.capture(
            self.driver,
            f"{request.node.name}_{screenshot_suffix}",
        )
        return self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            checkpoint_detail,
            screenshot_path,
        )

    def record_progress_bar_evidence(
        self,
        request,
        evidence_screenshots,
        checkpoint,
        expected_stage_colors,
        previous_signature=None,
    ):
        progress_bar, state = self.inspect_progress_bar()
        actual_stage_colors = {
            stage: [segment["color"] for segment in state["stages"][stage]]
            for stage in ("QAR", "RWG", "PIT")
        }

        def matches_expected(actual, expected):
            if not expected or len(actual) < len(expected):
                return False

            def satisfies(actual_color, expected_color):
                allowed = (
                    {expected_color}
                    if isinstance(expected_color, str)
                    else set(expected_color)
                )
                if actual_color in allowed:
                    return True
                # The app's "in-progress, not yet confirmed" marker renders
                # as a brand-accent color (classified 'purple') rather than
                # the plain green segment fill; treat it as an alternate
                # rendering of 'green' wherever green is an accepted color.
                return actual_color == "purple" and "green" in allowed

            # How many segments a stage draws follows the workflow's
            # configured iteration count for that stage, which is an
            # environment setting rather than something this flow drives:
            # PIT, for instance, renders two segments where it used to
            # render one. The expectation therefore describes the leading
            # segments this flow actually advances, and any extra trailing
            # segment only has to be a plain not-yet-reached 'gray' or repeat
            # the last expected color.
            trailing = expected[-1]
            trailing_options = (
                {trailing} if isinstance(trailing, str) else set(trailing)
            ) | {"gray"}
            padded = list(expected) + [
                tuple(trailing_options)
            ] * (len(actual) - len(expected))
            return all(
                satisfies(actual_color, expected_color)
                for actual_color, expected_color in zip(actual, padded)
            )

        assert all(
            matches_expected(actual_stage_colors[stage], expected_stage_colors[stage])
            for stage in ("QAR", "RWG", "PIT")
        ), (
            f"{checkpoint} progress colors were {actual_stage_colors}; "
            f"expected segment options {expected_stage_colors}."
        )
        signature = self.progress_signature(state)
        if previous_signature is not None:
            assert signature != previous_signature, (
                f"The progress bar did not change at {checkpoint}: {signature}."
            )

        screenshot_dir = Path.cwd() / "screenshots"
        screenshot_dir.mkdir(exist_ok=True)
        safe_checkpoint = re.sub(r"[^A-Za-z0-9 ]+", "", checkpoint).strip()
        timestamp = datetime.now().strftime("%Y %m %d %H %M %S")
        screenshot_path = screenshot_dir / f"{safe_checkpoint} {timestamp}.png"
        self.driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});",
            progress_bar,
        )
        progress_bar.screenshot(str(screenshot_path))
        description = self.describe_progress(checkpoint, state)
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            description,
            screenshot_path,
        )
        return signature, str(screenshot_path), description

    def test_e2e_teacher_third_rwg_iteration_rejects_items_then_pit_publishes_approved_only(
        self,
        request,
    ):
        request.node.user_properties.append(
            ("result_checkpoint", "fresh teacher Excel upload for isolated triple-iteration flow")
        )
        upload_page = TripleIterationUploadItemFilePage(self.driver)
        rwg_page = TripleIterationRWGReviewQueuePage(self.driver)
        pit_page = TripleIterationPITReviewQueuePage(self.driver)
        evidence_screenshots = []
        upload_page.set_evidence_recorder(
            lambda detail: self.capture_checkpoint_evidence(
                request,
                evidence_screenshots,
                detail,
                f"popup_{len(evidence_screenshots) + 1:02d}",
            )
        )
        rwg_page.set_evidence_recorder(
            lambda detail: self.capture_checkpoint_evidence(
                request,
                evidence_screenshots,
                detail,
                f"popup_{len(evidence_screenshots) + 1:02d}",
            )
        )
        pit_page.set_evidence_recorder(
            lambda detail: self.capture_checkpoint_evidence(
                request,
                evidence_screenshots,
                detail,
                f"popup_{len(evidence_screenshots) + 1:02d}",
            )
        )
        default_teacher = ReadConfig.get_username()
        teacher_username = next(
            (
                username
                for username in ReadConfig.get_role_usernames("teacher")
                if username.casefold() != default_teacher.casefold()
            ),
            default_teacher,
        )

        fresh_set = self.create_fresh_teacher_item_set(
            request,
            teacher_username,
            upload_page,
            evidence_screenshots,
        )
        item_set_id = fresh_set["item_set_id"]
        item_ids = fresh_set["item_ids"]
        item_set_url = fresh_set["item_set_url"]
        rwg_username = fresh_set["rwg_username"]
        evidence_screenshot = fresh_set["screenshot"]
        request.node.user_properties.extend(
            [
                ("item_set_id", item_set_id),
                ("item_ids", ", ".join(item_ids)),
                ("teacher_username", teacher_username),
            ]
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"Fresh item set {item_set_id} is visible with {len(item_ids)} items and moved "
            f"to the first RWG review bucket assigned to {rwg_username}.",
            evidence_screenshot,
        )
        progress_signature, evidence_screenshot, _ = self.record_progress_bar_evidence(
            request,
            evidence_screenshots,
            "Progress bar after QAR and before the first RWG review",
            {
                "QAR": [("blue", "green")],
                "RWG": [("green", "gray"), "gray", "gray"],
                "PIT": ["gray"],
            },
        )

        request.node.user_properties.append(
            ("result_checkpoint", "RWG iteration 1 revises two items and approves two items")
        )
        rwg_username = self.resolve_rwg_username_for_item_set(
            upload_page, item_set_id, rwg_username, item_set_url
        )
        sent_back_round_1, rwg_approved_item_ids = (
            rwg_page.revise_some_and_approve_rest_as_rwg(
                item_set_id,
                item_ids,
                item_set_url,
                revision_count=2,
            )
        )
        rejected_item_ids = list(sent_back_round_1)
        pit_allowed_item_ids = list(rwg_approved_item_ids)
        evidence_screenshot = rwg_page.capture_review_screenshot(
            request.node.name,
            "rwg_iteration_1_mixed_result",
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"RWG iteration 1 submitted: {len(sent_back_round_1)} items moved to the "
            f"Teacher revision bucket and {len(rwg_approved_item_ids)} items remained approved.",
            evidence_screenshot,
        )

        request.node.user_properties.append(
            ("result_checkpoint", "teacher edit and resubmit after RWG iteration 1")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(teacher_username)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        progress_signature, evidence_screenshot, _ = self.record_progress_bar_evidence(
            request,
            evidence_screenshots,
            "Progress bar after the first RWG review",
            {
                "QAR": [("blue", "green")],
                "RWG": [
                    ("blue", "green"),
                    ("green", "gray"),
                    "gray",
                ],
                "PIT": ["gray"],
            },
            previous_signature=progress_signature,
        )
        teacher_revised_round_1 = upload_page.revise_items_in_open_item_set(item_set_id)
        self.assert_exact_item_ids(
            teacher_revised_round_1,
            rejected_item_ids,
            "Teacher first revision",
        )
        first_resubmit_message = upload_page.rerun_qar_if_enabled()
        upload_page.verify_item_set_from_sets_module(item_set_id, item_ids)
        item_set_url = self.driver.current_url
        evidence_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_teacher_revision_1_resubmitted"
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"Teacher revision 1 saved for {len(teacher_revised_round_1)} items; "
            f"{first_resubmit_message} The set moved back to the RWG iteration 2 bucket.",
            evidence_screenshot,
        )
        try:
            rwg_username = ReadConfig.get_all_user_username(
                upload_page.require_item_set_assignee("rwg")
            )
        except AssertionError:
            pass

        request.node.user_properties.append(
            ("result_checkpoint", "RWG iteration 2 sends the same teacher items for revision")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(rwg_username)
        sent_back_round_2 = rwg_page.send_item_set_back_as_rwg(
            item_set_id,
            rejected_item_ids,
            item_set_url,
        )
        self.assert_exact_item_ids(
            sent_back_round_2,
            rejected_item_ids,
            "RWG second send-back",
        )
        evidence_screenshot = rwg_page.capture_review_screenshot(
            request.node.name,
            "rwg_iteration_2_revision",
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"RWG iteration 2 submitted: the same {len(sent_back_round_2)} items moved "
            "back to the Teacher revision bucket.",
            evidence_screenshot,
        )

        request.node.user_properties.append(
            ("result_checkpoint", "teacher edit and resubmit after RWG iteration 2")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(teacher_username)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        progress_signature, evidence_screenshot, _ = self.record_progress_bar_evidence(
            request,
            evidence_screenshots,
            "Progress bar after the second RWG review",
            {
                "QAR": [("blue", "green")],
                "RWG": [
                    ("blue", "green"),
                    ("blue", "green"),
                    ("green", "gray"),
                ],
                "PIT": ["gray"],
            },
            previous_signature=progress_signature,
        )
        teacher_revised_round_2 = upload_page.revise_items_in_open_item_set(item_set_id)
        self.assert_exact_item_ids(
            teacher_revised_round_2,
            rejected_item_ids,
            "Teacher second revision",
        )
        second_resubmit_message = upload_page.rerun_qar_if_enabled()
        upload_page.verify_item_set_from_sets_module(item_set_id, item_ids)
        item_set_url = self.driver.current_url
        evidence_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_teacher_revision_2_resubmitted"
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"Teacher revision 2 saved for {len(teacher_revised_round_2)} items; "
            f"{second_resubmit_message} The set moved back to the final RWG review bucket.",
            evidence_screenshot,
        )
        try:
            rwg_username = ReadConfig.get_all_user_username(
                upload_page.require_item_set_assignee("rwg")
            )
        except AssertionError:
            pass

        request.node.user_properties.append(
            ("result_checkpoint", "RWG iteration 3 rejects repeatedly revised teacher items")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(rwg_username)
        rwg_rejected_item_ids, final_rwg_approved_item_ids = (
            rwg_page.reject_revised_and_approve_remaining_as_rwg(
                item_set_id,
                rejected_item_ids,
                pit_allowed_item_ids,
                item_set_url,
            )
        )
        self.assert_exact_item_ids(
            rwg_rejected_item_ids,
            rejected_item_ids,
            "RWG terminal rejection",
        )
        evidence_screenshot = rwg_page.capture_review_screenshot(
            request.node.name,
            "rwg_iteration_3_rejected",
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"RWG iteration 3 submitted: {len(rwg_rejected_item_ids)} repeatedly revised "
            f"items changed to Rejected and {len(final_rwg_approved_item_ids)} approved "
            "items moved to the PIT review bucket.",
            evidence_screenshot,
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"Note: both rejected items ({', '.join(rejected_item_ids)}) are rejected "
            "together by RWG in this single iteration-3 action because they exhausted the "
            "3-iteration revision limit; this is not one item rejected directly by RWG and "
            "a separate item rejected purely for hitting the iteration limit.",
            evidence_screenshot,
        )

        request.node.user_properties.append(
            ("result_checkpoint", "4(a) verify third-iteration teacher items are Rejected by RWG")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(teacher_username)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        progress_signature, evidence_screenshot, _ = self.record_progress_bar_evidence(
            request,
            evidence_screenshots,
            "Progress bar after the third RWG review and rejection decision",
            {
                "QAR": [("blue", "green")],
                "RWG": [
                    ("blue", "green"),
                    ("blue", "green"),
                    ("blue", "green"),
                ],
                "PIT": [("green", "gray")],
            },
            previous_signature=progress_signature,
        )
        status_summary = upload_page.get_item_set_status_summary()
        teacher_statuses = self.get_teacher_item_statuses(item_set_id)
        assert int(status_summary.get("Rejected", 0)) == len(rejected_item_ids), (
            f"Expected {len(rejected_item_ids)} RWG-rejected items after iteration 3; "
            f"status was {upload_page.format_status_summary(status_summary)}."
        )
        for rejected_item_id in rejected_item_ids:
            assert teacher_statuses.get(rejected_item_id, "").casefold() == "rejected", (
                f"{rejected_item_id} was not visibly Rejected after the third RWG iteration: "
                f"{teacher_statuses}."
            )
        request.node.user_properties.append(
            ("rwg_rejected_item_ids", ", ".join(rejected_item_ids))
        )
        evidence_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_rwg_rejection_verified"
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"Teacher Sets view confirms {len(rejected_item_ids)} items are Rejected; "
            f"item-set status is {upload_page.format_status_summary(status_summary)}.",
            evidence_screenshot,
        )

        completed_pit_approvals = []
        pit_observations = []
        pit_item_set_status = ""
        for pit_index, pit_username in enumerate(ReadConfig.get_pit_usernames()[:3], start=1):
            request.node.user_properties.append(
                (
                    "result_checkpoint",
                    f"PIT {pit_index}/3 acts only on RWG-approved items; rejected IDs excluded",
                )
            )
            upload_page.reset_browser_session_to_login()
            self.login_as(pit_username)
            all_pit_ids, actionable_pit_ids, pit_statuses = (
                pit_page.approve_only_expected_items_as_pit(
                    item_set_id,
                    final_rwg_approved_item_ids,
                    rejected_item_ids,
                    item_set_url,
                )
            )
            pit_observations.append(
                f"{pit_username}: all={all_pit_ids}, actionable={actionable_pit_ids}, "
                f"statuses={pit_statuses}"
            )
            completed_pit_approvals.append(pit_username)
            pit_item_set_status = pit_page.get_item_set_queue_status(item_set_id)
            expected_pit_status = "Published" if pit_index == 3 else "Pending Review"
            assert pit_item_set_status.casefold() == expected_pit_status.casefold(), (
                f"After PIT vote {pit_index}, {item_set_id} status was "
                f"{pit_item_set_status!r}; expected {expected_pit_status!r}."
            )
            evidence_screenshot = pit_page.capture_review_screenshot(
                request.node.name,
                f"pit_{pit_index}_approved_only_non_rejected",
            )
            self.record_checkpoint_evidence(
                request,
                evidence_screenshots,
                f"PIT vote {pit_index}/3 submitted by {pit_username}: only "
                f"{len(actionable_pit_ids)} RWG-approved items were actionable; "
                f"rejected items stayed excluded and the set moved to {pit_item_set_status}.",
                evidence_screenshot,
            )

        assert len(completed_pit_approvals) == 3, (
            f"Expected PIT 3/3, completed {completed_pit_approvals}."
        )
        request.node.user_properties.append(
            ("pit_actionability", " | ".join(pit_observations))
        )

        request.node.user_properties.append(
            ("result_checkpoint", "published set retains rejected items and approved-only PIT path")
        )
        upload_page.reset_browser_session_to_login()
        self.login_as(teacher_username)
        upload_page.open_item_set_url_and_wait(item_set_url, item_set_id)
        final_summary = upload_page.get_item_set_status_summary()
        final_status_text = upload_page.format_status_summary(final_summary)
        progress_signature, evidence_screenshot, final_progress_description = (
            self.record_progress_bar_evidence(
                request,
                evidence_screenshots,
                "Progress bar after PIT publication",
                {
                    "QAR": [("blue", "green")],
                    "RWG": [
                        ("blue", "green"),
                        ("blue", "green"),
                        ("blue", "green"),
                    ],
                    "PIT": [("blue", "green")],
                },
                previous_signature=progress_signature,
            )
        )
        final_state_screenshot = upload_page.capture_sets_verification_screenshot(
            f"{request.node.name}_final_published_status"
        )
        self.record_checkpoint_evidence(
            request,
            evidence_screenshots,
            f"Final Teacher Sets view confirms publication: {final_status_text}; "
            f"{len(rejected_item_ids)} rejected items remained rejected after PIT 3/3.",
            final_state_screenshot,
        )
        evidence_screenshot = final_state_screenshot
        assert pit_item_set_status.casefold() == "published", (
            f"The PIT queue did not show {item_set_id} as Published after vote 3: "
            f"{pit_item_set_status!r}."
        )
        assert int(final_summary.get("Rejected", 0)) == len(rejected_item_ids), (
            f"Rejected count changed after PIT publication: {final_status_text}"
        )
        request.node.user_properties.extend(
            [
                ("post_approval_status", final_status_text),
                ("pit_approval_message", "Three PIT users approved only the two non-rejected items."),
                ("progress_bar_result", final_progress_description),
                ("qar_success_screenshot", evidence_screenshot),
                (
                    "result_description",
                    f"Fresh {item_set_id} passed: RWG rejected {len(rejected_item_ids)} "
                    f"teacher items at iteration 3, PIT excluded them, approved only "
                    f"{len(final_rwg_approved_item_ids)} items through 3/3, and published the set.",
                ),
            ]
        )
