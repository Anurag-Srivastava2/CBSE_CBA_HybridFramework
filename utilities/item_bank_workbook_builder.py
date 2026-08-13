from copy import copy as copy_cell_style
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utilities.question_bank_manager import get_questions
from utilities.read_config import ReadConfig


ITEM_COLUMN_COUNT = 24
DEFAULT_QUESTION_COUNT = 4


def _normalize_header(value):
    return "".join(character for character in str(value or "").casefold() if character.isalnum())


def _item_columns(worksheet):
    headers = {
        _normalize_header(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    aliases = {
        "sequence": ("sno", "serialnumber", "itemnumber"),
        "typology": ("typology", "questiontypology", "itemtype"),
        "question": ("question", "questiontext", "itemcontent"),
        "question_image": ("questionimage",),
        "option_1": ("option1", "optiona"),
        "option_2": ("option2", "optionb"),
        "option_3": ("option3", "optionc"),
        "option_4": ("option4", "optiond"),
        "image_1": ("image1",),
        "image_2": ("image2",),
        "image_3": ("image3",),
        "image_4": ("image4",),
        "answer": ("answer", "correctanswer", "answerkey"),
        "answer_image": ("answerimage",),
        "explanation": ("explanationremarks", "explanation", "rationale"),
        "marks": ("marks", "mark"),
    }
    resolved = {}
    for field, candidates in aliases.items():
        resolved[field] = next(
            (headers[candidate] for candidate in candidates if candidate in headers),
            None,
        )

    required = ("sequence", "typology", "question", "answer", "explanation", "marks")
    if any(resolved[field] is None for field in required):
        return None
    return resolved


def _copy_template_row(worksheet, source_row, target_row):
    for column in range(1, ITEM_COLUMN_COUNT + 1):
        source_cell = worksheet.cell(row=source_row, column=column)
        target_cell = worksheet.cell(row=target_row, column=column)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy_cell_style(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy_cell_style(source_cell.alignment)
        target_cell.protection = copy_cell_style(source_cell.protection)


def _format_item_area(worksheet, columns, first_row, last_row):
    column_widths = {
        "typology": 28,
        "question": 58,
        "question_image": 24,
        "option_1": 18,
        "option_2": 18,
        "option_3": 18,
        "option_4": 18,
        "image_1": 18,
        "image_2": 18,
        "image_3": 18,
        "image_4": 18,
        "answer": 18,
        "answer_image": 18,
        "explanation": 52,
        "marks": 10,
    }
    for field, width in column_widths.items():
        column = columns.get(field)
        if column is not None:
            worksheet.column_dimensions[get_column_letter(column)].width = width

    formatted_columns = [column for column in columns.values() if column is not None]
    for row in range(first_row, last_row + 1):
        worksheet.row_dimensions[row].height = max(worksheet.row_dimensions[row].height or 15, 48)
        for column in formatted_columns:
            cell = worksheet.cell(row=row, column=column)
            alignment = copy_cell_style(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment


def _write_item(worksheet, row, columns, item):
    worksheet.cell(row=row, column=columns["sequence"]).value = row - 1
    worksheet.cell(row=row, column=columns["typology"]).value = item["typology"]
    worksheet.cell(row=row, column=columns["question"]).value = item["question"]
    worksheet.cell(row=row, column=columns["answer"]).value = item["answer"]
    worksheet.cell(row=row, column=columns["explanation"]).value = item["explanation"]
    worksheet.cell(row=row, column=columns["marks"]).value = item["marks"]

    options = item.get("options", [])
    for option_index in range(4):
        column = columns.get(f"option_{option_index + 1}")
        if column is not None:
            worksheet.cell(row=row, column=column).value = (
                options[option_index] if option_index < len(options) else None
            )

    question_image = item.get("question_image")
    image_column = columns.get("question_image")
    if image_column is not None:
        worksheet.cell(row=row, column=image_column).value = (
            question_image["filename"] if question_image else None
        )
    for image_index in range(1, 5):
        option_image_column = columns.get(f"image_{image_index}")
        if option_image_column is not None:
            worksheet.cell(row=row, column=option_image_column).value = None
    answer_image_column = columns.get("answer_image")
    if answer_image_column is not None:
        worksheet.cell(row=row, column=answer_image_column).value = None


def populate_item_workbook(workbook_path, count=DEFAULT_QUESTION_COUNT, seed=None):
    """Populate every item-data worksheet and preserve template lookup ranges."""
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path)
    summaries = {}
    environment_key = ReadConfig.get_environment_key()

    for worksheet in workbook.worksheets:
        columns = _item_columns(worksheet)
        if columns is None:
            continue
        worksheet._images = []

        items = get_questions(count, env=environment_key)
        for offset, item in enumerate(items):
            row = offset + 2
            if row > 2:
                _copy_template_row(worksheet, 2, row)
            _write_item(worksheet, row, columns, item)

        _format_item_area(worksheet, columns, 2, len(items) + 1)

        first_unused_row = len(items) + 2
        for row in range(first_unused_row, worksheet.max_row + 1):
            for column in range(1, ITEM_COLUMN_COUNT + 1):
                worksheet.cell(row=row, column=column).value = None

        summaries[worksheet.title] = items

    if not summaries:
        workbook.close()
        raise ValueError("Workbook does not contain a recognized item-data worksheet.")
    workbook.save(workbook_path)

    workbook.close()
    return summaries


def build_item_workbook(template_path, output_path, count=DEFAULT_QUESTION_COUNT, seed=None):
    """Copy a template, populate it with 3-5 items per data sheet, and save it."""
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(template_path, output_path)
    summaries = populate_item_workbook(output_path, count=count, seed=seed)
    return output_path, summaries
