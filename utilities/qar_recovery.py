"""Reusable QAR feedback, correction, and bounded rerun workflow."""

from dataclasses import dataclass
from pathlib import Path
import re
from uuid import uuid4

from openpyxl import load_workbook

from utilities.question_bank_manager import QuestionBankExhaustedError, get_replacement_question
from utilities.read_config import ReadConfig


TERMINAL_QAR_FAILURES = {"rejected", "failed", "blocked"}


def _item_number(item_id):
    match = re.search(r"-i(\d+)$", str(item_id), re.IGNORECASE)
    return int(match.group(1)) if match else 1


def _workbook_item_records(workbook_path):
    workbook = load_workbook(
        Path(workbook_path),
        read_only=True,
        data_only=True,
    )
    records = {}
    item_number = 1
    for worksheet in workbook.worksheets:
        for row in range(2, worksheet.max_row + 1):
            question = str(worksheet.cell(row=row, column=11).value or "").strip()
            if not question:
                continue
            records[item_number] = {
                "typology": str(
                    worksheet.cell(row=row, column=10).value or ""
                ).strip(),
                "question": question,
                "options": [
                    str(worksheet.cell(row=row, column=column).value or "").strip()
                    for column in range(13, 17)
                ],
                "answer": str(
                    worksheet.cell(row=row, column=21).value or ""
                ).strip(),
            }
            item_number += 1
    workbook.close()
    return records


def build_workbook_qar_correction_factory(workbook_path, run_label="E2E"):
    """Build typology-preserving corrections from the exact uploaded workbook."""
    records = _workbook_item_records(workbook_path)
    correction_token = uuid4().hex[:10]

    def correction(item_id, retry_number, qar_feedback=None):
        record = records.get(_item_number(item_id))
        if not record:
            raise AssertionError(
                f"No uploaded workbook row was found for QAR item {item_id}."
            )

        feedback = qar_feedback or {}
        reason_names = set(feedback.get("failure_reasons", {}))

        if reason_names & {"Duplicate Detection", "Plagiarism Detection"}:
            try:
                replacement = get_replacement_question(
                    typology=record["typology"],
                    env=ReadConfig.get_environment_key(),
                )
                feedback_summary = ", ".join(reason_names) or feedback.get("status") or "QAR flag"
                return {
                    "question": replacement["question"],
                    "revision_note": (
                        f"Retry {retry_number}: replaced with an unused question-bank "
                        f"item of the same typology after {feedback_summary}."
                    ),
                }
            except QuestionBankExhaustedError:
                pass
            card_label = "Independent duplicate-safe practice card"
        elif reason_names & {"Grammar", "Clarity"}:
            card_label = "Clear-language practice card"
        elif reason_names & {"Hard Validation", "Metadata Alignment"}:
            card_label = "Validated Grade 1 practice card"
        else:
            card_label = "Revised Grade 1 practice card"

        typology = record["typology"].casefold()
        answer = record["answer"]
        card_id = f"{run_label}-{correction_token}-{retry_number}-{_item_number(item_id)}"
        if "true" in typology or typology in {"tof", "t/f"}:
            expects_true = answer.casefold() == "true"
            left, right = (9, 4) if expects_true else (4, 9)
            revised_question = (
                f"{card_label} {card_id}: One tray has {left} counters and "
                f"another tray has {right} counters. The first tray has more "
                "counters than the second tray. Is this statement true or false?"
            )
        elif "fill" in typology or typology in {"fitb", "fib"}:
            revised_question = (
                f"{card_label} {card_id}: The recorded answer is {answer}. "
                "Copy the recorded answer into the blank: ___."
            )
        elif "multiple choice" in typology or typology == "mcq":
            revised_question = (
                f"{card_label} {card_id}: Read the four options and select "
                "the option identified by the answer key."
            )
        elif any(
            label in typology
            for label in ("short answer", "very short", "long answer", "vsaq", "saq", "laq")
        ):
            revised_question = (
                f"{card_label} {card_id}: A schedule records an elapsed-time "
                f"answer of {answer}. What elapsed time is recorded?"
            )
        else:
            revised_question = (
                f"{card_label} {card_id}: {record['question']}"
            )

        feedback_summary = ", ".join(reason_names) or feedback.get("status") or "QAR flag"
        return {
            "question": revised_question,
            "revision_note": (
                f"Retry {retry_number}: reviewed {feedback_summary} and rewrote "
                "the item while preserving its typology and configured answer."
            ),
        }

    return correction


@dataclass(frozen=True)
class QARRecoveryResult:
    item_set_id: str
    retry_count: int
    revised_item_ids_by_retry: tuple
    rerun_messages: tuple
    final_statuses: dict


def recover_qar_need_improvement_items(
    *,
    page,
    item_set_id,
    item_ids,
    workbook_path,
    max_retries=3,
    run_label="E2E",
    correction_factory=None,
):
    """Open a set, correct every actionable QAR failure, and rerun QAR."""
    if max_retries < 1:
        raise ValueError("max_retries must be at least 1.")

    expected_item_ids = tuple(dict.fromkeys(item_ids))
    correction_factory = correction_factory or build_workbook_qar_correction_factory(
        workbook_path,
        run_label=run_label,
    )
    retry_statuses = {
        status.casefold() for status in page.QAR_RETRY_STATUS_LABELS
    }
    revised_history = []
    rerun_messages = []

    def read_item_statuses():
        """This set's item statuses, re-opening the set if they aren't shown.

        Statuses are scoped to item rows belonging to item_set_id, so a read
        that lands on a set-level table (the Sets list, upload history) comes
        back empty rather than reporting other sets' rows as this set's items.
        An empty read means the set's item table isn't on screen, so re-open
        it once before deciding anything from it.
        """
        statuses = page.get_qar_item_statuses(item_set_id)
        if statuses:
            return statuses
        page.verify_item_set_from_sets_module(item_set_id, expected_item_ids)
        statuses = page.get_qar_item_statuses(item_set_id)
        if not statuses:
            raise AssertionError(
                f"No item rows for {item_set_id} were readable after reopening "
                f"it; expected statuses for {list(expected_item_ids)}."
            )
        return statuses

    page.verify_item_set_from_sets_module(item_set_id, expected_item_ids)
    for retry_number in range(1, max_retries + 1):
        statuses = read_item_statuses()
        terminal_items = {
            item_id: status
            for item_id, status in statuses.items()
            if status.casefold() in TERMINAL_QAR_FAILURES
        }
        if terminal_items:
            raise AssertionError(
                f"QAR returned terminal, non-editable failures for {item_set_id}: "
                f"{terminal_items}."
            )

        failed_item_ids = [
            item_id
            for item_id, status in statuses.items()
            if status.casefold() in retry_statuses
        ]
        if not failed_item_ids:
            return QARRecoveryResult(
                item_set_id=item_set_id,
                retry_count=retry_number - 1,
                revised_item_ids_by_retry=tuple(revised_history),
                rerun_messages=tuple(rerun_messages),
                final_statuses=statuses,
            )

        revised_item_ids = page.revise_qar_need_improvement_items(
            item_set_id,
            failed_item_ids,
            correction_factory,
            retry_number,
        )
        revised_history.append(tuple(revised_item_ids))
        revised_keys = {
            page.compact_item_id(item_id).casefold()
            for item_id in revised_item_ids
        }
        missing_item_ids = [
            item_id
            for item_id in failed_item_ids
            if page.compact_item_id(item_id).casefold() not in revised_keys
        ]
        if missing_item_ids and retry_number == max_retries:
            raise AssertionError(
                f"QAR retry {retry_number} did not update every flagged item: "
                f"{missing_item_ids}."
            )
        if missing_item_ids:
            # An individual item's edit can fail to register (the same
            # silently-rejected save the rerun branch below describes). Retry
            # it on the next pass with a fresh correction rather than failing
            # the whole flow on the first miss - a genuinely stuck item still
            # fails, either on the last pass or at the post-loop check.
            rerun_messages.append(
                f"Retry {retry_number}: {missing_item_ids} did not take the "
                "correction; retrying them on the next pass."
            )
            continue

        rerun_message = str(page.rerun_qar_if_enabled() or "")
        if "disabled" in rerun_message.casefold() or "not available" in rerun_message.casefold():
            # Not every edited item necessarily persisted (a save can be
            # silently rejected for a specific typology). Treat this like
            # any other still-flagged item instead of failing outright -
            # the next iteration re-reads statuses and retries whatever is
            # still stuck with a fresh, uniquely-worded correction. Only
            # the post-loop check below fails the run if it never clears.
            rerun_messages.append(rerun_message)
            continue
        rerun_messages.append(rerun_message)
        page.verify_item_set_from_sets_module(item_set_id, expected_item_ids)

    final_statuses = read_item_statuses()
    remaining = {
        item_id: status
        for item_id, status in final_statuses.items()
        if status.casefold() in retry_statuses
    }
    if remaining:
        raise AssertionError(
            f"QAR items {remaining} still need improvement after "
            f"{max_retries} correction/rerun attempts."
        )
    return QARRecoveryResult(
        item_set_id=item_set_id,
        retry_count=max_retries,
        revised_item_ids_by_retry=tuple(revised_history),
        rerun_messages=tuple(rerun_messages),
        final_statuses=final_statuses,
    )
