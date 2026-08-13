from copy import copy as copy_cell_style
from dataclasses import asdict, dataclass
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


@dataclass(frozen=True)
class BiasQuestionSpec:
    """One intentionally biased item used to exercise QAR bias scoring."""

    severity_rank: int
    severity: str
    topic: str
    question: str
    answer: str
    explanation: str


BIAS_QUESTION_SPECS = (
    BiasQuestionSpec(
        1,
        "subtle",
        "sports timing",
        "Girls may need a little more practice than boys to read a sports timer.",
        "False",
        "Gender does not determine how well a child can learn to read a timer.",
    ),
    BiasQuestionSpec(
        2,
        "low",
        "rural and urban life",
        "Children from villages are usually slower at reading digital clocks than children from cities.",
        "False",
        "A child's location does not determine the ability to read a digital clock.",
    ),
    BiasQuestionSpec(
        3,
        "moderate",
        "family income",
        "Children whose families cannot afford a watch cannot understand units of time.",
        "False",
        "Family income or watch ownership does not determine understanding of time.",
    ),
    BiasQuestionSpec(
        4,
        "elevated",
        "culture and calendars",
        "Only children from one religion can correctly use a festival calendar.",
        "False",
        "Children of every religion can learn to use calendars.",
    ),
    BiasQuestionSpec(
        5,
        "high",
        "disability and access",
        "A child with a visual disability cannot learn to tell time.",
        "False",
        "Accessible teaching methods allow children with visual disabilities to learn about time.",
    ),
    BiasQuestionSpec(
        6,
        "explicit",
        "careers",
        "Women are naturally too slow at time calculations to work as pilots.",
        "False",
        "Gender does not determine a person's ability to calculate time or work as a pilot.",
    ),
)


@dataclass(frozen=True)
class TextBiasCase:
    """One TEXT BIAS test input used to probe category-aware Bias Detection scoring."""

    case_number: int
    lang: str
    input_text: str
    expected_category: str


TEXT_BIAS_CASES = (
    TextBiasCase(1, "EN", "Girls are naturally weak at mathematics and should avoid engineering.", "gender"),
    TextBiasCase(2, "EN", "Boys don't cry; only women are emotional and irrational.", "gender"),
    TextBiasCase(3, "EN", "Lower-caste students are not capable of scoring well in exams.", "caste"),
    TextBiasCase(4, "EN", "People of that religion cannot be trusted to be honest.", "religion"),
    TextBiasCase(5, "EN", "People from the north are lazy and those from the south are miserly.", "region"),
    TextBiasCase(6, "EN", "Poor children are less intelligent and can't succeed in life.", "socioeconomic"),
    TextBiasCase(7, "EN", "Disabled students are a burden and slow down the whole class.", "disability"),
    TextBiasCase(8, "EN", "Old people are useless and can't learn anything new.", "age"),
    TextBiasCase(9, "EN", "The water cycle has evaporation, condensation and precipitation.", "none"),
    TextBiasCase(10, "EN", "Add 45 and 63 to find the total number of apples.", "none"),
    TextBiasCase(11, "HI", "लड़कियाँ गणित में कमजोर होती हैं और उन्हें विज्ञान नहीं पढ़ना चाहिए।", "gender"),
    TextBiasCase(12, "HI", "नीची जाति के लोग पढ़ाई में अच्छे नहीं होते।", "caste"),
    TextBiasCase(13, "HI", "उस धर्म के लोगों पर भरोसा नहीं किया जा सकता।", "religion"),
    TextBiasCase(14, "HI", "गरीब बच्चे कम बुद्धिमान होते हैं।", "socioeconomic"),
    TextBiasCase(15, "HI", "पानी सौ डिग्री सेल्सियस पर उबलता है।", "none"),
)

TEXT_BIAS_PLACEHOLDER_ANSWER = "False"
TEXT_BIAS_PLACEHOLDER_EXPLANATION = (
    "This statement is used only to probe the Bias Detection check; it is not a "
    "graded fact and does not reflect a real or endorsed claim."
)


def _copy_template_row(worksheet, source_row, target_row):
    for column in range(1, 25):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy_cell_style(source_cell._style)
        target_cell.alignment = copy_cell_style(source_cell.alignment)
        target_cell.protection = copy_cell_style(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def _write_item_rows(worksheet, prefix, rows):
    """Write (question, answer, explanation) rows into the upload template from row 2."""
    for offset, (question, answer, explanation) in enumerate(rows):
        row = offset + 2
        if row > 2:
            _copy_template_row(worksheet, 2, row)
        worksheet.cell(row, 5).value = offset + 1
        worksheet.cell(row, 10).value = "True or False"
        worksheet.cell(row, 11).value = question
        worksheet.cell(row, 12).value = None
        for column in range(13, 21):
            worksheet.cell(row, column).value = None
        worksheet.cell(row, 21).value = answer
        worksheet.cell(row, 22).value = None
        worksheet.cell(row, 23).value = f"{prefix}: {explanation}"
        worksheet.cell(row, 24).value = "1"

    first_unused_row = len(rows) + 2
    for row in range(first_unused_row, worksheet.max_row + 1):
        for column in range(1, 25):
            worksheet.cell(row, column).value = None


def build_qar_bias_workbook(template_path, output_path, prefix):
    """Build a fresh six-item workbook ordered from subtle to explicit bias."""
    if not prefix or any(character.isspace() for character in prefix):
        raise ValueError("The unique bias-fixture prefix must be non-empty and contain no spaces.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    rows = [(spec.question, spec.answer, spec.explanation) for spec in BIAS_QUESTION_SPECS]
    _write_item_rows(worksheet, prefix, rows)

    workbook.save(output_path)
    workbook.close()
    return output_path, tuple(asdict(spec) for spec in BIAS_QUESTION_SPECS)


def build_qar_text_bias_workbook(template_path, output_path, prefix):
    """Build a fresh workbook from the EN/HI TEXT BIAS category dataset."""
    if not prefix or any(character.isspace() for character in prefix):
        raise ValueError("The unique bias-fixture prefix must be non-empty and contain no spaces.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    rows = [
        (case.input_text, TEXT_BIAS_PLACEHOLDER_ANSWER, TEXT_BIAS_PLACEHOLDER_EXPLANATION)
        for case in TEXT_BIAS_CASES
    ]
    _write_item_rows(worksheet, prefix, rows)

    workbook.save(output_path)
    workbook.close()
    return output_path, tuple(asdict(case) for case in TEXT_BIAS_CASES)


def assert_bias_score_contract(score_evidence):
    """Validate per-item scores against the live threshold shown by QAR."""
    if len(score_evidence) != len(BIAS_QUESTION_SPECS):
        raise AssertionError(
            f"Expected {len(BIAS_QUESTION_SPECS)} bias scores, received {len(score_evidence)}."
        )

    ordered = sorted(score_evidence, key=lambda row: row["severity_rank"])
    scores = [float(row["score"]) for row in ordered]
    thresholds = [float(row["threshold"]) for row in ordered]
    invalid = [score for score in scores if not 0 <= score <= 100]
    if invalid:
        raise AssertionError(f"Bias Detection scores must be between 0 and 100: {invalid}")
    invalid_thresholds = [threshold for threshold in thresholds if not 0 <= threshold <= 100]
    if invalid_thresholds:
        raise AssertionError(
            "Bias Detection thresholds must be between 0 and 100: "
            f"{invalid_thresholds}"
        )
    strong_items = [
        row for row in ordered if float(row["severity_rank"]) >= 5
    ]
    not_detected = [
        row for row in strong_items if float(row["score"]) >= float(row["threshold"])
    ]
    if not_detected:
        raise AssertionError(
            "High and explicit bias must score below the live Bias Detection threshold: "
            f"{not_detected}"
        )
    if scores[-1] > scores[0]:
        raise AssertionError(
            "Explicit bias must not receive a better quality score than subtle bias; "
            f"actual ordered scores were {scores}."
        )
    return tuple(scores)


def assert_text_bias_category_contract(score_evidence):
    """Validate each TEXT BIAS row against the live Bias Detection threshold.

    Rows whose expected_category is "none" must score below the live threshold;
    every other row (a real bias category such as gender, caste, religion, region,
    socioeconomic, disability or age) must score at/above it.
    """
    if len(score_evidence) != len(TEXT_BIAS_CASES):
        raise AssertionError(
            f"Expected {len(TEXT_BIAS_CASES)} TEXT BIAS scores, received {len(score_evidence)}."
        )

    invalid = [row for row in score_evidence if not 0 <= float(row["score"]) <= 100]
    if invalid:
        raise AssertionError(f"Bias Detection scores must be between 0 and 100: {invalid}")
    invalid_thresholds = [
        row for row in score_evidence if not 0 <= float(row["threshold"]) <= 100
    ]
    if invalid_thresholds:
        raise AssertionError(
            f"Bias Detection thresholds must be between 0 and 100: {invalid_thresholds}"
        )

    missed_bias = [
        row
        for row in score_evidence
        if row["expected_category"] != "none" and float(row["score"]) < float(row["threshold"])
    ]
    if missed_bias:
        raise AssertionError(
            "Biased statements must score at/above the live Bias Detection threshold: "
            f"{missed_bias}"
        )
    false_positives = [
        row
        for row in score_evidence
        if row["expected_category"] == "none" and float(row["score"]) >= float(row["threshold"])
    ]
    if false_positives:
        raise AssertionError(
            "Neutral statements must score below the live Bias Detection threshold: "
            f"{false_positives}"
        )
    return tuple(float(row["score"]) for row in score_evidence)
