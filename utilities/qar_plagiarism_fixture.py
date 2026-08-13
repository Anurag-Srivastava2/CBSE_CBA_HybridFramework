from copy import copy as copy_cell_style
from dataclasses import asdict, dataclass
from difflib import SequenceMatcher
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


EXPECTED_PLAGIARISM_THRESHOLD = 97.0


@dataclass(frozen=True)
class PublishedSourceItem:
    source_item_id: str
    source_pdf_page: int
    question: str
    answer: str


# Verbatim Grade 1 / Mathematics / Chapter 38 repository entries from the
# supplied item-bank-export.pdf. Keeping the chapter aligned prevents the L1
# Metadata check from short-circuiting the L3 Plagiarism check.
PUBLISHED_SOURCE_ITEMS = (
    PublishedSourceItem(
        "IS450-G1-Mathematics-Ch38-i3",
        1,
        "During the art-room inventory activity 1LPH5FTO-3, blue tray recorded 20 crayons, while green tray recorded 91 crayons. Is 20 > 91?",
        "False",
    ),
    PublishedSourceItem(
        "IS401-G1-Mathematics-Ch38-i4",
        1,
        "Is 1146 > 9330? Teacher triple iteration 4763102bec7f-4",
        "False",
    ),
    PublishedSourceItem(
        "IS401-G1-Mathematics-Ch38-i3",
        1,
        "Is 96843 > 2291? Teacher triple iteration 4763102bec7f-3",
        "True",
    ),
    PublishedSourceItem(
        "IS396-G1-Mathematics-Ch38-i2",
        2,
        "Updated revision for IS396-G1-Mathematics-Ch38-i2: Is 98765 > 12345?",
        "True",
    ),
    PublishedSourceItem(
        "IS396-G1-Mathematics-Ch38-i4",
        2,
        "Is 86980 > 88055? Teacher upload run 157566e6923a-4",
        "False",
    ),
    PublishedSourceItem(
        "IS396-G1-Mathematics-Ch38-i3",
        3,
        "Is 49534 > 36726? Teacher upload run 157566e6923a-3",
        "True",
    ),
)


def source_similarity(candidate, source):
    normalized_candidate = " ".join(str(candidate).casefold().split())
    normalized_source = " ".join(str(source).casefold().split())
    return SequenceMatcher(None, normalized_candidate, normalized_source).ratio() * 100


def _copy_template_row(worksheet, source_row, target_row):
    for column in range(1, 25):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy_cell_style(source_cell._style)
        target_cell.alignment = copy_cell_style(source_cell.alignment)
        target_cell.protection = copy_cell_style(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def build_qar_plagiarism_workbook(template_path, output_path, run_token):
    """Copy six published PDF items into a fresh upload workbook verbatim."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    evidence = []
    for offset, source in enumerate(PUBLISHED_SOURCE_ITEMS):
        row = offset + 2
        if row > 2:
            _copy_template_row(worksheet, 2, row)
        worksheet.cell(row, 5).value = offset + 1
        worksheet.cell(row, 10).value = "True or False"
        worksheet.cell(row, 11).value = source.question
        worksheet.cell(row, 12).value = None
        for column in range(13, 21):
            worksheet.cell(row, column).value = None
        worksheet.cell(row, 21).value = source.answer
        worksheet.cell(row, 22).value = None
        worksheet.cell(row, 23).value = (
            f"{run_token}: exact published-source plagiarism threshold fixture."
        )
        worksheet.cell(row, 24).value = "1"
        row_evidence = asdict(source)
        row_evidence["source_similarity"] = source_similarity(
            worksheet.cell(row, 11).value,
            source.question,
        )
        evidence.append(row_evidence)

    first_unused_row = len(PUBLISHED_SOURCE_ITEMS) + 2
    for row in range(first_unused_row, worksheet.max_row + 1):
        for column in range(1, 25):
            worksheet.cell(row, column).value = None

    workbook.save(output_path)
    workbook.close()
    return output_path, tuple(evidence)
