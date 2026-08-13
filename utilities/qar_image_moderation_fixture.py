"""Dynamic fixtures for probing the QAR Image Moderation check.

Every "positive" image is a plain, clearly-labelled synthetic placeholder
(shapes + text drawn with Pillow) -- none of them are real or realistic
depictions of nudity, gore, weapons, drugs or self-harm. Every "benign"
image is a small cartoon/icon look-alike for its category. Each test run
draws a fresh, randomized set: one of several scenario variants is chosen
per category/variant, with randomized colour and layout jitter, and the
item's question text is generated to match whichever scenario was picked --
so the question and its attached image stay related, and no two runs
produce byte-identical fixtures. The RNG is seeded from the run's own
unique prefix, so a single run's artifacts stay reproducible for debugging
even though different runs diverge.

These exist to test the upload/QAR plumbing (does an Image Moderation check
appear at all, and does BLOCK/ALLOW track a live threshold), not to
validate true detection accuracy against real sensitive imagery.
"""

import random
import re
from copy import copy as copy_cell_style
from dataclasses import asdict, dataclass, replace
from pathlib import Path
from shutil import copy2
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

CATEGORIES = ("nudity_sexual", "violence_gore", "weapon", "drugs", "self_harm")

_CATEGORY_TINT = {
    "nudity_sexual": (219, 174, 172),
    "violence_gore": (150, 30, 30),
    "weapon": (120, 120, 128),
    "drugs": (150, 110, 190),
    "self_harm": (90, 90, 150),
}

POSITIVE_LABELS = {
    "nudity_sexual": ("an explicit image", "explicit nudity content", "sexually explicit imagery"),
    "violence_gore": ("a graphic injury photo", "bloody gore content", "graphic violence imagery"),
    "weapon": ("a real firearm photo", "a real knife threat photo", "explicit weapon imagery"),
    "drugs": ("narcotics paraphernalia", "a syringe drug-use photo", "illegal drug imagery"),
    "self_harm": ("self-harm imagery", "self-injury content", "graphic self-harm imagery"),
}


@dataclass(frozen=True)
class DynamicImageCase:
    """One IMAGE MODERATION item generated for a single test run."""

    category: str
    variant: str  # "positive" (expect BLOCK) or "benign" (expect ALLOW)
    scenario_id: str
    label: str
    filename: str
    expected_outcome: str  # "BLOCK" or "ALLOW"
    question: str
    answer: str
    explanation: str


def _font(size):
    try:
        return ImageFont.truetype("arial.ttf", size)
    except OSError:
        return ImageFont.load_default()


def _jitter(rng, spread=12):
    return rng.randint(-spread, spread)


def _wrapped_text(draw, text, font, max_width):
    words = text.split()
    lines, current = [], ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if draw.textlength(candidate, font=font) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _shade(color, rng, spread=18):
    return tuple(max(0, min(255, channel + rng.randint(-spread, spread))) for channel in color)


def _draw_positive_placeholder(path, rng, category, label):
    background = _shade(_CATEGORY_TINT[category], rng)
    image = Image.new("RGB", (480, 320), background)
    draw = ImageDraw.Draw(image)
    draw.rectangle((8, 8, 471, 311), outline=(255, 255, 255), width=3)
    font = _font(20)
    lines = ["SYNTHETIC TEST PLACEHOLDER"]
    lines.extend(_wrapped_text(draw, f"category: {category} ({label})", font, 440))
    lines.append("(non-graphic, QA-only)")
    line_height = 26
    y = (320 - line_height * len(lines)) // 2
    for line in lines:
        width = draw.textlength(line, font=font)
        draw.text(((480 - width) / 2, y), line, fill=(255, 255, 255), font=font)
        y += line_height
    image.save(path)


def _caption(image, text):
    draw = ImageDraw.Draw(image)
    font = _font(16)
    width = draw.textlength(text, font=font)
    draw.text(((480 - width) / 2, 285), text, fill=(40, 40, 40), font=font)


def _draw_anatomy_diagram(path, rng, caption):
    image = Image.new("RGB", (480, 320), (235, 245, 250))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.ellipse((210 + dx, 40, 270 + dx, 100), outline=(60, 60, 60), width=4)
    draw.line((240 + dx, 100, 240 + dx, 200), fill=(60, 60, 60), width=4)
    draw.line((240 + dx, 120, 190 + dx, 170), fill=(60, 60, 60), width=4)
    draw.line((240 + dx, 120, 290 + dx, 170), fill=(60, 60, 60), width=4)
    draw.line((240 + dx, 200, 200 + dx, 270), fill=(60, 60, 60), width=4)
    draw.line((240 + dx, 200, 280 + dx, 270), fill=(60, 60, 60), width=4)
    draw.rectangle((205 + dx, 140, 275 + dx, 205), outline=(30, 120, 200), width=4)
    _caption(image, caption)
    image.save(path)


def _draw_swimwear_icon(path, rng, caption):
    image = Image.new("RGB", (480, 320), (235, 248, 250))
    draw = ImageDraw.Draw(image)
    tint = _shade((40, 130, 200), rng)
    dx = _jitter(rng)
    draw.polygon([(180 + dx, 100), (240 + dx, 140), (170 + dx, 150)], fill=tint)
    draw.polygon([(300 + dx, 100), (240 + dx, 140), (310 + dx, 150)], fill=tint)
    draw.rectangle((205 + dx, 150, 275 + dx, 190), fill=tint)
    draw.ellipse((225 + dx, 60, 255 + dx, 90), outline=(90, 60, 40), width=3)
    _caption(image, caption)
    image.save(path)


def _draw_body_outline(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 250, 245))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.rounded_rectangle((200 + dx, 60, 280 + dx, 240), radius=30, outline=(70, 70, 70), width=3)
    labels = [("Head", 70), ("Torso", 130), ("Limbs", 200)]
    font = _font(14)
    for text, y in labels:
        draw.line((285 + dx, y, 330 + dx, y), fill=(120, 120, 120), width=2)
        draw.text((335 + dx, y - 8), text, fill=(50, 50, 50), font=font)
    _caption(image, caption)
    image.save(path)


def _draw_ketchup_spill(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 250, 250))
    draw = ImageDraw.Draw(image)
    red = _shade((200, 30, 20), rng, spread=15)
    dx, dy = _jitter(rng), _jitter(rng)
    draw.ellipse((150, 90, 330, 230), outline=(120, 120, 120), width=3)
    draw.ellipse((190 + dx, 130 + dy, 260 + dx, 190 + dy), fill=red)
    draw.ellipse((230 + dx, 150 + dy, 290 + dx, 200 + dy), fill=red)
    _caption(image, caption)
    image.save(path)


def _draw_paint_splatter(path, rng, caption):
    image = Image.new("RGB", (480, 320), (245, 245, 240))
    draw = ImageDraw.Draw(image)
    red = _shade((190, 40, 40), rng, spread=20)
    draw.ellipse((130, 120, 350, 200), outline=(150, 120, 90), width=4)
    for _ in range(6):
        x = rng.randint(160, 320)
        y = rng.randint(130, 190)
        r = rng.randint(8, 22)
        draw.ellipse((x - r, y - r, x + r, y + r), fill=red)
    _caption(image, caption)
    image.save(path)


def _draw_nail_polish_spill(path, rng, caption):
    image = Image.new("RGB", (480, 320), (248, 246, 244))
    draw = ImageDraw.Draw(image)
    red = _shade((180, 20, 40), rng, spread=15)
    dx = _jitter(rng)
    draw.rectangle((210 + dx, 90, 240 + dx, 150), fill=(230, 230, 235), outline=(120, 120, 120))
    draw.polygon([(212 + dx, 90), (238 + dx, 90), (232 + dx, 70), (218 + dx, 70)], fill=(60, 60, 60))
    draw.ellipse((150 + dx, 150, 300 + dx, 200), fill=red)
    _caption(image, caption)
    image.save(path)


def _draw_kitchen_knife_recipe(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 248, 240))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.rectangle((80, 220, 400, 250), fill=(180, 140, 90))
    draw.polygon([(120 + dx, 220), (260 + dx, 220), (300 + dx, 150), (140 + dx, 150)], fill=(210, 210, 215))
    draw.rectangle((260 + dx, 200, 320 + dx, 222), fill=(90, 60, 40))
    draw.ellipse((330, 200, 360, 230), fill=_shade((80, 160, 60), rng))
    draw.ellipse((355, 195, 385, 225), fill=_shade((200, 60, 50), rng))
    _caption(image, caption)
    image.save(path)


def _draw_toy_water_gun(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 250, 250))
    draw = ImageDraw.Draw(image)
    body = _shade((60, 170, 230), rng)
    dx = _jitter(rng)
    draw.rounded_rectangle((150 + dx, 150, 320 + dx, 190), radius=15, fill=body)
    draw.rounded_rectangle((300 + dx, 130, 340 + dx, 170), radius=10, fill=body)
    draw.rectangle((330 + dx, 145, 360 + dx, 158), fill=(255, 150, 30))
    draw.rectangle((190 + dx, 190, 220 + dx, 230), fill=body)
    _caption(image, caption)
    image.save(path)


def _draw_vegetable_peeler(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 248, 240))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.rounded_rectangle((150 + dx, 150, 280 + dx, 175), radius=10, fill=(230, 230, 235))
    draw.rounded_rectangle((270 + dx, 145, 330 + dx, 180), radius=10, fill=(120, 90, 60))
    draw.ellipse((320, 130, 380, 190), fill=_shade((230, 150, 60), rng))
    _caption(image, caption)
    image.save(path)


def _draw_pill_strip(path, rng, caption):
    image = Image.new("RGB", (480, 320), (245, 245, 250))
    draw = ImageDraw.Draw(image)
    pill_color = _shade((240, 200, 90), rng, spread=25)
    for row in range(2):
        for col in range(5):
            x = 90 + col * 60
            y = 100 + row * 60
            draw.ellipse((x, y, x + 40, y + 40), fill=(230, 230, 235), outline=(150, 150, 160), width=2)
            draw.ellipse((x + 8, y + 8, x + 32, y + 32), fill=pill_color)
    _caption(image, caption)
    image.save(path)


def _draw_medicine_bottle(path, rng, caption):
    image = Image.new("RGB", (480, 320), (245, 248, 245))
    draw = ImageDraw.Draw(image)
    body = _shade((230, 235, 240), rng, spread=10)
    dx = _jitter(rng)
    draw.rounded_rectangle((200 + dx, 100, 280 + dx, 220), radius=10, fill=body, outline=(140, 140, 150), width=3)
    draw.rectangle((215 + dx, 85, 265 + dx, 105), fill=(150, 150, 160))
    draw.rectangle((210 + dx, 140, 270 + dx, 180), fill=(255, 255, 255), outline=(120, 120, 120))
    draw.line((220 + dx, 200, 260 + dx, 200), fill=(90, 160, 90), width=6)
    draw.line((240 + dx, 190, 240 + dx, 210), fill=(90, 160, 90), width=6)
    _caption(image, caption)
    image.save(path)


def _draw_multivitamin_jar(path, rng, caption):
    image = Image.new("RGB", (480, 320), (248, 248, 245))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.rounded_rectangle((180 + dx, 90, 300 + dx, 230), radius=18, outline=(120, 120, 130), width=3)
    draw.rectangle((190 + dx, 75, 290 + dx, 95), fill=(150, 150, 160))
    for _ in range(10):
        x = rng.randint(195 + dx, 285 + dx)
        y = rng.randint(110, 220)
        draw.ellipse((x, y, x + 14, y + 14), fill=_shade((240, 160, 80), rng, spread=30))
    _caption(image, caption)
    image.save(path)


def _draw_bandage(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 250, 245))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.rounded_rectangle((150 + dx, 110, 330 + dx, 190), radius=20, fill=(235, 210, 180), outline=(180, 150, 120), width=3)
    draw.rectangle((215 + dx, 130, 265 + dx, 170), fill=(255, 255, 255), outline=(200, 200, 200))
    draw.line((240 + dx, 135, 240 + dx, 165), fill=(200, 60, 50), width=4)
    draw.line((220 + dx, 150, 260 + dx, 150), fill=(200, 60, 50), width=4)
    _caption(image, caption)
    image.save(path)


def _draw_first_aid_kit(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 250, 245))
    draw = ImageDraw.Draw(image)
    dx = _jitter(rng)
    draw.rectangle((170 + dx, 110, 310 + dx, 210), fill=(255, 255, 255), outline=(150, 150, 150), width=3)
    draw.rectangle((170 + dx, 100, 310 + dx, 115), fill=(200, 30, 30))
    draw.rectangle((225 + dx, 135, 255 + dx, 185), fill=(200, 30, 30))
    draw.rectangle((205 + dx, 150, 275 + dx, 170), fill=(200, 30, 30))
    _caption(image, caption)
    image.save(path)


def _draw_elastic_bandage_roll(path, rng, caption):
    image = Image.new("RGB", (480, 320), (250, 250, 248))
    draw = ImageDraw.Draw(image)
    tan = _shade((225, 200, 170), rng, spread=15)
    dx = _jitter(rng)
    draw.ellipse((190 + dx, 110, 290 + dx, 210), fill=tan, outline=(180, 150, 120), width=3)
    draw.ellipse((225 + dx, 145, 255 + dx, 175), fill=(250, 250, 248))
    _caption(image, caption)
    image.save(path)


BENIGN_SCENARIOS = {
    "nudity_sexual": {
        "anatomy_diagram": {"caption": "labelled anatomy diagram (educational)", "drawer": _draw_anatomy_diagram},
        "swimwear_icon": {"caption": "swimwear icon (swim-class handout)", "drawer": _draw_swimwear_icon},
        "body_outline": {"caption": "labelled body-parts outline (textbook)", "drawer": _draw_body_outline},
    },
    "violence_gore": {
        "ketchup_spill": {"caption": "ketchup spill on a plate", "drawer": _draw_ketchup_spill},
        "paint_splatter": {"caption": "red paint on an art palette", "drawer": _draw_paint_splatter},
        "nail_polish_spill": {"caption": "spilled red nail polish", "drawer": _draw_nail_polish_spill},
    },
    "weapon": {
        "kitchen_knife_recipe": {"caption": "kitchen knife used in a recipe", "drawer": _draw_kitchen_knife_recipe},
        "toy_water_gun": {"caption": "brightly-coloured toy water gun", "drawer": _draw_toy_water_gun},
        "vegetable_peeler": {"caption": "vegetable peeler in the kitchen", "drawer": _draw_vegetable_peeler},
    },
    "drugs": {
        "pill_strip": {"caption": "vitamin pill strip", "drawer": _draw_pill_strip},
        "medicine_bottle": {"caption": "labelled medicine bottle", "drawer": _draw_medicine_bottle},
        "multivitamin_jar": {"caption": "jar of multivitamin tablets", "drawer": _draw_multivitamin_jar},
    },
    "self_harm": {
        "bandage": {"caption": "adhesive bandage / plaster", "drawer": _draw_bandage},
        "first_aid_kit": {"caption": "first-aid kit box", "drawer": _draw_first_aid_kit},
        "elastic_bandage_roll": {"caption": "rolled elastic support bandage", "drawer": _draw_elastic_bandage_roll},
    },
}


def _token(rng):
    return "".join(rng.choice("0123456789abcdef") for _ in range(6))


def build_dynamic_cases(rng, prefix):
    """Pick one scenario per category/variant and generate matching question text.

    Deterministic for a given rng seed, so re-running with the same prefix
    reproduces the same fixture -- but each fresh prefix (a new test run)
    draws its own random scenarios, images and question wording.
    """
    cases = []
    for category in CATEGORIES:
        label = rng.choice(POSITIVE_LABELS[category])
        token = _token(rng)
        filename = f"{category}_positive_{token}.png"
        cases.append(
            DynamicImageCase(
                category=category,
                variant="positive",
                scenario_id=label,
                label=label,
                filename=filename,
                expected_outcome="BLOCK",
                question=(
                    f"{prefix}: The attached image is a synthetic placeholder standing "
                    f"in for {label}."
                ),
                answer="True",
                explanation=(
                    f"{prefix}: This row intentionally attaches a synthetic, non-graphic "
                    f"placeholder for {label} to verify Image Moderation flags the "
                    f"{category} category."
                ),
            )
        )

        scenario_id, scenario = rng.choice(list(BENIGN_SCENARIOS[category].items()))
        token = _token(rng)
        filename = f"{category}_benign_{scenario_id}_{token}.png"
        caption = scenario["caption"]
        cases.append(
            DynamicImageCase(
                category=category,
                variant="benign",
                scenario_id=scenario_id,
                label=caption,
                filename=filename,
                expected_outcome="ALLOW",
                question=(
                    f"{prefix}: The attached image shows {caption}, not "
                    f"{category.replace('_', ' ')} content."
                ),
                answer="True",
                explanation=(
                    f"{prefix}: This row attaches a benign look-alike ({caption}) for the "
                    f"{category} category to verify Image Moderation does not flag it."
                ),
            )
        )
    return cases


def generate_image_moderation_images(output_dir, cases, rng):
    """Draw the images referenced by `cases` (from build_dynamic_cases) into output_dir."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    for case in cases:
        path = output_dir / case.filename
        if case.variant == "positive":
            _draw_positive_placeholder(path, rng, case.category, case.label)
        else:
            BENIGN_SCENARIOS[case.category][case.scenario_id]["drawer"](path, rng, case.label)
    return output_dir


REAL_IMAGE_FILENAME_PATTERN = re.compile(
    r"^(?P<category>[a-z]+)_(?P<tag>bad|ok)_(?P<description>.+)\.(jpg|jpeg|png)$",
    re.IGNORECASE,
)

# Curriculum-style True/False questions written for the specific curated
# real photos in a QA image set (see test-images.zip): one per filename, so
# the question text is actually about what the attached image shows rather
# than a generic "this is a test row" placeholder.
REAL_IMAGE_QUESTIONS = {
    "drugs_bad_cocaine.jpg": ("Cocaine is an illegal narcotic substance.", "True"),
    "drugs_ok_medicine.jpg": ("Medicines should only be taken as prescribed by a doctor.", "True"),
    "gore_ok_anatomy_model.jpg": (
        "Anatomy models are used in schools to help students learn about the human body.",
        "True",
    ),
    "gore_ok_skin_condition_foot.jpg": (
        "Doctors use photographs of skin conditions to help diagnose illnesses.",
        "True",
    ),
    "hate_bad_extremist_flags.jpg": (
        "The attached image displays flags associated with extremist and hate groups.",
        "True",
    ),
    "hate_bad_swastika.png": (
        "The swastika symbol shown was adopted by the Nazi party in the 20th century.",
        "True",
    ),
    "hate_bad_symbol_chart.jpg": (
        "The attached image is a reference chart of hate symbols used for identification purposes.",
        "True",
    ),
    "hate_ok_india_flag.jpg": (
        "The Indian national flag has three horizontal stripes of saffron, white and green.",
        "True",
    ),
    "hate_ok_om_flag.jpg": (
        "The Om symbol is considered sacred in Hindu, Buddhist and Jain traditions.",
        "True",
    ),
    "nudity_ok_bikini_sketch.jpg": (
        "Swimwear such as a bikini is designed for use in water activities.",
        "True",
    ),
    "nudity_ok_organ_swimsuit.jpg": (
        "Anatomy diagrams help students understand the organs of the human body.",
        "True",
    ),
    "nudity_ok_swimsuit.jpg": ("Swimsuits are worn for swimming and other water sports.", "True"),
    "sexual_ok_couple_hug.jpg": ("A hug is a common gesture of affection between people.", "True"),
    "sexual_ok_love_infographic.jpg": (
        "Infographics are used to present information in a visual format.",
        "True",
    ),
    "violence_bad_bloody_player.jpg": (
        "Sports injuries can sometimes result in visible bleeding.",
        "True",
    ),
    "violence_bad_girls_fight.jpg": (
        "Physical fights can lead to injuries and are discouraged in schools.",
        "True",
    ),
    "violence_ok_ketchup.jpg": ("Ketchup is a common condiment made primarily from tomatoes.", "True"),
    "violence_ok_minor_scrape.jpg": ("A minor scrape is a small injury to the skin's surface.", "True"),
    "violence_ok_red_paint.jpg": ("Red is one of the primary colours used in painting.", "True"),
    "weapon_bad_sword.jpg": (
        "A sword is a bladed weapon that was historically used in combat.",
        "True",
    ),
    "weapon_ok_kitchen_knife.jpg": (
        "A kitchen knife is a common tool used for cutting vegetables while cooking.",
        "True",
    ),
    "weapon_ok_toy_guns.jpg": (
        "The attached image shows toy guns, which are safe playthings and not real firearms.",
        "True",
    ),
}


def load_real_image_cases(source_zip_path, prefix):
    """Parse a curated <category>_<bad|ok>_<description>.<ext> image zip into cases.

    Filenames encode ground truth directly: "bad" images are expected to be
    BLOCKed by Image Moderation, "ok" images are expected to be ALLOWed.
    Each case's question is a real curriculum-style True/False item about
    what its specific attached photo shows (see REAL_IMAGE_QUESTIONS),
    falling back to a generic description-based question for any image not
    in that table.
    """
    if not prefix or any(character.isspace() for character in prefix):
        raise ValueError("The unique fixture prefix must be non-empty and contain no spaces.")

    cases = []
    with ZipFile(source_zip_path) as archive:
        names = [Path(info.filename).name for info in archive.infolist() if not info.is_dir()]

    for filename in sorted(names):
        match = REAL_IMAGE_FILENAME_PATTERN.match(filename)
        if not match:
            continue
        category = match.group("category").lower()
        tag = match.group("tag").lower()
        description = match.group("description").replace("_", " ")
        expected_outcome = "BLOCK" if tag == "bad" else "ALLOW"

        question_text, answer = REAL_IMAGE_QUESTIONS.get(
            filename,
            (f"The attached image shows {description} (category: {category}).", "True"),
        )
        cases.append(
            DynamicImageCase(
                category=category,
                variant="positive" if tag == "bad" else "benign",
                scenario_id=filename,
                label=description,
                filename=filename,
                expected_outcome=expected_outcome,
                question=f"{prefix}: {question_text}",
                answer=answer,
                explanation=(
                    f"{prefix}: Curated QA fixture image for the {category} category "
                    f"({tag}); expected Image Moderation outcome is {expected_outcome}."
                ),
            )
        )
    return cases


def build_neutral_named_images_zip(
    source_zip_path,
    cases,
    output_zip_path,
    neutral_start_index=1,
):
    """Copy the selected image bytes into one ZIP with content-neutral names.

    The returned cases contain the same neutral filenames, allowing the Excel
    fixture and ZIP to stay in lockstep while testing whether upload moderation
    uses filename text or actual image content.
    """
    source_zip_path = Path(source_zip_path)
    output_zip_path = Path(output_zip_path)
    output_zip_path.parent.mkdir(parents=True, exist_ok=True)
    renamed_cases = []

    with ZipFile(source_zip_path) as source_archive:
        source_entries = {
            Path(info.filename).name: info
            for info in source_archive.infolist()
            if not info.is_dir()
        }
        missing_images = [
            case.filename for case in cases if case.filename not in source_entries
        ]
        if missing_images:
            raise FileNotFoundError(
                f"Images referenced by the fixture were absent from {source_zip_path}: "
                f"{missing_images}"
            )

        with ZipFile(output_zip_path, "w", ZIP_DEFLATED) as output_archive:
            for index, case in enumerate(cases, start=neutral_start_index):
                suffix = Path(case.filename).suffix.casefold()
                neutral_filename = f"image_{index:03d}{suffix}"
                output_archive.writestr(
                    neutral_filename,
                    source_archive.read(source_entries[case.filename]),
                )
                renamed_cases.append(replace(case, filename=neutral_filename))

    return output_zip_path, tuple(renamed_cases)


def extract_real_images(source_zip_path, cases, output_dir):
    """Extract just the images referenced by `cases` from source_zip_path into output_dir."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    wanted = {case.filename for case in cases}
    with ZipFile(source_zip_path) as archive:
        for info in archive.infolist():
            name = Path(info.filename).name
            if name in wanted:
                with archive.open(info) as source, open(output_dir / name, "wb") as target:
                    target.write(source.read())
    return output_dir


def _copy_template_row(worksheet, source_row, target_row):
    for column in range(1, 25):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy_cell_style(source_cell._style)
        target_cell.alignment = copy_cell_style(source_cell.alignment)
        target_cell.protection = copy_cell_style(source_cell.protection)
        target_cell.number_format = source_cell.number_format


def build_image_moderation_workbook(template_path, output_path, prefix, cases):
    """Build a fresh workbook with one True/False item per case, referencing its
    image by filename in the "Question Image" column (column 12).

    "Image 1"-"Image 4" are option images and only valid for Multiple Choice
    Questions; a True/False row has no options, so its image must go in
    Question Image (or Answer Image) instead, or the upload is rejected
    ("Image 1 is an option image ... 'True or False' has no options").
    """
    if not prefix or any(character.isspace() for character in prefix):
        raise ValueError("The unique fixture prefix must be non-empty and contain no spaces.")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(template_path, output_path)
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    for offset, case in enumerate(cases):
        row = offset + 2
        if row > 2:
            _copy_template_row(worksheet, 2, row)
        worksheet.cell(row, 5).value = offset + 1
        worksheet.cell(row, 10).value = "True or False"
        worksheet.cell(row, 11).value = case.question
        worksheet.cell(row, 12).value = case.filename  # Question Image
        for column in range(13, 17):  # Option 1-4
            worksheet.cell(row, column).value = None
        for column in range(17, 21):  # Image 1-4
            worksheet.cell(row, column).value = None
        worksheet.cell(row, 21).value = case.answer
        worksheet.cell(row, 22).value = None
        worksheet.cell(row, 23).value = case.explanation
        worksheet.cell(row, 24).value = "1"

    first_unused_row = len(cases) + 2
    for row in range(first_unused_row, worksheet.max_row + 1):
        for column in range(1, 25):
            worksheet.cell(row, column).value = None

    workbook.save(output_path)
    workbook.close()
    return output_path, tuple(asdict(case) for case in cases)


def build_images_only_zip(image_dir, cases, output_zip_path):
    """Bundle just the referenced images (no workbook) into one .zip.

    The live app's upload flow is two-step: the .xlsx is uploaded and
    validated first ("Excel received. Please upload the images ZIP to
    finish creating the set."), then a second .zip containing only the
    images referenced by filename in the workbook's image columns is
    uploaded to the same dropzone to finish creating the item set.
    """
    image_dir = Path(image_dir)
    output_zip_path = Path(output_zip_path)
    output_zip_path.parent.mkdir(parents=True, exist_ok=True)

    with ZipFile(output_zip_path, "w", ZIP_DEFLATED) as archive:
        for case in cases:
            image_path = image_dir / case.filename
            archive.write(image_path, arcname=image_path.name)

    return output_zip_path


def assert_image_moderation_contract(score_evidence):
    """Validate each IMAGE MODERATION row against the live Image Moderation threshold.

    Rows expecting BLOCK must score at/above the live threshold; rows expecting
    ALLOW must score below it.
    """
    if not score_evidence:
        raise AssertionError("Expected at least one Image Moderation score, received none.")

    invalid = [row for row in score_evidence if not 0 <= float(row["score"]) <= 100]
    if invalid:
        raise AssertionError(f"Image Moderation scores must be between 0 and 100: {invalid}")
    invalid_thresholds = [
        row for row in score_evidence if not 0 <= float(row["threshold"]) <= 100
    ]
    if invalid_thresholds:
        raise AssertionError(
            f"Image Moderation thresholds must be between 0 and 100: {invalid_thresholds}"
        )

    missed_blocks = [
        row
        for row in score_evidence
        if row["expected_outcome"] == "BLOCK" and float(row["score"]) < float(row["threshold"])
    ]
    if missed_blocks:
        raise AssertionError(
            "BLOCK-expected images must score at/above the live Image Moderation threshold: "
            f"{missed_blocks}"
        )
    false_positives = [
        row
        for row in score_evidence
        if row["expected_outcome"] == "ALLOW" and float(row["score"]) >= float(row["threshold"])
    ]
    if false_positives:
        raise AssertionError(
            "ALLOW-expected images must score below the live Image Moderation threshold: "
            f"{false_positives}"
        )
    return tuple(float(row["score"]) for row in score_evidence)
