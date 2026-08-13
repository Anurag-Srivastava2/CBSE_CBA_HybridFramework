"""Normalize the CBSE RTM workbook into a version-controlled JSON manifest."""

import argparse
from datetime import date, datetime
import json
from pathlib import Path
import re

from openpyxl import load_workbook


HEADERS = {
    "tc_id": "TC ID",
    "fr_number": "FR Number",
    "feature": "Sub-Module / Feature",
    "scenario": "Test Scenario",
    "preconditions": "Pre-Conditions",
    "steps": "Test Steps",
    "expected_result": "Expected Result",
    "test_data": "Test Data",
    "priority": "Priority",
    "status": "Status",
    "business_rule": "Business Rule",
    "test_phase": "Test Phase",
}

AUTOMATION_LINKS = {
    # M4 QP Creation: test_question_paper_creation.py and
    # test_qp_builder_contracts.py were removed in favour of three focused
    # end-to-end suites (auto-generate section level, auto-generate item
    # level, manual build). Only the cases those suites genuinely assert are
    # linked below; the rest are back on the automation backlog.
    "TC-QPCM-05-P02": {
        "file": "tests/M4_QP_Creation/test_qp_autogenerate_section_preview.py",
        "test": "test_e2e_teacher_auto_generates_section_level_qp_and_previews_sets",
    },
    "TC-QPCM-06-P01": {
        "file": "tests/M4_QP_Creation/test_qp_autogenerate_section_preview.py",
        "test": "test_e2e_teacher_auto_generates_section_level_qp_and_previews_sets",
    },
    "TC-TCIB-01-P01": {
        "file": "tests/M5_Teacher_Contribution/test_teacher_manual_item_creation.py",
        "test": "test_tc_tcib_01_p01_contribution_dashboard_and_create_cta",
    },
    "TC-TCIB-01-P02": {
        "file": "tests/M5_Teacher_Contribution/test_teacher_manual_item_creation.py",
        "test": "test_tc_tcib_01_p02_dashboard_stat_counters_are_numeric",
    },
    "TC-TCIB-02-P01": {
        "file": "tests/M5_Teacher_Contribution/test_teacher_manual_item_creation.py",
        "test": "test_tc_tcib_02_p01_submit_locked_until_mandatory_item_complete",
    },
    "TC-TCIB-02-N01": {
        "file": "tests/M5_Teacher_Contribution/test_teacher_manual_item_creation.py",
        "test": "test_tc_tcib_02_n01_teacher_grade_subject_rbac_is_enforced",
    },
    # TC-TCIB-03-P01: test_tc_tcib_03_p01_teacher_item_is_created_in_ib2_not_ib1
    # was removed from the M5 suite — the Sets-module lookup after Submit-for-QAR
    # timed out and the case is back on the automation backlog.
    "TC-IBMM-01a-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-01a-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-01a-P03": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_bulk_upload_rbac.py",
        "test": "test_tc_ibmm_01a_p03_sme_sees_only_assigned_grade_subject_items",
    },
    "TC-IBMM-01a-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_bulk_upload_validation.py",
        "test": "test_tc_ibmm_01a_n01_non_xlsx_file_is_rejected",
    },
    "TC-IBMM-01a-N02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_bulk_upload_validation.py",
        "test": "test_tc_ibmm_01a_n02_modified_header_upload_fails",
    },
    "TC-IBMM-01b-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_manual_item_creation.py",
        "test": "test_sme_create_each_typology_manual_item_and_submit_for_qar_individually",
    },
    "TC-IBMM-01b-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_manual_item_validation.py",
        "test": "test_tc_ibmm_01b_p02_continue_locked_until_mandatory_item_complete",
    },
    "TC-IBMM-01b-P03": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_manual_item_validation.py",
        "test": "test_tc_ibmm_01b_p03_new_item_is_visible_in_draft_review",
    },
    "TC-IBMM-01b-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_manual_item_validation.py",
        "test": "test_tc_ibmm_01b_n01_out_of_scope_subject_is_not_available",
    },
    "TC-IBMM-01b-N02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_manual_item_validation.py",
        "test": "test_tc_ibmm_01b_n02_empty_item_content_shows_inline_error_on_blur",
    },
    "TC-IBMM-03-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_03_p01_created_items_receive_unique_12_character_ids",
    },
    "TC-IBMM-03-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-03-P03": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-03-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_bulk_upload_validation.py",
        "test": "test_tc_ibmm_03_n01_duplicate_item_content_is_rejected",
    },
    "TC-IBMM-04-M01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_04_m01_all_required_metadata_fields_are_enforced",
    },
    "TC-IBMM-04-M02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_04_m02_metadata_tags_support_exact_and_partial_search",
    },
    "TC-IBMM-04-M03": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_04_m03_invalid_stage_grade_combination_is_blocked",
    },
    "TC-IBMM-05-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_05_p01_all_controlled_typologies_are_available",
    },
    "TC-IBMM-05-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_05_p02_answer_controls_change_with_typology",
    },
    "TC-IBMM-05-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_05_n01_mcq_with_three_options_is_blocked",
    },
    "TC-IBMM-04-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_excel_template.py",
        "test": "test_tc_ibmm_04_p01_template_downloads_with_headers_and_version",
    },
    "TC-IBMM-04-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_excel_template.py",
        "test": "test_tc_ibmm_04_p02_template_has_controlled_field_validations",
    },
    "TC-IBMM-04-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_bulk_upload_validation.py",
        "test": "test_tc_ibmm_01a_n02_modified_header_upload_fails",
    },
    "TC-IBMM-04-N02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_excel_template.py",
        "test": "test_tc_ibmm_04_n02_old_template_version_shows_upgrade_notice",
    },
    "TC-IBMM-06-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_report_sections_statuses_and_sla",
    },
    "TC-IBMM-06-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_flags_paraphrase_bias_and_ambiguity",
    },
    "TC-IBMM-06-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_flags_paraphrase_bias_and_ambiguity",
    },
    "TC-IBMM-07-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_report_sections_statuses_and_sla",
    },
    "TC-IBMM-07-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_flags_paraphrase_bias_and_ambiguity",
    },
    "TC-IBMM-07-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_flags_paraphrase_bias_and_ambiguity",
    },
    "TC-IBMM-08-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_report_sections_statuses_and_sla",
    },
    "TC-IBMM-08-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_metadata_typology_ids.py",
        "test": "test_tc_ibmm_04_m03_invalid_stage_grade_combination_is_blocked",
    },
    "TC-IBMM-08-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_flags_paraphrase_bias_and_ambiguity",
    },
    "TC-IBMM-09-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_40_percent_failure_does_not_lock_set",
    },
    "TC-IBMM-09-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_70_percent_failure_provides_exception_report",
    },
    "TC-IBMM-09-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_exactly_60_percent_failure_locks_and_blocks_forwarding",
    },
    "TC-IBMM-09-N02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_exactly_60_percent_failure_locks_and_blocks_forwarding",
    },
    "TC-IBMM-10-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_report_sections_statuses_and_sla",
    },
    "TC-IBMM-10-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-10-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_flags_paraphrase_bias_and_ambiguity",
    },
    "TC-IBMM-11-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_11_p01_rwg_queue_contains_qar_cleared_items",
    },
    "TC-IBMM-11-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_11_p02_sirs_has_26_criteria_across_6_sections",
    },
    "TC-IBMM-11-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_11_n01_submit_disabled_with_zero_evaluated",
    },
    "TC-IBMM-12-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_12_p01_sme_feedback_has_iteration_counter",
    },
    "TC-IBMM-12-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-12-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_12_n01_item_disabled_after_third_rejection",
    },
    "TC-IBMM-13-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_13_p01_p02_sr_rwg_history_and_decisions",
    },
    "TC-IBMM-13-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_13_p01_p02_sr_rwg_history_and_decisions",
    },
    "TC-IBMM-13-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_review_role_contracts.py",
        "test": "test_tc_ibmm_13_n01_teacher_items_absent_from_sr_rwg_queue",
    },
    "TC-IBMM-16-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_version_rejection_contracts.py",
        "test": "test_tc_ibmm_16_p01_full_revision_history_is_visible",
    },
    "TC-IBMM-16-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_version_rejection_contracts.py",
        "test": "test_tc_ibmm_16_p02_reviewer_history_is_read_only",
    },
    "TC-IBMM-16-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-17-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_version_rejection_contracts.py",
        "test": "test_tc_ibmm_17_p01_admin_receives_three_strike_notification",
    },
    "TC-IBMM-17-P02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_version_rejection_contracts.py",
        "test": "test_tc_ibmm_17_p02_admin_can_view_rejection_history",
    },
    "TC-IBMM-17-N01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_version_rejection_contracts.py",
        "test": "test_tc_ibmm_17_n01_second_rejection_has_no_three_strike_alert",
    },
    "TC-IBMM-11-P03": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-11-N02": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
    "TC-IBMM-14-P01": {
        "file": "tests/M1_Item_Bank_Mgmt/test_e2e_sme_excel_upload_to_pit_publication.py",
        "test": "test_e2e_sme_excel_upload_qar_rwg_srrwg_pit_publish",
    },
}



AUTOMATION_LINKS.update({
    # M1 negative cases. These were only ever patched into the generated JSON,
    # so regenerating the manifest used to silently drop their automation links.
    "TC-NEG-M1-07": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_qar_rules.py",
        "test": "test_qar_exactly_60_percent_failure_locks_and_blocks_forwarding",
    },
    "TC-NEG-M1-08": {
        "file": "tests/M1_Item_Bank_Mgmt/test_sme_cross_rbac_api.py",
        "test": "test_tc_neg_m1_08_cross_sme_rbac_returns_403_forbidden",
    },
    "TC-NEG-M1-09": {
        "file": "tests/M1_Item_Bank_Mgmt/test_qar_resubmission_fresh_run.py",
        "test": "test_tc_neg_m1_09_resubmission_triggers_fresh_qar_run",
    },
    "TC-NEG-M1-10": {
        "file": "tests/M1_Item_Bank_Mgmt/test_upload_file_size_limit.py",
        "test": "test_tc_neg_m1_10_upload_file_size_limit_exceeded",
    },
    "TC-WPAD-01-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_01_p01_admin_creates_new_user_with_required_fields",
    },
    "TC-WPAD-01-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_01_p02_admin_deactivates_user_and_login_is_blocked",
    },
    "TC-WPAD-01-N01": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_01_n01_duplicate_email_is_rejected",
    },
    "TC-WPAD-01-N02": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_01_n02_duplicate_mobile_is_rejected",
    },
    "TC-WPAD-02-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_02_p01_sidebar_rbac_is_enforced_for_core_roles",
    },
    "TC-WPAD-02-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_02_p02_sme_grade_subject_restriction_is_enforced",
    },
    "TC-WPAD-02-N01": {
        "file": "tests/M2_Web_Portal_Admin/test_user_management_rbac.py",
        "test": "test_tc_wpad_02_n01_teacher_direct_admin_url_is_denied",
    },
    "TC-WPAD-03-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_03_p01_welcome_email_is_sent_within_60_seconds",
    },
    "TC-WPAD-03-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_03_p02_onboarding_link_is_single_use",
    },
    "TC-WPAD-03-N01": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_03_n01_onboarding_link_expires_after_24_hours",
    },
    "TC-WPAD-04-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_04_p01_otp_is_delivered_by_email_and_sms_within_60_seconds",
    },
    "TC-WPAD-04-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_04_p02_otp_expires_after_5_minutes",
    },
    "TC-WPAD-04-N01": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_04_n01_account_locks_after_three_invalid_otp_attempts",
    },
    "TC-WPAD-05-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_05_p01_idle_session_expires_after_10_minutes",
    },
    "TC-WPAD-05-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_05_p02_idle_warning_appears_at_8_minutes",
    },
    "TC-WPAD-05-P03": {
        "file": "tests/M2_Web_Portal_Admin/test_mfa_session_contracts.py",
        "test": "test_tc_wpad_05_p03_stay_active_resets_idle_timer",
    },
    "TC-WPAD-06-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_06_p01_admin_can_update_theme_without_deployment",
    },
    "TC-WPAD-06-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_06_p02_teacher_preview_and_publish_theme",
    },
    "TC-WPAD-07-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_07_p01_sme_dashboard_shows_item_stats_only",
    },
    "TC-WPAD-07-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_07_p02_pit_dashboard_has_quorum_without_sme_edit_widgets",
    },
    "TC-WPAD-08-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_08_p01_audit_logs_filter_by_user_date_and_action",
    },
    "TC-WPAD-08-N01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_08_n01_audit_logs_are_immutable",
    },
    "TC-WPAD-09-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_09_p01_admin_report_generates_within_5_seconds",
    },
    "TC-WPAD-09-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_09_p02_admin_report_downloads_csv_or_excel",
    },
    "TC-WPAD-10-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_10_p01_teacher_can_submit_support_ticket",
    },
    "TC-WPAD-10-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_10_p02_support_ticket_email_acknowledgement_within_48_hours",
    },
    "TC-WPAD-11-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_11_p01_new_subject_master_reflects_in_m1_and_m4",
    },
    "TC-WPAD-11-N01": {
        "file": "tests/M2_Web_Portal_Admin/test_portal_admin_features.py",
        "test": "test_tc_wpad_11_n01_delete_linked_subject_is_blocked",
    },
    "TC-WPAD-12-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_12_p01_otp_notification_email_and_sms_within_60_seconds",
    },
    "TC-WPAD-12-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_12_p02_qar_pass_notification_email_sms_and_panel",
    },
    "TC-WPAD-13-P01": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_13_p01_system_health_dashboard_shows_core_services",
    },
    "TC-WPAD-13-P02": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_13_p02_service_outage_alert_fires_within_5_minutes",
    },
    "TC-WPAD-PERF-01": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_perf_01_create_10_users_each_within_2_seconds",
    },
    "TC-WPAD-PERF-02": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_perf_02_audit_log_796_entries_filters_within_3_seconds",
    },
    "TC-WPAD-PERF-03": {
        "file": "tests/M2_Web_Portal_Admin/test_notifications_health_performance.py",
        "test": "test_tc_wpad_perf_03_100_user_load_has_no_session_or_rbac_leakage",
    },
})


# M5 Teacher Contribution: test_teacher_item_lifecycle_contracts.py was removed
# from the suite — its 25 contract/perf placeholders never created or progressed
# controlled teacher items, so they were permanently skipped. Those cases are
# back on the automation backlog.

def clean(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return " ".join(str(value).split())


def find_header_map(row):
    normalized = {clean(value): index for index, value in enumerate(row)}
    result = {}
    for key, expected in HEADERS.items():
        match = next(
            (index for label, index in normalized.items() if label.startswith(expected)),
            None,
        )
        if match is not None:
            result[key] = match
    return result


def classify(case, sheet_name):
    combined = " ".join(
        case[key]
        for key in ("tc_id", "feature", "scenario", "steps", "expected_result", "test_phase")
    ).lower()
    if "out of scope" in case["status"].lower():
        return "out_of_scope"
    if "perf" in case["tc_id"].lower() or "performance" in combined:
        return "performance"
    if "browser" in sheet_name.lower() or "responsive" in combined or "mobile" in combined:
        return "cross_browser"
    if any(
        token in combined
        for token in (
            "database",
            " db ",
            "email",
            "concurrent",
            "audit log",
            "lighthouse",
            "tims",
            "otp",
            "keycloak",
        )
    ):
        return "integration"
    return "ui"


def automation_component(case, sheet_name):
    if sheet_name.startswith("M5"):
        return "manual_item_creation"
    if sheet_name.startswith("M1"):
        feature = case["feature"].lower()
        if "bulk upload" in feature:
            return "upload_item_file"
        if "manual" in feature or "typology" in feature or "metadata" in feature:
            return "manual_item_creation"
        if any(role in feature for role in ("rwg", "sr.rwg", "pit")):
            return "review_queue"
        return "item_bank_management"
    if sheet_name.startswith("M2"):
        return "web_portal_admin"
    if sheet_name.startswith("M3"):
        return "item_testing_out_of_scope"
    if sheet_name.startswith("M4"):
        return "question_paper_creation"
    if sheet_name == "Browser & Responsiveness":
        return "cross_browser_responsiveness"
    return "unassigned"


def extract_cases(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    cases = []
    for worksheet in workbook.worksheets:
        header_row = None
        header_map = None
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            values = [clean(value) for value in row]
            if "TC ID" in values:
                header_row = row_number
                header_map = find_header_map(row)
                break
        if not header_map:
            continue

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            header_row + 1,
        ):
            tc_id = clean(row[header_map["tc_id"]])
            if not re.fullmatch(r"TC-[A-Za-z0-9-]+", tc_id):
                continue
            case = {
                key: (
                    clean(row[header_map[key]])
                    if key in header_map and header_map[key] < len(row)
                    else ""
                )
                for key in HEADERS
            }
            case.update(
                {
                    "module": worksheet.title,
                    "source_row": row_number,
                    "execution_mode": "",
                    "automation_component": "",
                    "automation": AUTOMATION_LINKS.get(tc_id),
                }
            )
            case["execution_mode"] = classify(case, worksheet.title)
            case["automation_component"] = automation_component(
                case,
                worksheet.title,
            )
            cases.append(case)
    return cases


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("rtm/rtm_test_cases.json"),
    )
    args = parser.parse_args()
    cases = extract_cases(args.workbook)
    payload = {
        "source_workbook": args.workbook.name,
        "case_count": len(cases),
        "cases": cases,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"Generated {len(cases)} RTM cases at {args.output}")


if __name__ == "__main__":
    main()
